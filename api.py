#!/usr/bin/env python3
# api.py
# FastAPI backend для MinskDvizh

import os
import io
import csv
import sqlite3
import httpx
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from typing import Optional
from collections import defaultdict

from fastapi import FastAPI, Query, HTTPException, UploadFile, File, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# ── Конфиг ──────────────────────────────────────────────────────────────────

DB_PATH      = os.getenv("DB_PATH", "/data/events_final.db")
MINSK_TZ     = timezone(timedelta(hours=3))
FRONTEND_URL = os.getenv("FRONTEND_URL", "*")
BOT_TOKEN    = os.getenv("TELEGRAM_BOT_TOKEN", "")
ADMIN_ID     = int(os.getenv("ADMIN_ID", "502917728"))

app = FastAPI(
    title="MinskDvizh API",
    description="API афиши Минска",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)

# ── БД ──────────────────────────────────────────────────────────────────────

@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()


def row_to_dict(row) -> dict:
    return dict(row)


# ── Pydantic схемы ───────────────────────────────────────────────────────────

class Event(BaseModel):
    id: int
    title: str
    details: Optional[str] = ""
    description: Optional[str] = ""
    event_date: str
    show_time: Optional[str] = ""
    end_time: Optional [str] = ""
    place: Optional[str] = ""
    location: Optional[str] = "Минск"
    price: Optional[str] = ""
    category: str
    source_url: Optional[str] = ""
    source_name: Optional[str] = ""


class EventsResponse(BaseModel):
    total: int
    page: int
    per_page: int
    events: list[Event]


class CategoryCounts(BaseModel):
    cinema: int = 0
    concert: int = 0
    theater: int = 0
    exhibition: int = 0
    kids: int = 0
    sport: int = 0
    party: int = 0
    free: int = 0
    excursion: int = 0
    market: int = 0
    masterclass: int = 0
    boardgames: int = 0
    broadcast: int = 0
    education: int = 0
    other: int = 0


# ── Вспомогательные функции ──────────────────────────────────────────────────

def now_minsk() -> datetime:
    return datetime.now(MINSK_TZ)


def today_str() -> str:
    return now_minsk().strftime("%Y-%m-%d")


def now_time_str() -> str:
    return now_minsk().strftime("%H:%M")


def get_weekend_dates() -> tuple[str, str]:
    today = now_minsk()
    days_until_saturday = (5 - today.weekday()) % 7 or 7
    saturday = today + timedelta(days=days_until_saturday)
    sunday = saturday + timedelta(days=1)
    return saturday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d")


def fetch_events(
    where_clauses: list[str],
    params: list,
    order: str = "event_date, show_time, title",
    limit: int = 500,
) -> list[dict]:
    where = " AND ".join(where_clauses) if where_clauses else "1=1"
    sql = f"""
        SELECT id, title, details, description, event_date, show_time, end_time,
               place, location, price, category, source_url, source_name
        FROM events
        WHERE {where}
        ORDER BY {order}
        LIMIT {limit}
    """
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(sql, params)
        return [row_to_dict(r) for r in cursor.fetchall()]


def paginate(items: list, page: int, per_page: int) -> tuple[list, int]:
    total = len(items)
    start = (page - 1) * per_page
    end = start + per_page
    return items[start:end], total

def _build_time_filter(date_filter: str, today: str, now_time: str) -> tuple[str, list]:
    """Возвращает SQL условие для фильтрации прошедших событий и параметры.
    Логика: нет show_time → всегда показываем.
            есть end_time → фильтруем по end_time > now.
            нет end_time  → фильтруем по show_time > now.
    """
    if date_filter != today:
        return "", []

    return """
        (
            show_time = '' OR show_time IS NULL
            OR (
                (end_time != '' AND end_time IS NOT NULL AND end_time > ?)
                OR
                ((end_time = '' OR end_time IS NULL) AND show_time > ?)
            )
        )
    """, [now_time, now_time]


# ── Health ───────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "time": now_minsk().isoformat()}


# ── Категории ────────────────────────────────────────────────────────────────

@app.get("/api/categories/counts", response_model=CategoryCounts)
def categories_counts():
    """Количество актуальных событий по каждой категории."""
    today = today_str()
    with get_db() as conn:
        cursor = conn.cursor()
        
        # Обычные категории - ВСЕ события (и платные, и бесплатные)
        cursor.execute(
            "SELECT category, COUNT(*) as cnt FROM events WHERE event_date >= ? GROUP BY category",
            (today,)
        )
        data = {row["category"]: row["cnt"] for row in cursor.fetchall()}
        
        # Категория free - все бесплатные события (независимо от категории)
        cursor.execute(
            "SELECT COUNT(*) FROM events WHERE event_date >= ? AND price = 'Бесплатно'",
            (today,)
        )
        free_count = cursor.fetchone()[0]
        data["free"] = free_count
        
    return CategoryCounts(**data)

# ── Даты с событиями (для календаря) ────────────────────────────────────────

@app.get("/api/calendar/dates")
def calendar_dates(
    category: Optional[str] = Query(None),
    months_ahead: int = Query(3, ge=1, le=12),
):
    """Список дат у которых есть события (для подсветки в календаре)."""
    today = today_str()
    until = (now_minsk() + timedelta(days=30 * months_ahead)).strftime("%Y-%m-%d")
    params: list = [today, until]
    where = ["event_date >= ?", "event_date <= ?"]
    
    # КАТЕГОРИЯ FREE - ОСОБАЯ ОБРАБОТКА
    if category == "free":
        where.append("price = 'Бесплатно'")
    elif category and category != "all":
        where.append("category = ?")
        params.append(category)
        
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            f"SELECT DISTINCT event_date FROM events WHERE {' AND '.join(where)} ORDER BY event_date",
            params,
        )
        dates = [row["event_date"] for row in cursor.fetchall()]
    return {"dates": dates}


# ── Последнее обновление ─────────────────────────────────────────────────────

@app.get("/api/last-updated")
def last_updated():
    """Когда последний раз обновлялась база (по минимальной дате добавления)."""
    # Простая эвристика — самая свежая запись в БД
    with get_db() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT MAX(created_at) as ts FROM events")
            row = cursor.fetchone()
            ts = row["ts"] if row and row["ts"] else None
        except Exception:
            ts = None
    return {"last_updated": ts, "schedule": "ежедневно в 06:00"}


# ── Основной эндпоинт событий ────────────────────────────────────────────────

@app.get("/api/events", response_model=EventsResponse)
def get_events(
    category: Optional[str] = Query(None, description="Категория: cinema/concert/theater/..."),
    date: Optional[str]     = Query(None, description="Конкретная дата YYYY-MM-DD"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str]   = Query(None),
    search: Optional[str]    = Query(None, description="Поиск по названию/месту"),
    page: int                = Query(1, ge=1),
    per_page: int            = Query(10, ge=1, le=500),
):
    today = today_str()
    now_t = now_time_str()

    where: list[str] = []
    params: list = []

    # Дата
    if date:
        where.append("event_date = ?")
        params.append(date)
        # Для сегодня — фильтруем прошедшие сеансы с учётом end_time
        if date == today:
            time_filter, time_params = _build_time_filter(date, today, now_t)
            where.append(time_filter)
            params.extend(time_params)
    elif date_from and date_to:
        where.append("event_date BETWEEN ? AND ?")
        params.extend([date_from, date_to])
        # Если начало диапазона — сегодня, фильтруем время
        if date_from == today:
            time_filter, time_params = _build_time_filter(date_from, today, now_t)
            where.append(time_filter)
            params.extend(time_params)
    elif date_from:
        where.append("event_date >= ?")
        params.append(date_from)
        if date_from == today:
            time_filter, time_params = _build_time_filter(date_from, today, now_t)
            where.append(time_filter)
            params.extend(time_params)
    else:
        # По умолчанию — только будущие
        where.append("event_date >= ?")
        params.append(today)
        # Фильтруем сегодняшние события по времени
        where.append("(event_date > ? OR (show_time = '' OR show_time IS NULL OR (end_time != '' AND end_time IS NOT NULL AND end_time > ?) OR ((end_time = '' OR end_time IS NULL) AND show_time > ?)))")
        params.extend([today, now_t, now_t])

    # КАТЕГОРИЯ FREE - ОСОБАЯ ОБРАБОТКА
    if category == "free":
        # Показываем ВСЕ события с ценой "Бесплатно"
        where.append("price = 'Бесплатно'")
        # НЕ добавляем условие по category
    elif category and category != "all":
        where.append("category = ?")
        params.append(category)

    # Поиск
    if search and len(search.strip()) >= 2:
        q = search.strip()
        current_year = now_minsk().year

        import re as _re
        date_match = _re.match(r'^(\d{1,2})\.(\d{1,2})(?:\.(\d{4}))?$', q)
        if date_match:
            # При поиске по дате — заменяем date-фильтры
            day, month = date_match.group(1).zfill(2), date_match.group(2).zfill(2)
            year = date_match.group(3) or str(current_year)
            search_date = f"{year}-{month}-{day}"
            where.clear()
            params.clear()
            where.append("event_date = ?")
            params.append(search_date)
        else:
            # SQLite LOWER() не работает с кириллицей — генерируем варианты через Python
            ql = q.lower()
            qc = (ql[0].upper() + ql[1:]) if ql else ql  # "концерт" → "Концерт"
            spl = f"%{ql}%"   # нижний регистр (Python корректно обрабатывает кириллицу)
            spc = f"%{qc}%"   # с заглавной первой буквой
            where.append(
                "(title LIKE ? OR title LIKE ? "
                "OR place LIKE ? OR place LIKE ? "
                "OR details LIKE ? OR details LIKE ? "
                "OR description LIKE ? OR description LIKE ? "
                "OR category LIKE ?)"
            )
            params.extend([spl, spc, spl, spc, spl, spc, spl, spc, spl])

    events = fetch_events(where, params)
    page_events, total = paginate(events, page, per_page)

    return EventsResponse(
        total=total,
        page=page,
        per_page=per_page,
        events=[Event(**e) for e in page_events],
    )


# ── Шорткаты ─────────────────────────────────────────────────────────────────

@app.get("/api/events/today", response_model=EventsResponse)
def events_today(
    category: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=500),
):
    today = today_str()
    now_t = now_time_str()
    where = ["event_date = ?"]
    params: list = [today]

    # Фильтр времени с учётом end_time
    time_filter, time_params = _build_time_filter(today, today, now_t)
    where.append(time_filter)
    params.extend(time_params)
    
    # КАТЕГОРИЯ FREE - ОСОБАЯ ОБРАБОТКА
    if category == "free":
        where.append("price = 'Бесплатно'")
    elif category and category != "all":
        where.append("category = ?")
        params.append(category)
        
    events = fetch_events(where, params)
    page_events, total = paginate(events, page, per_page)
    return EventsResponse(total=total, page=page, per_page=per_page,
                          events=[Event(**e) for e in page_events])


@app.get("/api/events/tomorrow", response_model=EventsResponse)
def events_tomorrow(
    category: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=500),
):
    tomorrow = (now_minsk() + timedelta(days=1)).strftime("%Y-%m-%d")
    where = ["event_date = ?"]
    params: list = [tomorrow]
    
    # КАТЕГОРИЯ FREE - ОСОБАЯ ОБРАБОТКА
    if category == "free":
        where.append("price = 'Бесплатно'")
    elif category and category != "all":
        where.append("category = ?")
        params.append(category)
        
    events = fetch_events(where, params)
    page_events, total = paginate(events, page, per_page)
    return EventsResponse(total=total, page=page, per_page=per_page,
                          events=[Event(**e) for e in page_events])


@app.get("/api/events/weekend", response_model=EventsResponse)
def events_weekend(
    category: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=500),
):
    saturday, sunday = get_weekend_dates()
    where = ["event_date IN (?, ?)"]
    params: list = [saturday, sunday]
    
    # КАТЕГОРИЯ FREE - ОСОБАЯ ОБРАБОТКА
    if category == "free":
        where.append("price = 'Бесплатно'")
    elif category and category != "all":
        where.append("category = ?")
        params.append(category)
        
    events = fetch_events(where, params)
    page_events, total = paginate(events, page, per_page)
    return EventsResponse(total=total, page=page, per_page=per_page,
                          events=[Event(**e) for e in page_events])


@app.get("/api/events/upcoming", response_model=EventsResponse)
def events_upcoming(
    category: Optional[str] = Query(None),
    days: int = Query(30, ge=1, le=90),
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=500),
):
    today = today_str()
    now_t = now_time_str()
    until = (now_minsk() + timedelta(days=days)).strftime("%Y-%m-%d")
    where = [
        "event_date BETWEEN ? AND ?",
        "(event_date > ? OR (show_time = '' OR show_time IS NULL OR (end_time != '' AND end_time IS NOT NULL AND end_time > ?) OR ((end_time = '' OR end_time IS NULL) AND show_time > ?)))",
    ]
    params: list = [today, until, today, now_t, now_t]
    
    # КАТЕГОРИЯ FREE - ОСОБАЯ ОБРАБОТКА
    if category == "free":
        where.append("price = 'Бесплатно'")
    elif category and category != "all":
        where.append("category = ?")
        params.append(category)
        
    events = fetch_events(where, params)
    page_events, total = paginate(events, page, per_page)
    return EventsResponse(total=total, page=page, per_page=per_page,
                          events=[Event(**e) for e in page_events])


# ── Сабмит события от пользователя сайта ────────────────────────────────────

class EventSubmit(BaseModel):
    title: str
    category: str
    event_date: str
    event_date_to: Optional[str] = None
    show_time: Optional[str] = ""
    end_time: Optional[str] = ""
    place: str
    address: Optional[str] = ""
    price: Optional[str] = ""
    details: Optional[str] = ""
    description: Optional[str] = ""
    source_url: Optional[str] = ""
    tg_user_id: Optional[int] = None      # ID пользователя из Telegram WebApp
    tg_username: Optional[str] = None
    tg_first_name: Optional[str] = None
    is_promo: Optional[bool] = False       # Промо-публикация в канал после одобрения


def _dates_in_range(date_from: str, date_to: str) -> list[str]:
    """Возвращает список дат включительно от date_from до date_to."""
    from datetime import date as _date
    start = _date.fromisoformat(date_from)
    end = _date.fromisoformat(date_to)
    result = []
    current = start
    while current <= end:
        result.append(current.isoformat())
        current += timedelta(days=1)
    return result


@app.post("/api/events/submit")
def submit_event(event: EventSubmit):
    """Принимает событие → валидирует даты → проверяет дубликат → сохраняет в pending_events."""
    import re as _re
    from datetime import date as _date

    # ── Валидация дат ────────────────────────────────────────────────────────
    today_d = _date.today()

    try:
        d_from = _date.fromisoformat(event.event_date)
    except Exception:
        raise HTTPException(status_code=400, detail="Некорректная дата начала. Формат: YYYY-MM-DD")

    if d_from < today_d:
        raise HTTPException(
            status_code=400,
            detail=f"Дата {d_from.strftime('%d.%m.%Y')} уже в прошлом. Сегодня {today_d.strftime('%d.%m.%Y')}"
        )

    if event.event_date_to:
        try:
            d_to = _date.fromisoformat(event.event_date_to)
        except Exception:
            raise HTTPException(status_code=400, detail="Некорректная дата окончания. Формат: YYYY-MM-DD")
        if d_to < today_d:
            raise HTTPException(status_code=400, detail=f"Дата окончания {d_to.strftime('%d.%m.%Y')} уже в прошлом.")
        if d_to < d_from:
            raise HTTPException(status_code=400, detail="Дата окончания раньше даты начала.")
        if d_to == d_from:
            raise HTTPException(status_code=400, detail="Для одного дня не указывайте дату окончания.")
        if (d_to - d_from).days > 90:
            raise HTTPException(status_code=400, detail="Период не может быть больше 90 дней.")
        # НЕ создаем список дат, только проверяем период
        is_period = True
        event_date_value = f"{event.event_date}|{event.event_date_to}"
    else:
        is_period = False
        event_date_value = event.event_date

    # ── Проверка дубликата ───────────────────────────────────────────────────
    def _norm(s: str) -> str:
        return _re.sub(r"[\s\-—–,\.!?]+", " ", (s or "").lower()).strip()

    t_norm = _norm(event.title)
    p_norm = _norm(event.place)

    # Для проверки дубликата используем первую дату (если период)
    check_date = event.event_date

    with get_db() as conn:
        # Уровень 1: title + date + place
        rows = conn.execute(
            "SELECT id, title, event_date, place FROM events "
            "WHERE event_date = ? AND LOWER(title) LIKE ? AND LOWER(COALESCE(place,'')) LIKE ?",
            (check_date, f"%{t_norm[:20]}%", f"%{p_norm[:20]}%")
        ).fetchall()
        for r in rows:
            if _norm(r["title"]) == t_norm and _norm(r["place"] or "") == p_norm:
                try:
                    from datetime import datetime as _dt
                    d_fmt = _dt.strptime(r["event_date"], "%Y-%m-%d").strftime("%d.%m.%Y")
                except Exception:
                    d_fmt = r["event_date"]
                raise HTTPException(status_code=409, detail=f"Событие уже есть в афише: «{r['title']}» {d_fmt}")

        # Уровень 2: title + date (без места)
        rows = conn.execute(
            "SELECT id, title, event_date, place FROM events "
            "WHERE event_date = ? AND LOWER(title) LIKE ?",
            (check_date, f"%{t_norm[:20]}%")
        ).fetchall()
        for r in rows:
            if _norm(r["title"]) == t_norm:
                try:
                    from datetime import datetime as _dt
                    d_fmt = _dt.strptime(r["event_date"], "%Y-%m-%d").strftime("%d.%m.%Y")
                except Exception:
                    d_fmt = r["event_date"]
                raise HTTPException(status_code=409, detail=f"Событие уже есть в афише: «{r['title']}» {d_fmt}")

        # Проверяем pending_events
        rows_p = conn.execute(
            "SELECT id, title, event_date FROM pending_events "
            "WHERE event_date LIKE ? AND status NOT IN ('rejected','approved') AND LOWER(title) LIKE ?",
            (f"%{check_date}%", f"%{t_norm[:20]}%")
        ).fetchall()
        for r in rows_p:
            if _norm(r["title"]) == t_norm:
                raise HTTPException(status_code=409, detail=f"Событие уже отправлено на модерацию: «{r['title']}»")

    try:
        user_id   = event.tg_user_id or 0
        username  = event.tg_username or "web_user"
        first_name = event.tg_first_name or "Web"
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Сохраняем ОДНУ запись, даже если это период
            cursor.execute("""
                INSERT INTO pending_events
                    (user_id, username, first_name, title, event_date, show_time, end_time,
                     place, address, category, details, description, price, source_url,
                     is_promo, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                user_id, username, first_name,
                event.title,
                event_date_value,
                event.show_time or "",
                event.end_time or "",
                event.place,
                event.address or "",
                event.category,
                event.details or "",
                event.description or "",
                event.price or "",
                event.source_url or "",
                1 if event.is_promo else 0,
                "pending",
                datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"),
            ))
            pending_id = cursor.lastrowid
            
            # Логируем в user_stats
            conn.execute(
                "INSERT INTO user_stats (user_id, username, first_name, action, detail, created_at) VALUES (?,?,?,?,?,?)",
                (user_id, username, first_name, "web_submit_event", 
                 f"{event.title} ({event_date_value})",
                 datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"))
            )
            conn.commit()
        
        # Формируем информацию о датах для уведомления
        days_info = ""
        date_info = event.event_date
        if is_period:
            days_count = (d_to - d_from).days + 1
            days_info = f" ({days_count} дней)"
            date_info = f"{event.event_date} → {event.event_date_to}"

        # Формируем строку времени для уведомления
        time_str = ""
        if event.show_time:
            time_str = f" ⏰ {event.show_time}"
            if event.end_time:
                time_str += f"-{event.end_time}"

        # Уведомление админу в Telegram
        if BOT_TOKEN and ADMIN_ID:
            lines = [
                "🌐 <b>Новое событие с сайта</b>",
                f"📌 <b>{event.title}</b>",
                f"🗂 Категория: {event.category}",
            ]
            if event.details:
                lines.append(f"📝 Формат: {event.details}")
            lines += [
                f"📅 Дата: {date_info}{days_info}{time_str}",
                f"🏢 Место: {event.place}" + (f", {event.address}" if event.address else ""),
            ]
            if event.price:
                lines.append(f"💰 Цена: {event.price}")
            if event.description:
                lines.append(f"📝 {event.description[:500]}")
            if event.source_url:
                lines.append(f"🔗 {event.source_url}")
            lines.append(f"\n<i>ID в очереди: #{pending_id}</i>")
            if event.is_promo:
                lines.append("📣 <b>Запрошена промо-публикация в канале</b>")
            text = "\n".join(lines)
            
            try:
                with httpx.Client(timeout=5) as client:
                    client.post(
                        f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                        json={
                            "chat_id": ADMIN_ID,
                            "text": text,
                            "parse_mode": "HTML",
                            "reply_markup": {
                                "inline_keyboard": [
                                    [
                                        {"text": "✅ Одобрить", "callback_data": f"mod_approve_{pending_id}"},
                                        {"text": "❌ Отклонить", "callback_data": f"mod_reject_{pending_id}"},
                                    ],
                                    [
                                        {"text": "✏️ Редактировать", "callback_data": f"mod_edit_{pending_id}"},
                                    ],
                                ]
                            }
                        }
                    )
            except Exception:
                pass  # не блокируем ответ если TG недоступен

        return {"ok": True, "message": "Событие отправлено на модерацию"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Webapp ping — трекинг открытий Mini App ─────────────────────────────────

class WebappPingRequest(BaseModel):
    user_id: Optional[int] = None
    username: Optional[str] = None
    first_name: Optional[str] = None

@app.post("/api/webapp-ping")
def webapp_ping(req: WebappPingRequest):
    """Сайт вызывает при открытии — логируем для статистики."""
    try:
        with get_db() as conn:
            conn.execute(
                "INSERT INTO user_stats (user_id, username, first_name, action, detail, created_at) VALUES (?,?,?,?,?,?)",
                (req.user_id or 0, req.username or "", req.first_name or "",
                 "webapp_ping", "web",
                 datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"))
            )
            conn.commit()
    except Exception:
        pass
    return {"ok": True}


# ── Одно событие ─────────────────────────────────────────────────────────────

@app.get("/api/events/{event_id}", response_model=Event)
def get_event(event_id: int):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """SELECT id, title, details, description, event_date, show_time, end_time,
                      place, location, price, category, source_url, source_name
               FROM events WHERE id = ?""",
            (event_id,)
        )
        row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Event not found")
    return Event(**row_to_dict(row))

# ── Pydantic схемы для подписок ───────────────────────────────────────────────

class SubscriptionRequest(BaseModel):
    user_id: int
    category: str
    date_type: str  # "upcoming", "daily", "weekly"


class SubscriptionResponse(BaseModel):
    subscriptions: list[dict]  # [{"category": "concert", "date_type": "upcoming"}, ...]


# ── Эндпоинты для подписок ────────────────────────────────────────────────────

@app.get("/api/ical")
def get_ical(
    title: str = Query(...),
    date: str = Query(...),          # YYYYMMDD
    time: Optional[str] = Query(None),    # HH:MM
    end_time: Optional[str] = Query(None),
    venue: Optional[str] = Query(None),
    url: Optional[str] = Query(None),
    description: Optional[str] = Query(None),
):
    """Генерирует .ics файл события для импорта в календарь."""
    def _dt(d: str, t: Optional[str]) -> str:
        if t:
            h, m = t.split(":")
            return f"{d}T{h.zfill(2)}{m.zfill(2)}00"
        return d

    dtstart = _dt(date, time)
    dtend = _dt(date, end_time or time)
    now = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    uid = f"minskdvizh-{date}-{title[:30].replace(' ', '-')}@minskdvizh"

    def esc(s: str) -> str:
        return s.replace("\\", "\\\\").replace(",", "\\,").replace("\n", "\\n")

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//MinskDvizh//MinskDvizh//RU",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "BEGIN:VEVENT",
        f"UID:{uid}",
        f"DTSTAMP:{now}",
        f"DTSTART:{dtstart}",
        f"DTEND:{dtend}",
        f"SUMMARY:{esc(title)}",
    ]
    if description:
        lines.append(f"DESCRIPTION:{esc(description)}")
    if venue:
        lines.append(f"LOCATION:{esc(venue)}")
    if url:
        lines.append(f"URL:{url}")
    lines += ["END:VEVENT", "END:VCALENDAR"]

    ics_content = "\r\n".join(lines)
    safe_title = "".join(c if ord(c) < 128 and (c.isalnum() or c in " _-") else "" for c in title)[:40].strip().replace(" ", "_")
    filename = f"{safe_title or 'event'}.ics"

    return Response(
        content=ics_content,
        media_type="text/calendar; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/subscriptions", response_model=SubscriptionResponse)
def get_subscriptions(user_id: int = Query(..., description="ID пользователя Telegram")):
    """
    Получить все активные подписки пользователя.
    Вызывается из фронта при загрузке страницы подписок.
    """
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT category, date_type FROM subscriptions WHERE user_id = ? AND status = 'active'",
            (user_id,)
        )
        subs = [{"category": row[0], "date_type": row[1]} for row in cursor.fetchall()]
    
    return {"subscriptions": subs}


@app.post("/api/subscriptions/add")
def add_subscription(req: SubscriptionRequest):
    """
    Добавить или активировать подписку.
    INSERT OR REPLACE гарантирует, что даже если была неактивная — станет активной.
    """
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT OR REPLACE INTO subscriptions 
            (user_id, category, date_type, status) 
            VALUES (?, ?, ?, 'active')
        """, (req.user_id, req.category, req.date_type))
        conn.commit()
    
    # Логируем для статистики
    try:
        with get_db() as conn:
            conn.execute(
                "INSERT INTO user_stats (user_id, action, detail, created_at) VALUES (?, ?, ?, ?)",
                (req.user_id, "web_subscribe", f"{req.category}_{req.date_type}",
                 datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"))
            )
            conn.commit()
    except Exception:
        pass
    
    return {"ok": True, "message": "Подписка добавлена"}


@app.post("/api/subscriptions/remove")
def remove_subscription(req: SubscriptionRequest):
    """
    Деактивировать подписку (мягкое удаление).
    """
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE subscriptions SET status='inactive' WHERE user_id = ? AND category = ? AND date_type = ?",
            (req.user_id, req.category, req.date_type)
        )
        conn.commit()
    
    # Логируем для статистики
    try:
        with get_db() as conn:
            conn.execute(
                "INSERT INTO user_stats (user_id, action, detail, created_at) VALUES (?, ?, ?, ?)",
                (req.user_id, "web_unsubscribe", f"{req.category}_{req.date_type}",
                 datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"))
            )
            conn.commit()
    except Exception:
        pass
    
    return {"ok": True, "message": "Подписка удалена"}


@app.get("/api/subscriptions/categories")
def get_available_categories():
    """
    Список категорий, на которые можно подписаться (те, у которых есть события).
    """
    today = today_str()
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT DISTINCT category FROM events WHERE event_date >= ?",
            (today,)
        )
        categories = [row[0] for row in cursor.fetchall()]
    
    # Добавляем все возможные date_type
    date_types = ["upcoming", "daily", "weekly"]
    
    return {
        "categories": categories,
        "date_types": date_types,
        "descriptions": {
            "upcoming": "🔔 Все новые события",
            "daily": "📅 Ежедневный дайджест",
            "weekly": "📆 Дайджест на выходные"
        }
    }


# ── Флеш-подписки ─────────────────────────────────────────────────────────────

class FlashSubscriptionRequest(BaseModel):
    user_id: int
    query: str


class FlashSubscriptionRemoveRequest(BaseModel):
    user_id: int
    flash_id: int


@app.get("/api/flash-subscriptions")
def get_flash_subscriptions(user_id: int = Query(..., description="ID пользователя Telegram")):
    """Получить все активные флеш-подписки пользователя."""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, query, created_at FROM flash_subscriptions WHERE user_id = ? AND status = 'active' ORDER BY created_at DESC",
            (user_id,)
        )
        subs = [{"id": row[0], "query": row[1], "created_at": row[2]} for row in cursor.fetchall()]
    return {"flash_subscriptions": subs}


@app.post("/api/flash-subscriptions/add")
def add_flash_subscription(req: FlashSubscriptionRequest):
    """Добавить флеш-подписку. Возвращает 409 если такая уже есть."""
    if not req.query or len(req.query.strip()) < 2:
        raise HTTPException(status_code=400, detail="Запрос слишком короткий (минимум 2 символа)")

    query_clean = req.query.strip()

    with get_db() as conn:
        # Проверяем дубликат
        existing = conn.execute(
            "SELECT id FROM flash_subscriptions WHERE user_id = ? AND LOWER(query) = LOWER(?) AND status = 'active'",
            (req.user_id, query_clean)
        ).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail="Вы уже подписаны на этот запрос")

        conn.execute(
            "INSERT INTO flash_subscriptions (user_id, query, created_at, status) VALUES (?, ?, ?, 'active')",
            (req.user_id, query_clean, datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()

    # Логируем
    try:
        with get_db() as conn:
            conn.execute(
                "INSERT INTO user_stats (user_id, action, detail, created_at) VALUES (?, ?, ?, ?)",
                (req.user_id, "web_flash_subscribe", query_clean,
                 datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"))
            )
            conn.commit()
    except Exception:
        pass

    return {"ok": True, "message": "Флеш-подписка оформлена"}


@app.post("/api/flash-subscriptions/remove")
def remove_flash_subscription(req: FlashSubscriptionRemoveRequest):
    """Деактивировать флеш-подписку по id."""
    with get_db() as conn:
        conn.execute(
            "UPDATE flash_subscriptions SET status = 'inactive' WHERE id = ? AND user_id = ?",
            (req.flash_id, req.user_id)
        )
        conn.commit()

    # Логируем
    try:
        with get_db() as conn:
            conn.execute(
                "INSERT INTO user_stats (user_id, action, detail, created_at) VALUES (?, ?, ?, ?)",
                (req.user_id, "web_flash_unsubscribe", str(req.flash_id),
                 datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"))
            )
            conn.commit()
    except Exception:
        pass

    return {"ok": True, "message": "Флеш-подписка отменена"}


# ==================== ПАКЕТНАЯ ЗАГРУЗКА СОБЫТИЙ ====================

BATCH_CATEGORY_MAP = {
    "кино": "cinema", "cinema": "cinema",
    "концерт": "concert", "концерты": "concert", "concert": "concert",
    "театр": "theater", "theater": "theater",
    "выставка": "exhibition", "выставки": "exhibition", "exhibition": "exhibition",
    "детям": "kids", "дети": "kids", "kids": "kids",
    "спорт": "sport", "sport": "sport",
    "движ": "party", "вечеринка": "party", "party": "party",
    "бесплатно": "free", "free": "free",
    "экскурсия": "excursion", "excursion": "excursion",
    "маркет": "market", "market": "market",
    "мастер-класс": "masterclass", "masterclass": "masterclass",
    "настолки": "boardgames", "boardgames": "boardgames",
    "трансляция": "broadcast", "broadcast": "broadcast",
    "обучение": "education", "education": "education",
    "квиз": "quiz", "quiz": "quiz",
}

BATCH_TEMPLATE_HEADERS = [
    "title", "details", "category", "event_date", "show_time",
    "place", "address", "price", "description", "source_url", "is_promo"
]

BATCH_TEMPLATE_EXAMPLE = [
    "Концерт джаза", "Вечер живой музыки для всех", "concert",
    "15.05.2026", "19:00", "Джаз-клуб Blue Note", "ул. Немига, 3",
    "от 20 руб", "Программа из классики и современного джаза", "https://example.com", "0"
]


def _batch_parse_date(raw: str):
    """Парсит дату или период. Возвращает (value, error).
    Поддерживает форматы:
      - YYYY-MM-DD (ISO, в т.ч. datetime из Excel: "2026-03-29 00:00:00")
      - ДД.ММ.ГГГГ
      - ДД.ММ.ГГГГ-ДД.ММ.ГГГГ (период)
    """
    import re as _re
    from datetime import date as _date
    raw = (raw or "").strip()
    if not raw:
        return None, "пустая дата"
    # Excel/openpyxl возвращает datetime как "2026-03-29 00:00:00" — берём только дату
    excel_dt = _re.match(r"^(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}(:\d{2})?$", raw)
    if excel_dt:
        raw = excel_dt.group(1)
    if _re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
        try:
            d = _date.fromisoformat(raw)
            if d < _date.today():
                return None, f"дата {raw} в прошлом"
            return raw, None
        except Exception:
            return None, f"невалидная дата {raw}"
    pm = _re.match(r"^(\d{1,2}\.\d{1,2}\.\d{4})-(\d{1,2}\.\d{1,2}\.\d{4})$", raw)
    if pm:
        def _pd(s):
            d, mo, y = s.split(".")
            return _date(int(y), int(mo), int(d))
        try:
            d1, d2 = _pd(pm.group(1)), _pd(pm.group(2))
        except Exception:
            return None, f"невалидный период {raw}"
        if d2 < d1:
            return None, "конец периода раньше начала"
        if d1 < _date.today():
            return None, "дата начала в прошлом"
        if (d2 - d1).days > 90:
            return None, "период > 90 дней"
        return f"{d1.isoformat()}|{d2.isoformat()}", None
    m = _re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", raw)
    if m:
        day, month, year = m.groups()
        try:
            d = _date(int(year), int(month), int(day))
        except Exception:
            return None, f"невалидная дата {raw}"
        if d < _date.today():
            return None, f"дата {raw} в прошлом"
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}", None
    return None, f"не распознан формат: {raw}"


def _batch_parse_time(raw: str) -> tuple:
    """Парсит ЧЧ:ММ или ЧЧ:ММ-ЧЧ:ММ. Возвращает (show_time, end_time).
    Обрабатывает Excel-формат "19:00:00" → "19:00".
    """
    import re as _re
    raw = (raw or "").strip()
    if not raw:
        return "", ""
    # Excel может отдать "19:00:00" — обрезаем секунды
    raw = _re.sub(r"^(\d{1,2}:\d{2}):\d{2}$", r"\1", raw)
    rng = _re.match(r"^(\d{1,2}:\d{2})-(\d{1,2}:\d{2})$", raw)
    if rng:
        return rng.group(1), rng.group(2)
    if _re.match(r"^\d{1,2}:\d{2}$", raw):
        return raw, ""
    return raw, ""


def _batch_check_duplicate(title: str, event_date: str, place: str) -> bool:
    """Проверяет дубликат в events и pending_events. Возвращает True если дубль."""
    import re as _re
    def _norm(s):
        return _re.sub(r"[\s\-—–,\.!?]+", " ", (s or "").lower()).strip()
    t_norm = _norm(title)
    p_norm = _norm(place)
    date_chk = event_date.split("|")[0].strip() if "|" in (event_date or "") else event_date
    if not date_chk or not t_norm:
        return False
    with get_db() as conn:
        rows = conn.execute(
            "SELECT title, place FROM events WHERE event_date = ? AND LOWER(title) LIKE ?",
            (date_chk, f"%{t_norm[:20]}%")
        ).fetchall()
        for r in rows:
            if _norm(r[0]) == t_norm:
                return True
        rows_p = conn.execute(
            "SELECT title FROM pending_events "
            "WHERE event_date LIKE ? AND status NOT IN ('rejected','approved') AND LOWER(title) LIKE ?",
            (f"%{date_chk}%", f"%{t_norm[:20]}%")
        ).fetchall()
        for r in rows_p:
            if _norm(r[0]) == t_norm:
                return True
    return False


def _read_batch_rows(data: bytes, filename: str) -> tuple:
    """Читает xlsx/xls/csv, возвращает (rows: list[dict], error: str)."""
    fname = filename.lower()
    try:
        if fname.endswith(".csv"):
            text = data.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            return [dict(r) for r in reader], ""
        elif fname.endswith(".xlsx") or fname.endswith(".xls"):
            import openpyxl
            buf = io.BytesIO(data)
            wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return [], "Файл пустой"
            headers = [str(h).strip().lower() if h else "" for h in all_rows[0]]
            result = []
            for row in all_rows[1:]:
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                result.append({
                    headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                    for i in range(min(len(headers), len(row)))
                })
            return result, ""
        else:
            return [], f"Неподдерживаемый формат: {filename}"
    except ImportError:
        return [], "openpyxl не установлен — используйте CSV"
    except Exception as e:
        return [], f"Ошибка чтения файла: {e}"


class BatchUploadRequest(BaseModel):
    """Метаданные пользователя при batch-загрузке через API."""
    tg_user_id: Optional[int] = None
    tg_username: Optional[str] = None
    tg_first_name: Optional[str] = None


@app.post("/api/events/batch")
async def batch_upload_events(
    file: UploadFile = File(...),
    tg_user_id: Optional[int] = Query(None),
    tg_username: Optional[str] = Query(None),
    tg_first_name: Optional[str] = Query(None),
):
    """
    Пакетная загрузка событий из xlsx/xls/csv.
    Файл содержит столбцы: title, details, category, event_date, show_time,
    place, address, price, description, source_url.
    Каждое валидное событие попадает в pending_events со статусом pending.
    Возвращает детальный отчёт.
    """
    fname = file.filename or "upload"

    if not (fname.lower().endswith(".xlsx") or fname.lower().endswith(".xls") or fname.lower().endswith(".csv")):
        raise HTTPException(status_code=400, detail="Поддерживаются только .xlsx, .xls, .csv")

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Файл слишком большой, максимум 5 МБ")

    rows, err = _read_batch_rows(contents, fname)
    if err:
        raise HTTPException(status_code=400, detail=err)
    if not rows:
        raise HTTPException(status_code=400, detail="Файл пустой или не содержит данных")
    if len(rows) > 100:
        raise HTTPException(status_code=400, detail="Максимум 100 событий за раз")

    user_id = tg_user_id or 0
    username = tg_username or "web_user"
    first_name = tg_first_name or "Web"
    now_str = datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S")

    seen_in_file: set = set()
    results = []

    for row_num, raw_row in enumerate(rows, start=2):
        row = {(k or "").strip().lower(): v for k, v in raw_row.items()}

        title = row.get("title", "").strip()
        if not title or len(title) < 3:
            results.append({"row": row_num, "title": title[:30], "status": "error", "reason": "пустое или короткое название"})
            continue

        date_raw = row.get("event_date", "") or row.get("date", "")
        event_date, date_err = _batch_parse_date(date_raw)
        if date_err:
            results.append({"row": row_num, "title": title[:30], "status": "error", "reason": date_err})
            continue

        time_raw = row.get("show_time", "") or row.get("time", "")
        show_time, end_time = _batch_parse_time(time_raw)

        place = row.get("place", "").strip()
        if not place:
            results.append({"row": row_num, "title": title[:30], "status": "error", "reason": "не указано место"})
            continue

        cat_raw = row.get("category", "").strip().lower()
        category = BATCH_CATEGORY_MAP.get(cat_raw, "other")

        details     = (row.get("details", "") or "")[:300]
        description = (row.get("description", "") or "")[:1000]
        address     = (row.get("address", "") or "").strip()
        price       = (row.get("price", "") or "").strip()
        source_url  = (row.get("source_url", "") or row.get("url", "") or "").strip()
        # is_promo: "1", "да", "yes", "true", "+" → 1, всё остальное → 0
        promo_raw = (row.get("is_promo", "") or "").strip().lower()
        is_promo = 1 if promo_raw in ("1", "да", "yes", "true", "+") else 0

        # Дубликат внутри файла
        key = (title.lower(), event_date, place.lower())
        if key in seen_in_file:
            results.append({"row": row_num, "title": title[:30], "status": "error", "reason": "дубликат внутри файла"})
            continue
        seen_in_file.add(key)

        # Дубликат в БД
        if _batch_check_duplicate(title, event_date, place):
            results.append({"row": row_num, "title": title[:30], "status": "error", "reason": "уже есть в афише/очереди"})
            continue

        # Сохраняем в pending_events
        try:
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO pending_events
                        (user_id, username, first_name, title, event_date, show_time, end_time,
                         place, address, category, details, description, price, source_url,
                         is_promo, status, created_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'pending',?)
                """, (
                    user_id, username, first_name,
                    title, event_date, show_time, end_time,
                    place, address, category,
                    details, description, price, source_url,
                    is_promo, now_str,
                ))
                conn.execute(
                    "INSERT INTO user_stats (user_id, username, first_name, action, detail, created_at) VALUES (?,?,?,?,?,?)",
                    (user_id, username, first_name, "web_batch_upload",
                     f"{title} ({event_date})", now_str)
                )
                conn.commit()
                pending_id = cursor.lastrowid
            results.append({"row": row_num, "title": title[:30], "status": "accepted", "id": pending_id})
        except Exception as e:
            results.append({"row": row_num, "title": title[:30], "status": "error", "reason": str(e)})

    accepted = [r for r in results if r["status"] == "accepted"]
    errors   = [r for r in results if r["status"] == "error"]

    # Уведомляем админа если есть принятые
    if accepted and BOT_TOKEN and ADMIN_ID:
        try:
            msg = (
                f"\U0001f4e6 <b>Пакетная загрузка с сайта</b>\n"
                f"\U0001f464 {first_name} (@{username}, ID {user_id})\n"
                f"\U0001f4e5 Принято: <b>{len(accepted)}</b> | "
                f"\u26a0\ufe0f Пропущено: <b>{len(errors)}</b>\n"
                f"\U0001f4c1 Файл: {fname}\n\n"
                f"\U0001f50d /pending — просмотр очереди"
            )
            with httpx.Client(timeout=5) as client:
                client.post(
                    f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                    json={"chat_id": ADMIN_ID, "text": msg, "parse_mode": "HTML"}
                )
        except Exception:
            pass

    return {
        "success": True,
        "total": len(rows),
        "accepted": len(accepted),
        "errors": len(errors),
        "results": results,
    }


@app.get("/api/events/batch/template")
async def get_batch_template(format: str = Query("xlsx", description="Формат: xlsx или csv")):
    """
    Возвращает шаблон для пакетной загрузки событий.
    ?format=xlsx (по умолчанию) или ?format=csv
    """
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(BATCH_TEMPLATE_HEADERS)
        writer.writerow(BATCH_TEMPLATE_EXAMPLE)
        return Response(
            content=output.getvalue().encode("utf-8-sig"),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=minskdvizh_template.csv"},
        )

    # xlsx
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "События"

        header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(BATCH_TEMPLATE_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        example_fill = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
        for col, val in enumerate(BATCH_TEMPLATE_EXAMPLE, 1):
            ws.cell(row=2, column=col, value=val).fill = example_fill

        widths = [35, 35, 15, 12, 8, 25, 25, 15, 40, 30, 10]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        ws2 = wb.create_sheet("Подсказки")
        hints = [
            ("Поле", "Обязательное", "Формат / пример"),
            ("title",       "ДА",  "Название события"),
            ("details",     "ДА",  "Краткое описание — для кого и что"),
            ("category",    "ДА",  "concert / theater / cinema / exhibition / kids / sport / party / free / excursion / market / masterclass / boardgames / broadcast / education / quiz"),
            ("event_date",  "ДА",  "ДД.ММ.ГГГГ  или  ДД.ММ.ГГГГ-ДД.ММ.ГГГГ (период)"),
            ("show_time",   "ДА",  "ЧЧ:ММ  или  ЧЧ:ММ-ЧЧ:ММ"),
            ("place",       "ДА",  "Название площадки"),
            ("address",     "нет", "ул. Пример, 1"),
            ("price",       "нет", "от 20 руб  /  Бесплатно"),
            ("description", "нет", "Подробное описание, программа"),
            ("source_url",  "нет", "https://..."),
            ("is_promo",    "нет", "1 = анонсировать в канал и подписчикам, 0 = нет (по умолчанию)"),
        ]
        for row_idx, row_data in enumerate(hints, 1):
            for col_idx, val in enumerate(row_data, 1):
                ws2.cell(row=row_idx, column=col_idx, value=val)
        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 70

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=minskdvizh_template.xlsx"},
        )
    except ImportError:
        # Fallback на CSV если openpyxl не установлен
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(BATCH_TEMPLATE_HEADERS)
        writer.writerow(BATCH_TEMPLATE_EXAMPLE)
        return Response(
            content=output.getvalue().encode("utf-8-sig"),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=minskdvizh_template.csv"},
        )
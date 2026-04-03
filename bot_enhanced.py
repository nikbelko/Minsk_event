#!/usr/bin/env python3
# bot_enhanced.py
# Бот-афиша Минск

import logging
import os
import re
import sys
import json
import sqlite3
import csv
import io
import tempfile
from contextlib import contextmanager
from collections import defaultdict
from datetime import datetime, timedelta

from dotenv import load_dotenv
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    LabeledPrice,
    ReplyKeyboardMarkup,
    WebAppInfo,
)
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
    MessageHandler,
    InlineQueryHandler,
    filters,
)

load_dotenv()

from config import (
    MINSK_TZ, DB_PATH, ADMIN_ID,
    VENUE_OPEN_TIME, VENUE_CLOSE_TIME,
    BATCH_TEMPLATE_HEADERS, BATCH_TEMPLATE_EXAMPLE, BATCH_CATEGORY_MAP,
)

import asyncio
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from telegram.constants import ParseMode
from telegram.ext import PreCheckoutQueryHandler

# ---------------------- Конфиг и логирование ----------------------

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

TOKEN       = os.getenv("TELEGRAM_BOT_TOKEN")
DB_NAME     = DB_PATH  # алиас для совместимости с кодом бота
WEB_APP_URL = os.getenv("WEB_APP_URL", "https://minskdvizh-web.up.railway.app")
CHANNEL_ID  = os.getenv("CHANNEL_ID", "")   # @MinskDvizh или -100xxxxxxxxxx

DONATION_ENABLED = True
DONATION_SUGGESTIONS = [10, 50, 100, 500]
DONATION_CURRENCY = "XTR"  # Telegram Stars

PER_PAGE = 10
SEARCH_MULTIPLIER = 3

CATEGORY_EMOJI = {
    "cinema": "🎬",
    "concert": "🎵",
    "theater": "🎭",
    "exhibition": "🖼️",
    "kids": "🧸",
    "sport": "⚽",
    "free": "🆓",
    "party": "🌟",
    "excursion": "🗺️",
    "market": "🛍️",
    "masterclass": "🎨",
    "boardgames": "🎲",
    "broadcast": "📺",
    "education": "📚",
    "quiz": "❓",
}

CATEGORY_NAMES = {
    "cinema": "🎬 Кино",
    "concert": "🎵 Концерты",
    "theater": "🎭 Театр",
    "exhibition": "🖼️ Выставки",
    "kids": "🧸 Детям",
    "sport": "⚽ Спорт",
    "free": "🆓 Бесплатно",
    "party": "🌟 Движ",
    "excursion": "🗺️ Экскурсии",
    "market": "🛍️ Маркеты",
    "masterclass": "🎨 Мастер-классы",
    "boardgames": "🎲 Настолки",
    "broadcast": "📺 Трансляции",
    "education": "📚 Обучение",
    "quiz": "❓ Квизы",
}

# ---------------------- Работа с БД ----------------------


@contextmanager
def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    conn.create_function("pylow", 1, lambda s: s.lower() if s else "")
    try:
        yield conn
    finally:
        conn.close()


def init_db():
    # Создаём директорию если не существует (первый запуск на Volume)
    db_dir = os.path.dirname(DB_NAME)
    if db_dir and not os.path.exists(db_dir):
        os.makedirs(db_dir, exist_ok=True)

    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS pending_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                username TEXT,
                first_name TEXT,
                title TEXT,
                event_date TEXT,
                show_time TEXT,
                place TEXT,
                category TEXT,
                description TEXT,
                price TEXT,
                address TEXT DEFAULT '',
                source_url TEXT DEFAULT '',
                status TEXT DEFAULT 'pending',
                created_at TEXT NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS subscriptions (
                user_id INTEGER,
                category TEXT,
                date_type TEXT,
                PRIMARY KEY (user_id, category, date_type)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS user_stats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                username TEXT,
                first_name TEXT,
                action TEXT NOT NULL,
                detail TEXT,
                created_at TEXT NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS flash_subscriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                query TEXT NOT NULL,
                created_at TEXT NOT NULL,
                status TEXT DEFAULT 'active'
            )
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_flash_user ON flash_subscriptions(user_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_flash_status ON flash_subscriptions(status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_user_stats_user_id ON user_stats(user_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_user_stats_created_at ON user_stats(created_at)")
        # Миграции для существующих БД
        for col_sql in [
            "ALTER TABLE pending_events ADD COLUMN details TEXT DEFAULT ''",
            "ALTER TABLE pending_events ADD COLUMN end_time TEXT DEFAULT ''",
            "ALTER TABLE pending_events ADD COLUMN is_promo INTEGER DEFAULT 0",
            "ALTER TABLE events ADD COLUMN end_time TEXT DEFAULT ''",
            "ALTER TABLE subscriptions ADD COLUMN status TEXT DEFAULT 'active'",
            "ALTER TABLE flash_subscriptions ADD COLUMN last_notified_at TEXT DEFAULT ''",
        ]:
            try:
                cursor.execute(col_sql)
            except Exception:
                pass
        conn.commit()
        try:
            cursor.execute("""
                DELETE FROM events
                WHERE event_date < DATE('now', '-7 days')
                AND (source_name IS NULL OR source_name != 'user_submitted')
            """)
            conn.commit()
        except Exception:
            pass
        for _migration in [
            "ALTER TABLE pending_events ADD COLUMN address TEXT DEFAULT ''",
            "ALTER TABLE pending_events ADD COLUMN source_url TEXT DEFAULT ''",
        ]:
            try:
                cursor.execute(_migration); conn.commit()
            except Exception:
                pass


def log_user_action(user_id: int, username: str | None, first_name: str | None, action: str, detail: str | None = None):
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO user_stats (user_id, username, first_name, action, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (user_id, username, first_name, action, detail, datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S")),
            )
            conn.commit()
    except Exception as e:
        logger.error(f"Ошибка логирования: {e}")

def _build_time_filter(date_filter: str, today: str, now_time: str) -> tuple[str, list]:
    """Возвращает чистое SQL условие БЕЗ 'AND' для фильтрации прошедших событий.
    - Для where.append(): добавлять напрямую, ' AND '.join() сам добавит AND.
    - Для query +=: использовать query += ' AND ' + time_filter.
    Логика: нет show_time → показываем только в VENUE_OPEN_TIME–VENUE_CLOSE_TIME
                            (выставки, музеи и т.п. ночью закрыты).
            есть end_time → фильтруем по end_time > now.
            нет end_time  → фильтруем по show_time > now.
    """
    if date_filter != today:
        return "", []

    venue_open = VENUE_OPEN_TIME <= now_time < VENUE_CLOSE_TIME

    if venue_open:
        return (
            "(show_time = '' OR show_time IS NULL "
            "OR ((end_time != '' AND end_time IS NOT NULL AND end_time > ?) "
            "OR ((end_time = '' OR end_time IS NULL) AND show_time > ?)))"
        ), [now_time, now_time]
    else:
        # Вне рабочих часов — только события с явным временем
        return (
            "((end_time != '' AND end_time IS NOT NULL AND end_time > ?) "
            "OR ((end_time = '' OR end_time IS NULL) AND show_time != '' AND show_time IS NOT NULL AND show_time > ?))"
        ), [now_time, now_time]
def get_stats_data(exclude_admin: bool = True) -> dict:
    today = datetime.now(MINSK_TZ).strftime("%Y-%m-%d")
    admin_filter = ADMIN_ID if exclude_admin else -1  # -1 никогда не совпадёт
    with get_db_connection() as conn:
        cursor = conn.cursor()

        # ── Всего пользователей (all-time) ──────────────────────────
        cursor.execute("SELECT COUNT(DISTINCT user_id) FROM user_stats WHERE user_id != ?", (admin_filter,))
        total_users = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM user_stats WHERE user_id != ?", (admin_filter,))
        total_actions = cursor.fetchone()[0]
        # Сколько дней проект живёт (с первого пользователя)
        cursor.execute("SELECT MIN(DATE(created_at)) FROM user_stats WHERE user_id != ?", (admin_filter,))
        first_date_row = cursor.fetchone()[0]
        if first_date_row:
            from datetime import date as _date
            days_alive = (datetime.now(MINSK_TZ).date() - datetime.strptime(first_date_row, "%Y-%m-%d").date()).days
        else:
            days_alive = 0

        # ── DAU / WAU / MAU ─────────────────────────────────────────
        cursor.execute("""SELECT COUNT(DISTINCT user_id) FROM user_stats
            WHERE user_id != ? AND created_at LIKE ?""", (admin_filter, f"{today}%"))
        dau = cursor.fetchone()[0]
        cursor.execute("""SELECT COUNT(*) FROM user_stats
            WHERE user_id != ? AND created_at LIKE ?""", (admin_filter, f"{today}%"))
        actions_today = cursor.fetchone()[0]
        cursor.execute("""SELECT COUNT(DISTINCT user_id) FROM user_stats
            WHERE user_id != ? AND created_at >= DATE('now', '-7 days')""", (admin_filter,))
        wau = cursor.fetchone()[0]
        cursor.execute("""SELECT COUNT(DISTINCT user_id) FROM user_stats
            WHERE user_id != ? AND created_at >= DATE('now', '-30 days')""", (admin_filter,))
        mau = cursor.fetchone()[0]

        # ── Новые пользователи ───────────────────────────────────────
        cursor.execute("""SELECT COUNT(*) FROM (
            SELECT user_id FROM user_stats WHERE user_id != ?
            GROUP BY user_id HAVING MIN(DATE(created_at)) = ?
        )""", (admin_filter, today))
        new_today = cursor.fetchone()[0]
        cursor.execute("""SELECT COUNT(*) FROM (
            SELECT user_id FROM user_stats WHERE user_id != ?
            GROUP BY user_id HAVING MIN(DATE(created_at)) >= DATE('now', '-30 days')
        )""", (admin_filter,))
        new_30d = cursor.fetchone()[0]

        # ── WebApp DAU / WAU / MAU ───────────────────────────────────
        wa_filter = "action IN ('open_webapp', 'webapp_ping') AND user_id != ?"
        cursor.execute(f"SELECT COUNT(DISTINCT user_id) FROM user_stats WHERE {wa_filter}", (admin_filter,))
        webapp_total = cursor.fetchone()[0]
        cursor.execute(f"SELECT COUNT(DISTINCT user_id) FROM user_stats WHERE {wa_filter} AND created_at LIKE ?",
                       (admin_filter, f"{today}%"))
        webapp_dau = cursor.fetchone()[0]
        cursor.execute(f"SELECT COUNT(DISTINCT user_id) FROM user_stats WHERE {wa_filter} AND created_at >= DATE('now', '-7 days')",
                       (admin_filter,))
        webapp_wau = cursor.fetchone()[0]
        cursor.execute(f"SELECT COUNT(DISTINCT user_id) FROM user_stats WHERE {wa_filter} AND created_at >= DATE('now', '-30 days')",
                       (admin_filter,))
        webapp_mau = cursor.fetchone()[0]

        # ── Активность по дням (30 дней) ────────────────────────────
        cursor.execute("""
            SELECT
                DATE(u.created_at) as day,
                COUNT(*) as cnt,
                COUNT(DISTINCT u.user_id) as users,
                COUNT(DISTINCT CASE WHEN DATE(u.created_at) = fv.first_day THEN u.user_id END) as new_users
            FROM user_stats u
            LEFT JOIN (
                SELECT user_id, MIN(DATE(created_at)) as first_day
                FROM user_stats WHERE user_id != ? GROUP BY user_id
            ) fv ON u.user_id = fv.user_id
            WHERE u.created_at >= DATE('now', '-30 days') AND u.user_id != ?
            GROUP BY day ORDER BY day DESC
        """, (admin_filter, admin_filter))
        daily_activity = cursor.fetchall()

        cursor.execute("""SELECT DATE(created_at) as day, COUNT(DISTINCT user_id) as users
            FROM user_stats WHERE action IN ('open_webapp','webapp_ping') AND user_id != ?
            AND created_at >= DATE('now', '-30 days') GROUP BY day""", (admin_filter,))
        webapp_by_day = {r["day"]: r["users"] for r in cursor.fetchall()}

        # ── Активность по месяцам ─────────────────────────────────────
        cursor.execute("""
            SELECT
                strftime('%Y-%m', u.created_at) as month,
                COUNT(*) as cnt,
                COUNT(DISTINCT u.user_id) as users,
                COUNT(DISTINCT CASE
                    WHEN strftime('%Y-%m', u.created_at) = strftime('%Y-%m', fv.first_day)
                    THEN u.user_id END) as new_users
            FROM user_stats u
            LEFT JOIN (
                SELECT user_id, MIN(DATE(created_at)) as first_day
                FROM user_stats WHERE user_id != ? GROUP BY user_id
            ) fv ON u.user_id = fv.user_id
            WHERE u.user_id != ?
            GROUP BY month ORDER BY month DESC LIMIT 12
        """, (admin_filter, admin_filter))
        monthly_activity = cursor.fetchall()

        cursor.execute("""SELECT strftime('%Y-%m', created_at) as month, COUNT(DISTINCT user_id) as users
            FROM user_stats WHERE action IN ('open_webapp','webapp_ping') AND user_id != ?
            GROUP BY month""", (admin_filter,))
        webapp_by_month = {r["month"]: r["users"] for r in cursor.fetchall()}

        # ── Топ действий ─────────────────────────────────────────────
        cursor.execute("""SELECT action, COUNT(*) as cnt FROM user_stats
            WHERE user_id != ? GROUP BY action ORDER BY cnt DESC LIMIT 10""", (admin_filter,))
        top_actions = cursor.fetchall()

        # ── База событий ─────────────────────────────────────────────
        cursor.execute("SELECT COUNT(*) FROM events WHERE event_date >= ?", (today,))
        events_count = cursor.fetchone()[0]
        cursor.execute("SELECT source_name, COUNT(*) as cnt FROM events WHERE event_date >= ? GROUP BY source_name ORDER BY cnt DESC", (today,))
        events_by_source = cursor.fetchall()

        # ── Подписки (категории) ─────────────────────────────────────
        cursor.execute("""SELECT COUNT(DISTINCT user_id) FROM subscriptions
            WHERE user_id != ? AND (status IS NULL OR status = 'active')""", (admin_filter,))
        subscribers_count = cursor.fetchone()[0]
        cursor.execute("""SELECT COUNT(*) FROM subscriptions
            WHERE user_id != ? AND (status IS NULL OR status = 'active')""", (admin_filter,))
        subscriptions_total = cursor.fetchone()[0]

        # ── Флеш-подписки ────────────────────────────────────────────
        cursor.execute("""SELECT COUNT(*) FROM flash_subscriptions
            WHERE status = 'active' AND user_id != ?""", (admin_filter,))
        flash_total = cursor.fetchone()[0]
        cursor.execute("""SELECT COUNT(DISTINCT user_id) FROM flash_subscriptions
            WHERE status = 'active' AND user_id != ?""", (admin_filter,))
        flash_users = cursor.fetchone()[0]
        cursor.execute("""SELECT COUNT(*) FROM flash_subscriptions
            WHERE status = 'active' AND user_id != ? AND DATE(created_at) = ?""", (admin_filter, today))
        flash_new_today = cursor.fetchone()[0]
        cursor.execute("""SELECT COUNT(*) FROM flash_subscriptions
            WHERE status = 'active' AND user_id != ? AND created_at >= DATE('now', '-30 days')""", (admin_filter,))
        flash_new_30d = cursor.fetchone()[0]
        cursor.execute("""SELECT COUNT(DISTINCT user_id) FROM flash_subscriptions
            WHERE user_id != ? AND last_notified_at != '' AND last_notified_at >= DATE('now', '-30 days')""", (admin_filter,))
        flash_notified_users_30d = cursor.fetchone()[0]

        return {
            "total_users": total_users,
            "days_alive": days_alive,
            "total_actions": total_actions,
            "dau": dau,
            "wau": wau,
            "mau": mau,
            "actions_today": actions_today,
            "new_today": new_today,
            "new_30d": new_30d,
            "webapp_total": webapp_total,
            "webapp_dau": webapp_dau,
            "webapp_wau": webapp_wau,
            "webapp_mau": webapp_mau,
            "daily_activity": daily_activity,
            "monthly_activity": monthly_activity,
            "webapp_by_day": webapp_by_day,
            "webapp_by_month": webapp_by_month,
            "top_actions": top_actions,
            "events_count": events_count,
            "events_by_source": events_by_source,
            "subscribers_count": subscribers_count,
            "subscriptions_total": subscriptions_total,
            "flash_total": flash_total,
            "flash_users": flash_users,
            "flash_new_today": flash_new_today,
            "flash_new_30d": flash_new_30d,
            "flash_notified_users_30d": flash_notified_users_30d,
        }


def get_raw_events_count_by_category() -> dict:
    """Полное кол-во строк в БД по категориям (для 'О проекте')."""
    today = datetime.now(MINSK_TZ).strftime("%Y-%m-%d")
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT category, COUNT(*) FROM events WHERE event_date >= ? GROUP BY category",
            (today,),
        )
        return {row[0]: row[1] for row in cursor.fetchall()}


def get_events_count_by_category() -> dict:
    """Кол-во сгруппированных событий по категориям."""
    today = datetime.now(MINSK_TZ).strftime("%Y-%m-%d")
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Обычные категории (все события, включая бесплатные)
        cursor.execute("""
            SELECT category, COUNT(*) FROM (
                SELECT DISTINCT category, title, COALESCE(place, '') as place 
                FROM events
                WHERE category IS NOT NULL AND category != '' AND category != 'cinema'
                AND event_date >= ?
            ) GROUP BY category
        """, (today,))
        result = {row[0]: row[1] for row in cursor.fetchall()}
        
        # Кино (отдельная логика)
        cursor.execute("""
            SELECT COUNT(*) FROM (
                SELECT DISTINCT title, event_date FROM events
                WHERE category = 'cinema' AND event_date >= ?
            )
        """, (today,))
        result["cinema"] = cursor.fetchone()[0]
        
        # FREE - ВСЕ бесплатные события (независимо от категории)
        cursor.execute("""
            SELECT COUNT(*) FROM (
                SELECT DISTINCT title, COALESCE(place, ''), event_date 
                FROM events
                WHERE event_date >= ? AND price = 'Бесплатно'
            )
        """, (today,))
        result["free"] = cursor.fetchone()[0]
        
        return result


def search_events_by_title(query: str, limit: int = 20):
    today = datetime.now(MINSK_TZ).strftime("%Y-%m-%d")
    q = query.lower()
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, title, details, description, event_date, show_time, end_time,
                   place, location, price, category, source_url
            FROM events
            WHERE (pylow(title) LIKE ? OR pylow(details) LIKE ? OR pylow(place) LIKE ?)
              AND event_date >= ?
            ORDER BY event_date, show_time, title
            LIMIT ?
        """, (f"%{q}%", f"%{q}%", f"%{q}%", today, limit * SEARCH_MULTIPLIER))
        return cursor.fetchall()


def search_events_by_date_raw(date_str: str):
    current_year = datetime.now(MINSK_TZ).year
    date_str = date_str.strip()
    if re.match(r"^\d{1,2}\.\d{1,2}\.\d{4}$", date_str):
        day, month, year = date_str.split(".")
    elif re.match(r"^\d{1,2}\.\d{1,2}$", date_str):
        day, month = date_str.split(".")
        year = str(current_year)
    else:
        return None, None, "неверный_формат"
    day, month = day.zfill(2), month.zfill(2)
    search_date = f"{year}-{month}-{day}"
    formatted_date = f"{day}.{month}.{year}"
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, title, details, description, event_date, show_time, end_time,
                   place, location, price, category, source_url
            FROM events WHERE event_date = ? ORDER BY show_time, title LIMIT 300
        """, (search_date,))
        events = cursor.fetchall()
    return (events, formatted_date, "найдены") if events else ([], formatted_date, "нет_событий")


def get_events_by_date_and_category(target_date: datetime, category: str | None = None):
    """События на дату. Для сегодня фильтрует прошедшие сеансы (по времени Минска)."""
    now_minsk = datetime.now(MINSK_TZ)
    date_str = target_date.strftime("%Y-%m-%d")
    today_str = now_minsk.strftime("%Y-%m-%d")
    now_time = now_minsk.strftime("%H:%M")
    
    with get_db_connection() as conn:
        cursor = conn.cursor()

        # Вспомогательная функция для фильтра времени
        def add_time_filter(query, params, date_str, now_time):
            """Добавляет условие для фильтрации прошедших сеансов."""
            venue_open = VENUE_OPEN_TIME <= now_time < VENUE_CLOSE_TIME
            if venue_open:
                time_filter = """
                    AND (
                        show_time = '' OR show_time IS NULL
                        OR (
                            (end_time != '' AND end_time IS NOT NULL AND end_time > ?)
                            OR
                            ((end_time = '' OR end_time IS NULL) AND show_time > ?)
                        )
                    )
                """
            else:
                time_filter = """
                    AND (
                        (end_time != '' AND end_time IS NOT NULL AND end_time > ?)
                        OR
                        ((end_time = '' OR end_time IS NULL) AND show_time != '' AND show_time IS NOT NULL AND show_time > ?)
                    )
                """
            query += time_filter
            params.append(now_time)  # для end_time
            params.append(now_time)  # для show_time
            return query, params
        
        # ОСОБЫЙ СЛУЧАЙ: категория "free" показывает ВСЕ бесплатные события
        if category == "free":
            query = """
                SELECT id, title, details, description, event_date, show_time, end_time,
                       place, location, price, category, source_url
                FROM events 
                WHERE event_date = ? AND price = 'Бесплатно'
            """
            params = [date_str]
            
            # Для сегодня — исключаем прошедшие сеансы
            if date_str == today_str:
                time_filter, time_params = _build_time_filter(date_str, today_str, now_time)
                if time_filter:
                    query += " AND " + time_filter
                params.extend(time_params)
            
            query += " ORDER BY show_time, title"
            cursor.execute(query, params)
            return cursor.fetchall()
        
        # Обычная категория (не free)
        query = """
            SELECT id, title, details, description, event_date, show_time, end_time,
                   place, location, price, category, source_url
            FROM events WHERE event_date = ?
        """
        params = [date_str]
        
        if category and category != "all":
            query += " AND category = ?"
            params.append(category)
        
        if date_str == today_str:
            time_filter, time_params = _build_time_filter(date_str, today_str, now_time)
            if time_filter:
                query += " AND " + time_filter
            params.extend(time_params)
        
        query += " ORDER BY show_time, title"
        cursor.execute(query, params)
        return cursor.fetchall()


def get_upcoming_events(limit: int = 20, category: str | None = None):
    now_minsk = datetime.now(MINSK_TZ)
    today = now_minsk.strftime("%Y-%m-%d")
    now_time = now_minsk.strftime("%H:%M")

    with get_db_connection() as conn:
        cursor = conn.cursor()

        # Условие для фильтрации прошедших событий
        venue_open = VENUE_OPEN_TIME <= now_time < VENUE_CLOSE_TIME
        if venue_open:
            time_filter = """
                AND (
                    event_date > ?
                    OR (
                        show_time = '' OR show_time IS NULL
                        OR (
                            (end_time != '' AND end_time IS NOT NULL AND end_time > ?)
                            OR
                            ((end_time = '' OR end_time IS NULL) AND show_time > ?)
                        )
                    )
                )
            """
        else:
            time_filter = """
                AND (
                    event_date > ?
                    OR (
                        (end_time != '' AND end_time IS NOT NULL AND end_time > ?)
                        OR
                        ((end_time = '' OR end_time IS NULL) AND show_time != '' AND show_time IS NOT NULL AND show_time > ?)
                    )
                )
            """
        
        # ОСОБЫЙ СЛУЧАЙ: категория "free" показывает ВСЕ бесплатные события
        if category == "free":
            cursor.execute(f"""
                SELECT id, title, details, description, event_date, show_time, end_time,
                       place, location, price, category, source_url
                FROM events 
                WHERE event_date >= ? AND price = 'Бесплатно'
                {time_filter}
                ORDER BY event_date, show_time, title LIMIT ?
            """, (today, today, now_time, now_time, limit * SEARCH_MULTIPLIER))
            return cursor.fetchall()
        
        # Обычная категория (не free)
        if category and category != "all":
            cursor.execute(f"""
                SELECT id, title, details, description, event_date, show_time, end_time,
                       place, location, price, category, source_url
                FROM events WHERE event_date >= ? AND category = ?
                {time_filter}
                ORDER BY event_date, show_time, title LIMIT ?
            """, (today, category, today, now_time, now_time, limit * SEARCH_MULTIPLIER))
        else:
            cursor.execute(f"""
                SELECT id, title, details, description, event_date, show_time, end_time,
                       place, location, price, category, source_url
                FROM events WHERE event_date >= ?
                {time_filter}
                ORDER BY event_date, show_time, title LIMIT ?
            """, (today, today, now_time, now_time, limit * SEARCH_MULTIPLIER))
        
        return cursor.fetchall()


def get_weekend_events(category: str | None = None):
    """
    Возвращает события на ближайшие выходные (суббота и воскресенье).
    Для категории free: все события с ценой 'Бесплатно'
    Для остальных категорий: события по категории
    """
    today = datetime.now(MINSK_TZ)
    days_until_saturday = (5 - today.weekday()) % 7 or 7
    saturday = today + timedelta(days=days_until_saturday)
    sunday = saturday + timedelta(days=1)
    saturday_str, sunday_str = saturday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d")
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # ОСОБЫЙ СЛУЧАЙ: категория "free" показывает ВСЕ бесплатные события
        if category == "free":
            cursor.execute("""
                SELECT id, title, details, description, event_date, show_time, end_time,
                       place, location, price, category, source_url
                FROM events 
                WHERE event_date IN (?, ?) AND price = 'Бесплатно'
                ORDER BY event_date, show_time, title
            """, (saturday_str, sunday_str))
            return cursor.fetchall(), saturday, sunday
        
        # Обычная категория (не free)
        if category and category != "all":
            cursor.execute("""
                SELECT id, title, details, description, event_date, show_time, end_time,
                       place, location, price, category, source_url
                FROM events WHERE event_date IN (?, ?) AND category = ?
                ORDER BY event_date, show_time, title
            """, (saturday_str, sunday_str, category))
        else:
            cursor.execute("""
                SELECT id, title, details, description, event_date, show_time, end_time,
                       place, location, price, category, source_url
                FROM events WHERE event_date IN (?, ?)
                ORDER BY event_date, show_time, title
            """, (saturday_str, sunday_str))
        
        return cursor.fetchall(), saturday, sunday


def filter_events_by_category(events, category: str):
    """Фильтрует события по категории. Для free фильтрует по цене."""
    if category == "free":
        return [e for e in events if e.get("price") == "Бесплатно"]
    return [e for e in events if e.get("category") == category]


def add_subscription(user_id: int, category: str, date_type: str):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT OR IGNORE INTO subscriptions (user_id, category, date_type, status) VALUES (?, ?, ?, 'active')",
            (user_id, category, date_type),
        )
        conn.commit()


def remove_subscription(user_id: int, category: str, date_type: str):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        # Ставим status='inactive' вместо DELETE — история сохраняется
        cursor.execute(
            "UPDATE subscriptions SET status='inactive' WHERE user_id = ? AND category = ? AND date_type = ?",
            (user_id, category, date_type)
        )
        conn.commit()


def get_user_subscriptions(user_id: int):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT category, date_type FROM subscriptions WHERE user_id = ? AND (status IS NULL OR status = 'active')", (user_id,))
        return cursor.fetchall()


def get_all_subscribers() -> dict:
    """Возвращает {(category, date_type): [user_id, ...]} — только активные подписки."""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT user_id, category, date_type FROM subscriptions "
            "WHERE status IS NULL OR status = 'active'"
        )
        result = defaultdict(list)
        for row in cursor.fetchall():
            result[(row["category"], row["date_type"])].append(row["user_id"])
        return result


# ── Флеш-подписки ──────────────────────────────────────────────────────────

def add_flash_subscription(user_id: int, query: str) -> bool:
    """Добавляет флеш-подписку. Возвращает False если такая уже есть."""
    with get_db_connection() as conn:
        # Проверяем дубликат
        existing = conn.execute(
            "SELECT id FROM flash_subscriptions WHERE user_id=? AND LOWER(query)=LOWER(?) AND status='active'",
            (user_id, query)
        ).fetchone()
        if existing:
            return False
        conn.execute(
            "INSERT INTO flash_subscriptions (user_id, query, created_at, status) VALUES (?,?,?,'active')",
            (user_id, query, datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
        return True


def remove_flash_subscription(flash_id: int, user_id: int):
    with get_db_connection() as conn:
        conn.execute(
            "UPDATE flash_subscriptions SET status='inactive' WHERE id=? AND user_id=?",
            (flash_id, user_id)
        )
        conn.commit()


def get_user_flash_subscriptions(user_id: int) -> list:
    with get_db_connection() as conn:
        return conn.execute(
            "SELECT id, query, created_at FROM flash_subscriptions WHERE user_id=? AND status='active' ORDER BY created_at DESC",
            (user_id,)
        ).fetchall()


def get_all_flash_subscriptions() -> list:
    """Все активные флеш-подписки для проверки после парсинга."""
    with get_db_connection() as conn:
        return conn.execute(
            "SELECT id, user_id, query FROM flash_subscriptions WHERE status='active'"
        ).fetchall()


async def check_flash_subscriptions(bot) -> int:
    """Проверяет новые события против всех флеш-подписок и рассылает совпадения.
    Каждая подписка — отдельное сообщение пользователю.
    Повторно не уведомляет: сравнивает created_at события с last_notified_at подписки.
    Вызывается после каждого парсинга."""
    subs = get_all_flash_subscriptions()
    if not subs:
        return 0

    today = datetime.now(MINSK_TZ).strftime("%Y-%m-%d")
    sent_total = 0
    import html as _html
    from telegram.error import RetryAfter

    with get_db_connection() as conn:
        for sub in subs:
            user_id = sub["user_id"]
            q = sub["query"]
            ql = q.lower()
            sp = f"%{q}%"
            spl = f"%{ql}%"

            # Берём last_notified_at для этой подписки (может отсутствовать в старых записях)
            sub_row = conn.execute(
                "SELECT last_notified_at FROM flash_subscriptions WHERE id = ?",
                (sub["id"],)
            ).fetchone()
            last_notified = (sub_row["last_notified_at"] or "") if sub_row else ""

            # Показываем только события добавленные ПОСЛЕ последнего уведомления
            # Для новых подписок (last_notified пуст) — показываем все актуальные события
            date_condition = "AND event_date >= ?" if not last_notified else "AND event_date >= ? AND (created_at IS NULL OR created_at > ?)"
            date_params = (today,) if not last_notified else (today, last_notified)

            rows = conn.execute(f"""
                SELECT DISTINCT id, title, details, event_date, show_time, place, price, category, source_url
                FROM events
                WHERE event_date >= ?
                AND (LOWER(title) LIKE ? OR title LIKE ?
                     OR LOWER(details) LIKE ? OR details LIKE ?)
                {"AND (created_at IS NULL OR created_at > ?)" if last_notified else ""}
                ORDER BY event_date, show_time
                LIMIT 5
            """, (today, spl, sp, spl, sp, *((last_notified,) if last_notified else ()))).fetchall()

            if not rows:
                continue

            lines = [f"⚡ <b>Флеш-подписка: «{_html.escape(q)}»</b>\n"]
            for e in rows:
                cat_emoji = CATEGORY_EMOJI.get(e["category"] or "", "🎉")
                title = _html.escape(e["title"] or "")
                try:
                    date_str = datetime.strptime(e["event_date"], "%Y-%m-%d").strftime("%d.%m.%Y")
                except Exception:
                    date_str = e["event_date"] or ""
                time_str = f" ⏰ {e['show_time']}" if e["show_time"] else ""
                place_str = f"\n🏢 {_html.escape(e['place'])}" if e["place"] else ""
                price_str = f" | 💰 {e['price']}" if e["price"] else ""
                url = e["source_url"] or ""
                title_link = f"<a href=\"{url}\">{title}</a>" if url else title
                lines.append(f"{cat_emoji} {title_link}\n📅 {date_str}{time_str}{place_str}{price_str}\n")

            lines.append("👉 @Minskdvizh_bot")
            text = "\n".join(lines)

            sub_id = sub["id"]
            confirm_keyboard = InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ Да, нашёл!", callback_data=f"flash_found_{sub_id}"),
                InlineKeyboardButton("🔄 Нет, искать дальше", callback_data=f"flash_continue_{sub_id}"),
            ]])

            try:
                await bot.send_message(
                    chat_id=user_id, text=text,
                    parse_mode="HTML", disable_web_page_preview=True,
                    reply_markup=confirm_keyboard,
                )
                sent_total += 1
                # Обновляем время последнего уведомления
                conn.execute(
                    "UPDATE flash_subscriptions SET last_notified_at = ? WHERE id = ?",
                    (datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"), sub["id"])
                )
                conn.commit()
                await asyncio.sleep(0.1)
            except RetryAfter as e:
                logger.warning(f"Флеш-рассылка RetryAfter {e.retry_after}с для {user_id}")
                await asyncio.sleep(e.retry_after + 1)
            except Exception as e:
                logger.warning(f"Флеш-рассылка {user_id} «{q}»: {e}")

    logger.info(f"⚡ Флеш-подписки: разослано {sent_total} уведомлений")
    return sent_total


# ---------------------- Форматирование ----------------------


def format_event_text(event) -> str:
    import html as _html
    title = _html.escape(event["title"] or "")
    _emoji = CATEGORY_EMOJI.get(event.get("category") or "", "🎉")
    text = f"{_emoji} <b>{title}</b>"
    if event["details"]:
        details = event["details"][:177] + "..." if len(event["details"]) > 180 else event["details"]
        text += f"\n📝 {_html.escape(details)}"
    if event["event_date"]:
        text += f"\n📅 {datetime.strptime(event['event_date'], '%Y-%m-%d').strftime('%d.%m.%Y')}"
    if event["show_time"]:
        _t = event["show_time"]
        try:
            _et = event["end_time"] or ""
        except Exception:
            _et = ""
        if _et:
            text += f" ⏰ {_t}–{_et}"
        else:
            text += f" ⏰ {_t}"
    if event["place"] and event["place"] != "Кинотеатр":
        text += f"\n🏢 {event['place']}"
    if event["price"]:
        text += f"\n💰 {event['price']}"
    return text


def group_cinema_events(events):
    """Группировка: title → date → place → [сеансы].
    Ключ пагинации = (title, date) — один фильм в один день = одна запись."""
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for event in events:
        if event["category"] == "cinema":
            grouped[event["title"]][event["event_date"]][event["place"]].append(
                {"time": event["show_time"], "details": event["details"]}
            )
    return grouped


def format_grouped_cinema_events(grouped):
    """Каждый элемент result = один фильм в один день (все кинотеатры внутри).
    Пагинация по этому списку даёт 10 фильмов на страницу."""
    result = []
    for title, dates in grouped.items():
        for date, cinemas in dates.items():
            details = ""
            film_url = ""
            for seances in cinemas.values():
                for s in seances:
                    if not details and s.get("details"):
                        details = s["details"]
                    if not film_url and s.get("url"):
                        film_url = s["url"]
            text = f"🎬 <b>{title}</b>"
            if details:
                details = details[:100] + "..." if len(details) > 100 else details
                text += f"\n🎭 {details}"
            text += f"\n📅 {datetime.strptime(date, '%Y-%m-%d').strftime('%d.%m.%Y')}"
            for place, seances in cinemas.items():
                times = sorted([s["time"] for s in seances if s["time"]])
                times_str = ", ".join(times) if times else "—"
                text += f"\n📍 {place}: {times_str}"
            result.append((text, film_url))
    return result


# ---------------------- Пагинация + категории ----------------------


def group_other_events(events: list) -> list:
    from collections import OrderedDict, defaultdict
    EMOJI_MAP = CATEGORY_EMOJI
    grouped = OrderedDict()
    title_to_key = {}

    for e in events:
        title = e.get("title", "")
        place = e.get("place", "")

        if place:
            key = (title, place)
            title_to_key.setdefault(title, key)
        else:
            key = title_to_key.get(title, (title, ""))

        if key not in grouped:
            grouped[key] = {
                "title": title, "place": place,
                "price": e.get("price", ""), "category": e.get("category", ""),
                "source_url": e.get("source_url", ""), "dates": []
            }
            if place:
                title_to_key[title] = key
        else:
            if not grouped[key]["place"] and place:
                grouped[key]["place"] = place
            if not grouped[key]["price"] and e.get("price"):
                grouped[key]["price"] = e["price"]
            if not grouped[key]["source_url"] and e.get("source_url"):
                grouped[key]["source_url"] = e["source_url"]

        # Сохраняем show_time и end_time вместе
        show_time = e.get("show_time", "")
        end_time = e.get("end_time", "")
        grouped[key]["dates"].append((
            e.get("event_date", ""),
            show_time,
            end_time
        ))
        if not grouped[key]["price"] and e.get("price"):
            grouped[key]["price"] = e["price"]

    result = []
    for g in grouped.values():
        cat_emoji = EMOJI_MAP.get(g["category"], "🎉")
        text = f"{cat_emoji} <b>{g['title']}</b>"
        if g["place"]:
            text += f"\n🏢 {g['place']}"
        if g["price"]:
            text += f"\n💰 {g['price']}"

        # Группируем по времени (show_time + end_time)
        dates_sorted = sorted(set(g["dates"]))  # [(date_str, show_time, end_time), ...]
        by_time = defaultdict(list)
        for date_str, show_time, end_time in dates_sorted:
            time_key = f"{show_time}|{end_time}" if end_time else show_time
            by_time[time_key].append(date_str)

        def make_ranges(date_strs):
            try:
                ds = sorted(datetime.strptime(d, "%Y-%m-%d") for d in date_strs)
            except Exception:
                return date_strs
            ranges = []
            start = end = ds[0]
            for d in ds[1:]:
                if (d - end).days == 1:
                    end = d
                else:
                    ranges.append((start, end))
                    start = end = d
            ranges.append((start, end))
            result = []
            for s, e in ranges:
                if s == e:
                    result.append(s.strftime("%d.%m.%Y"))
                elif s.month == e.month and s.year == e.year:
                    result.append(f"{s.strftime('%d')}–{e.strftime('%d.%m.%Y')}")
                else:
                    result.append(f"{s.strftime('%d.%m')}–{e.strftime('%d.%m.%Y')}")
            return result

        for time_key, date_strs in by_time.items():
            if "|" in time_key:
                show_time, end_time = time_key.split("|")
                time_display = f"{show_time}–{end_time}"
            else:
                time_display = time_key
            
            ranges = make_ranges(date_strs)
            for r in ranges:
                text += f"\n📅 {r}" + (f" ⏰ {time_display}" if time_display else "")

        # _sort_key: (date, show_time) — минимальный первый сеанс
        # min() по кортежу (date_str, show_time, end_time) → берём date и show_time
        if g["dates"]:
            first = min(g["dates"])   # (date_str, show_time, end_time)
            first_date = first[0]
            first_time = first[1] or ""
        else:
            first_date = "9999"
            first_time = ""

        result.append({
            "_pre_formatted": True, "text": text,
            "url": g["source_url"], "category": g["category"],
            "_sort_key": (first_date, first_time)
        })

    result.sort(key=lambda x: x.get("_sort_key", ("9999", "")))
    return result


def pre_group_for_pagination(events: list) -> list:
    """Группирует события ДО пагинации:
    - кино: title+date → все кинотеатры/сеансы
    - остальные: title+place → все даты/времена"""
    events = [dict(e) if not isinstance(e, dict) else e for e in events]
    cinema = [e for e in events if e.get("category") == "cinema"]
    other  = [e for e in events if e.get("category") != "cinema"]
    result = []
    if cinema:
        # Строим индекс title → (min_date, min_time) из сырых событий
        cinema_sort: dict[str, tuple] = {}
        for e in cinema:
            key = e.get("title", "")
            dt = (e.get("event_date", "9999-12-31"), e.get("show_time", ""))
            if key not in cinema_sort or dt < cinema_sort[key]:
                cinema_sort[key] = dt
        grouped_items = format_grouped_cinema_events(group_cinema_events(cinema))
        for t, u in grouped_items:
            # Заголовок фильма — первая жирная строка
            m = re.search(r"<b>(.*?)</b>", t)
            title_key = m.group(1) if m else ""
            sort_key = cinema_sort.get(title_key, ("9999-12-31", ""))
            result.append({"_pre_formatted": True, "text": t, "url": u,
                           "category": "cinema", "_sort_key": sort_key})
    if other:
        result.extend(group_other_events(other))
    # Сортируем всё вместе по дате/времени
    result.sort(key=lambda x: x.get("_sort_key", ("9999", "")))
    return result


def set_pagination(context: ContextTypes.DEFAULT_TYPE, events, title: str, date_info: str | None = None,
                   share_query: str = ""):
    raw = [dict(e) if not isinstance(e, dict) else e for e in events]
    context.user_data["pagination"] = {
        "events": raw, "page": 0, "per_page": PER_PAGE,
        "title": title, "date_info": date_info,
        "share_query": share_query,
    }


def build_page_keyboard(data: dict):
    """Клавиатура: фильтры категорий + навигация ◀ 1/5 ▶."""
    events = data.get("_grouped", data["events"])
    page = data["page"]
    per_page = data["per_page"]
    total = len(events)
    max_page = max(0, (total - 1) // per_page)
    keyboard = []
    
    # Считаем уникальные события (как после группировки) — title+place
    category_counts = defaultdict(int)
    _seen_cats: dict = defaultdict(set)
    
    for e in events:
        cat = e.get("category") if e.get("category") else ("cinema" if e.get("_pre_formatted") else None)
        if not cat:
            continue
        if e.get("_pre_formatted"):
            category_counts[cat] += 1
        else:
            key = (e.get("title", ""), e.get("place") or "")
            if key not in _seen_cats[cat]:
                _seen_cats[cat].add(key)
                category_counts[cat] += 1
    
    # Считаем бесплатные события из raw_events по price='Бесплатно' —
    # сюда попадают события ЛЮБОЙ категории с этой ценой, не только category='free'.
    # Всегда перезаписываем category_counts["free"], чтобы не занижать счётчик
    # в случае когда среди событий есть и category='free' (1 шт.) и price='Бесплатно'
    # у событий других категорий (ещё N шт.).
    raw_events = data.get("events", [])
    free_count = len({
        (e.get("title", ""), e.get("event_date", ""), e.get("place") or "")
        for e in raw_events
        if (e.get("price") or "") == "Бесплатно"
    })
    if free_count > 0:
        category_counts["free"] = free_count  # перезаписываем всегда
    
    # Кнопки фильтрации по категориям
    if len(category_counts) > 1:
        row = []
        for cat_key, cat_name in CATEGORY_NAMES.items():
            if cat_key in category_counts:
                count = category_counts[cat_key]
                btn_text = f"{cat_name} ({count})" if count > 0 else cat_name
                row.append(InlineKeyboardButton(btn_text, callback_data=f"filter_{cat_key}"))
                if len(row) == 2:
                    keyboard.append(row)
                    row = []
        if row:
            keyboard.append(row)
    
    # Кнопки пагинации
    if max_page > 0:
        keyboard.append([
            InlineKeyboardButton("◀️", callback_data="page_prev") if page > 0 else InlineKeyboardButton(" ", callback_data="page_noop"),
            InlineKeyboardButton(f"{page + 1}/{max_page + 1}", callback_data="page_noop"),
            InlineKeyboardButton("▶️", callback_data="page_next") if page < max_page else InlineKeyboardButton(" ", callback_data="page_noop"),
        ])
    
    # Кнопка поделиться
    share_q = data.get("share_query") or ""
    keyboard.append([
        InlineKeyboardButton("📤 Поделиться подборкой", switch_inline_query=share_q)
    ])
    
    return InlineKeyboardMarkup(keyboard) if keyboard else None

async def show_page(update_or_query, context: ContextTypes.DEFAULT_TYPE):
    data = context.user_data.get("pagination")
    if not data:
        msg = "Данные не найдены. Попробуйте запрос заново."
        if isinstance(update_or_query, Update):
            await update_or_query.message.reply_text(msg)
        else:
            await update_or_query.answer(msg, show_alert=True)
        return
    raw_events = data["events"]
    # Группируем сырые события; кешируем чтобы не пересчитывать при листании
    cache_key = len(raw_events)  # простой ключ — количество сырых событий
    if data.get("_grouped_key") != cache_key or "_grouped" not in data:
        data["_grouped"] = pre_group_for_pagination(raw_events)
        data["_grouped_key"] = cache_key
    events = data["_grouped"]

    page, per_page = data["page"], data["per_page"]
    total = len(events)
    if total == 0:
        msg = "😕 Событий не найдено."
        if isinstance(update_or_query, Update):
            await update_or_query.message.reply_text(msg)
        else:
            await update_or_query.answer()
            await update_or_query.message.reply_text(msg)
        return
    max_page = (total - 1) // per_page
    page = max(0, min(page, max_page))
    data["page"] = page
    chunk = events[page * per_page:(page + 1) * per_page]
    if isinstance(update_or_query, Update):
        await update_or_query.message.chat.send_action(action="typing")
        send = update_or_query.message.reply_text
    else:
        await update_or_query.answer()
        send = update_or_query.message.reply_text
    lines = []
    if data.get("title"): lines.append(data["title"])
    if data.get("date_info"): lines.append(data["date_info"])
    lines.append(f"Найдено: {total} | Стр. {page + 1}/{max_page + 1}")
    lines.append("")
    for item in chunk:
        if item.get("_pre_formatted"):
            film_url = item.get("url") or "https://afisha.relax.by/kino/minsk/"
            lines.append(item["text"] + f"\n🔗 <a href=\"{film_url}\">Подробнее</a>")
        else:
            url = item.get("source_url", "")
            suffix = f"\n🔗 <a href=\"{url}\">Подробнее</a>" if url else ""
            lines.append(format_event_text(item) + suffix)
        lines.append("")
    text = "\n".join(lines).strip()
    keyboard = build_page_keyboard(data)
    if len(text) <= 4096:
        await send(text, reply_markup=keyboard, parse_mode="HTML", disable_web_page_preview=True)
    else:
        # Текст >4000 — делим на части, склеивая события в блоки до 4000 символов
        header = f"{data.get('title', '')}\nНайдено: {total} | Стр. {page + 1}/{max_page + 1}\n"
        all_texts = []
        for item in chunk:
            if item.get("_pre_formatted"):
                film_url = item.get("url") or "https://afisha.relax.by/kino/minsk/"
                all_texts.append(item["text"] + f"\n🔗 <a href=\"{film_url}\">Подробнее</a>")
            else:
                url = item.get("source_url", "") or ""
                suffix = f"\n🔗 <a href=\"{url}\">Подробнее</a>" if url else ""
                all_texts.append(format_event_text(item) + suffix)
        # Склеиваем в сообщения до 4000 символов
        parts = []
        current = header
        for t in all_texts:
            candidate = current + "\n" + t + "\n"
            if len(candidate) > 4096 and current != header:
                parts.append(current.strip())
                current = header + t + "\n"
            else:
                current = candidate
        if current.strip() != header.strip():
            parts.append(current.strip())
        for idx, part in enumerate(parts):
            is_last = idx == len(parts) - 1
            await send(part, reply_markup=keyboard if is_last else None,
                       parse_mode="HTML", disable_web_page_preview=True)


# ---------------------- Календарь ----------------------


def build_calendar_keyboard(year: int, month: int, available_dates: set) -> InlineKeyboardMarkup:
    import calendar as cal_module
    MONTH_NAMES = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                   "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    keyboard = [
        [InlineKeyboardButton("◀", callback_data=f"cal_prev_{year}_{month}"),
         InlineKeyboardButton(f"{MONTH_NAMES[month]} {year}", callback_data="page_noop"),
         InlineKeyboardButton("▶", callback_data=f"cal_next_{year}_{month}")],
        [InlineKeyboardButton(d, callback_data="page_noop") for d in ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]]
    ]
    from datetime import date as date_cls
    today_date = date_cls.today()
    for week in cal_module.monthcalendar(year, month):
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(" ", callback_data="page_noop"))
            else:
                this_date = date_cls(year, month, day)
                if this_date < today_date:
                    row.append(InlineKeyboardButton(" ", callback_data="page_noop"))
                else:
                    date_str = f"{year}-{month:02d}-{day:02d}"
                    if date_str in available_dates:
                        row.append(InlineKeyboardButton(str(day), callback_data=f"cal_day_{year}_{month}_{day}"))
                    else:
                        row.append(InlineKeyboardButton(f"·{day}", callback_data="page_noop"))
        keyboard.append(row)
    return InlineKeyboardMarkup(keyboard)


def get_available_dates() -> set:
    today = datetime.now(MINSK_TZ).strftime("%Y-%m-%d")
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT event_date FROM events WHERE event_date >= ?", (today,))
        return {row["event_date"] for row in cursor.fetchall()}


async def show_calendar(update_or_query, context: ContextTypes.DEFAULT_TYPE, year: int = None, month: int = None):
    now = datetime.now(MINSK_TZ)
    year = year or now.year
    month = month or now.month
    keyboard = build_calendar_keyboard(year, month, get_available_dates())
    text = "🗓 Выберите дату (активны даты с событиями):"
    if isinstance(update_or_query, Update):
        await update_or_query.message.reply_text(text, reply_markup=keyboard)
    else:
        await update_or_query.answer()
        try:
            await update_or_query.edit_message_text(text, reply_markup=keyboard)
        except Exception:
            await update_or_query.message.reply_text(text, reply_markup=keyboard)


# ---------------------- UI-хелперы ----------------------


def get_reply_main_menu():
    return ReplyKeyboardMarkup([
        ["📅 Сегодня", "📆 Завтра"],
        ["⏰ Ближайшие", "🎉 Выходные"],
        ["🗓 Календарь", "🎯 Категории"],
        ["ℹ️ О проекте", "⭐ Поддержать"],
    ], resize_keyboard=True)


async def show_main_menu(chat_id: int, context: ContextTypes.DEFAULT_TYPE | None = None, send_method=None):
    text = "🎉 **Главное меню**\n\nВыберите действие:"
    kwargs = {"reply_markup": get_reply_main_menu(), "parse_mode": "Markdown"}
    if send_method:
        await send_method(text, **kwargs)
    else:
        await context.bot.send_message(chat_id=chat_id, text=text, **kwargs)


async def show_categories_menu(query, context: ContextTypes.DEFAULT_TYPE):
    await query.answer()
    counts = get_events_count_by_category()
    # Строим список: сначала категории из CATEGORY_NAMES (в правильном порядке),
    # потом неизвестные категории из БД
    ordered = {}
    for cat in CATEGORY_NAMES:
        if counts.get(cat, 0) > 0:
            ordered[cat] = CATEGORY_NAMES[cat]
    for cat in counts:
        if cat not in ordered and counts[cat] > 0:
            # Неизвестная категория — показываем с emoji из CATEGORY_EMOJI или дефолт
            emoji = CATEGORY_EMOJI.get(cat, "📌")
            ordered[cat] = f"{emoji} {cat.replace('_', ' ').capitalize()}"
    keyboard = []
    row = []
    for cat, name in ordered.items():
        n = counts[cat]
        label = f"{name} ({n})"
        row.append(InlineKeyboardButton(label, callback_data=f"cat_{cat}"))
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    if not keyboard:
        await query.edit_message_text("😔 Пока нет доступных событий.", parse_mode="Markdown")
        return
    await query.edit_message_text("🎯 **Выберите категорию:**", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")


async def show_date_options(update_or_query, category_name: str):
    display_name = CATEGORY_NAMES.get(category_name, category_name)
    keyboard = [
        [InlineKeyboardButton("📅 Сегодня", callback_data=f"date_today_{category_name}"),
         InlineKeyboardButton("📆 Завтра", callback_data=f"date_tomorrow_{category_name}")],
        [InlineKeyboardButton("⏰ Ближайшие", callback_data=f"date_upcoming_{category_name}"),
         InlineKeyboardButton("🎉 Выходные", callback_data=f"date_weekend_{category_name}")],
        [InlineKeyboardButton("◀️ Назад к категориям", callback_data="show_categories")],
    ]
    text = f"📌 **{display_name}**\n\nВыберите дату для поиска:"
    if isinstance(update_or_query, Update):
        await update_or_query.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")
    else:
        await update_or_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")


# ---------------------- Подписки ----------------------


async def send_subscription_prompt(query_or_update, category: str, date_type: str):
    display_name = CATEGORY_NAMES.get(category, category)
    dt_names = {"today": "на сегодня", "tomorrow": "на завтра", "upcoming": "на ближайшие дни", "weekend": "на выходные"}
    keyboard = [[InlineKeyboardButton("🔔 Подписаться", callback_data=f"sub_{category}_{date_type}")]]
    text = f"🔔 Подписаться на {display_name} {dt_names.get(date_type, '')}?"
    send = query_or_update.message.reply_text if isinstance(query_or_update, Update) else query_or_update.message.reply_text
    await send(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")


def _build_subs_keyboard(subs: list, flash_subs: list = None) -> tuple:
    """Строит текст и клавиатуру экрана подписок (галочки)."""
    dt_names = {"today": "сегодня", "tomorrow": "завтра", "upcoming": "ближайшие", "weekend": "выходные"}
    keyboard = []
    flash_subs = flash_subs or []

    if not subs and not flash_subs:
        text = (
            "🔔 <b>Мои подписки</b>\n\n"
            "У вас нет активных подписок.\n"
            "Подписаться можно через меню 🎯 Категории → выбрать категорию → выбрать период."
        )
        keyboard.append([InlineKeyboardButton("🎯 Перейти к категориям", callback_data="show_categories")])
        return text, InlineKeyboardMarkup(keyboard)

    lines = ["🔔 <b>Мои подписки</b>\n"]

    if subs:
        lines.append("Нажмите ✅ чтобы <b>отписаться</b>:\n")
        for s in subs:
            cat_name = CATEGORY_NAMES.get(s["category"], s["category"])
            dt_name = dt_names.get(s["date_type"], s["date_type"])
            keyboard.append([
                InlineKeyboardButton(
                    f"✅  {cat_name} / {dt_name}",
                    callback_data=f"unsub_{s['category']}_{s['date_type']}"
                )
            ])

    if flash_subs:
        lines.append("\n⚡ <b>Флеш-подписки</b> (нажмите чтобы отписаться):\n")
        for f in flash_subs:
            q = f["query"]
            keyboard.append([
                InlineKeyboardButton(
                    f"⚡ {q[:35]}",
                    callback_data=f"flash_unsub_{f['id']}"
                )
            ])

    keyboard.append([InlineKeyboardButton("➕ Добавить подписки", callback_data="show_categories")])
    return "\n".join(lines), InlineKeyboardMarkup(keyboard)


async def show_subscriptions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    subs = get_user_subscriptions(user_id)
    flash_subs = get_user_flash_subscriptions(user_id)
    text, keyboard = _build_subs_keyboard(subs, flash_subs)
    await update.message.reply_text(text, reply_markup=keyboard, parse_mode="HTML")


async def show_subscriptions_query(query, context: ContextTypes.DEFAULT_TYPE):
    """Версия для callback — обновляет то же сообщение."""
    user_id = query.from_user.id
    subs = get_user_subscriptions(user_id)
    flash_subs = get_user_flash_subscriptions(user_id)
    text, keyboard = _build_subs_keyboard(subs, flash_subs)
    try:
        await query.edit_message_text(text, reply_markup=keyboard, parse_mode="HTML")
    except Exception:
        await query.message.reply_text(text, reply_markup=keyboard, parse_mode="HTML")
async def send_subscriptions_digest(bot, date_type: str):
    """Рассылает дайджест подписчикам после обновления парсеров.
    Каждая категория — отдельное сообщение каждому подписчику.
    Если событий нет — не отправляем."""
    logger.info(f"📬 Рассылка дайджеста: {date_type}")
    subscribers = get_all_subscribers()
    today = datetime.now(MINSK_TZ)
    tomorrow = today + timedelta(days=1)
    sent_count, error_count = 0, 0

    categories_with_subs = {cat for (cat, dt) in subscribers.keys() if dt == date_type}

    for category in categories_with_subs:
        user_ids = subscribers.get((category, date_type), [])
        if not user_ids:
            continue

        if date_type == "today":
            events = get_events_by_date_and_category(today, category)
            period_label = f"сегодня ({today.strftime('%d.%m')})"
        elif date_type == "tomorrow":
            events = get_events_by_date_and_category(tomorrow, category)
            period_label = f"завтра ({tomorrow.strftime('%d.%m')})"
        elif date_type == "upcoming":
            events = get_upcoming_events(limit=20, category=category)
            period_label = "ближайшие дни"
        elif date_type == "weekend":
            events, saturday, sunday = get_weekend_events(category=category)
            period_label = f"выходные ({saturday.strftime('%d.%m')}–{sunday.strftime('%d.%m')})"
        else:
            continue

        # Не отправляем если нет событий
        if not events:
            logger.info(f"  ↩ {category}/{date_type}: нет событий, пропускаем")
            continue

        display_name = CATEGORY_NAMES.get(category, category)
        events_list = [dict(e) if not isinstance(e, dict) else e for e in events]

        # Группируем как в боте
        if category == "cinema":
            grouped_items = format_grouped_cinema_events(group_cinema_events(events_list[:10]))
            event_lines = []
            for text, url in grouped_items[:5]:
                link = f"\n🔗 <a href=\"{url}\">Подробнее</a>" if url else ""
                event_lines.append(text + link)
        else:
            grouped_items = group_other_events(events_list[:10])
            event_lines = []
            for item in grouped_items[:5]:
                link = f"\n🔗 <a href=\"{item['url']}\">Подробнее</a>" if item.get("url") else ""
                event_lines.append(item["text"] + link)

        lines = [
            "🔔 С добрым утром! Пора начинать новый 🌟 Dvizh!\n",
            f"🔔 <b>{display_name} на {period_label}</b> — {len(events)} событий\n",
        ] + event_lines

        if len(events) > 5:
            lines.append(f"\n<i>...и ещё {len(events) - 5} событий. Откройте бот для просмотра всех.</i>")

        message_text = "\n\n".join(lines)
        if len(message_text) > 4096:
            message_text = message_text[:4040] + "\n\n<i>...открой бот чтобы увидеть все.</i>"

        unsubscribe_keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("🔕 Отписаться", callback_data=f"unsub_{category}_{date_type}")
        ]])
        from telegram.error import RetryAfter
        for user_id in user_ids:
            try:
                await bot.send_message(chat_id=user_id, text=message_text,
                                       reply_markup=unsubscribe_keyboard,
                                       parse_mode="HTML", disable_web_page_preview=True)
                sent_count += 1
                await asyncio.sleep(0.1)
            except RetryAfter as e:
                logger.warning(f"Дайджест RetryAfter {e.retry_after}с для {user_id}")
                await asyncio.sleep(e.retry_after + 1)
                try:
                    await bot.send_message(chat_id=user_id, text=message_text,
                                           reply_markup=unsubscribe_keyboard,
                                           parse_mode="HTML", disable_web_page_preview=True)
                    sent_count += 1
                except Exception as retry_e:
                    error_count += 1
                    logger.warning(f"Повтор после RetryAfter не удался {user_id}: {retry_e}")
            except Exception as e:
                error_count += 1
                logger.warning(f"Не удалось отправить подписчику {user_id}: {e}")

    logger.info(f"📬 Рассылка завершена: отправлено {sent_count}, ошибок {error_count}")
    return sent_count, error_count


# ---------------------- Статистика ----------------------


async def show_pending_list(update_or_query, context: ContextTypes.DEFAULT_TYPE):
    """Список событий на модерации (status=pending/edited)."""
    if hasattr(update_or_query, 'from_user'):
        uid = update_or_query.from_user.id
    else:
        uid = update_or_query.effective_user.id
    if uid != ADMIN_ID:
        return

    with get_db_connection() as conn:
        rows = conn.execute("""
            SELECT id, title, event_date, place, category, status, first_name, username, created_at
            FROM pending_events WHERE status IN ('pending','edited')
            ORDER BY created_at ASC
        """).fetchall()

    if not rows:
        text = "✅ <b>Очередь модерации пуста</b>\n\nНет событий ожидающих проверки."
        await update_or_query.message.reply_text(text, parse_mode="HTML")
        return

    import html as _html
    lines = [f"📋 <b>На модерации: {len(rows)} событий</b>\n"]
    keyboard = []
    for r in rows:
        try:
            ed = r["event_date"] or ""
            d = datetime.strptime(ed.split("|")[0].strip(), "%Y-%m-%d").strftime("%d.%m.%Y")
            if "|" in ed:
                d += "–" + datetime.strptime(ed.split("|")[1].strip(), "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            d = r["event_date"] or "?"
        cat_emoji = CATEGORY_EMOJI.get(r["category"] or "", "📌")
        mark = "🆕" if r["status"] == "pending" else "✏️"
        uname = f"@{r['username']}" if r["username"] else (r["first_name"] or "?")
        place_str = f" • {_html.escape(r['place'])}" if r["place"] else ""
        lines.append(
            f"{mark} <b>#{r['id']}</b> {cat_emoji} {_html.escape(r['title'] or '—')}\n"
            f"   📅 {d}{place_str} | 👤 {_html.escape(uname)}"
        )
        keyboard.append([
            InlineKeyboardButton(f"✅#{r['id']}", callback_data=f"mod_approve_{r['id']}"),
            InlineKeyboardButton(f"❌#{r['id']}", callback_data=f"mod_reject_{r['id']}"),
            InlineKeyboardButton(f"✏️#{r['id']}", callback_data=f"mod_edit_{r['id']}"),
        ])
    keyboard.append([
        InlineKeyboardButton("✅ Принять все", callback_data="adm_approve_all"),
        InlineKeyboardButton("❌ Отклонить все", callback_data="adm_reject_all"),
    ])

    text = "\n".join(lines)[:4096]
    kbd = InlineKeyboardMarkup(keyboard)
    # Всегда reply_text — чтобы /admin панель оставалась видна
    await update_or_query.message.reply_text(text, reply_markup=kbd, parse_mode="HTML")


async def pending_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /pending — очередь модерации."""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Нет доступа.")
        return
    await show_pending_list(update, context)


async def admin_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """🔧 /admin — панель администратора."""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Нет доступа.")
        return
    with get_db_connection() as conn:
        _pcnt = conn.execute(
            "SELECT COUNT(*) FROM pending_events WHERE status IN ('pending','edited')"
        ).fetchone()[0]
    _plabel = f"📋 Модерация ({_pcnt})" if _pcnt else "📋 Модерация"
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton(_plabel, callback_data="adm_pending")],
        [InlineKeyboardButton("📊 Статистика", callback_data="adm_stats"),
         InlineKeyboardButton("📈 Кл. статистика", callback_data="adm_ustats")],
        [InlineKeyboardButton("🔄 Обновить парсеры", callback_data="adm_update"),
         InlineKeyboardButton("🗄 Скачать базу", callback_data="adm_download")],
        [InlineKeyboardButton("📢 Пост: сегодня", callback_data="adm_post_today"),
         InlineKeyboardButton("🎉 Пост: выходные", callback_data="adm_post_weekend")],
        [InlineKeyboardButton("🧹 Чистка дубликатов", callback_data="adm_dedup")],
        [InlineKeyboardButton("🗑 Удалить событие", callback_data="adm_del_prompt")],
    ])
    await update.message.reply_text(
        "🔧 <b>Панель администратора</b>\n\nВыберите действие:",
        reply_markup=keyboard,
        parse_mode="HTML"
    )


async def download_db(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Скачать БД — только для админа."""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Нет доступа.")
        return
    try:
        await update.message.reply_document(
            document=open(DB_NAME, "rb"),
            filename="events_final.db",
            caption=f"🗄 База данных\n📅 {datetime.now(MINSK_TZ).strftime('%d.%m.%Y %H:%M')}",
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {e}")


def _format_stats(stats: dict, title: str) -> str:
    import html as _html
    MN = {"01":"янв","02":"фев","03":"мар","04":"апр","05":"май","06":"июн",
          "07":"июл","08":"авг","09":"сен","10":"окт","11":"ноя","12":"дек"}

    total   = stats["total_users"]
    dau     = stats["dau"]
    wau     = stats["wau"]
    mau     = stats["mau"]
    wa_dau  = stats["webapp_dau"]
    wa_wau  = stats["webapp_wau"]
    wa_mau  = stats["webapp_mau"]
    wa_all  = stats["webapp_total"]

    acts_total  = stats["total_actions"]
    acts_today  = stats["actions_today"]
    apd = round(acts_today / dau, 1) if dau else 0   # actions per DAU

    lines = [f"<b>{title}</b>", ""]

    # ── Пользователи ──────────────────────────────────────────────
    lines += [
        "👤 <b>Пользователи</b>",
        f"  All-time: <b>{total}</b>   (за {stats['days_alive']} дн, +{stats['new_30d']} за 30 дн)",
        f"  DAU:  <b>{dau}</b>  (+{stats['new_today']} новых сегодня)",
        f"  WAU:  <b>{wau}</b>",
        f"  MAU:  <b>{mau}</b>",
        "",
        "🌐 <b>WebApp</b>",
        f"  All-time: <b>{wa_all}</b>",
        f"  DAU: <b>{wa_dau}</b>  WAU: <b>{wa_wau}</b>  MAU: <b>{wa_mau}</b>",
        "",
    ]

    # ── Контент ───────────────────────────────────────────────────
    lines += [
        "🗂 <b>База событий</b>",
        f"  Предстоящих: <b>{stats['events_count']}</b>",
    ]
    for row in stats.get("events_by_source", []):
        lines.append(f"  · {row['source_name']}: {row['cnt']}")
    lines.append("")

    # ── Подписки ──────────────────────────────────────────────────
    lines += [
        "🔔 <b>Подписки на категории</b>",
        f"  Подписчиков: <b>{stats['subscribers_count']}</b>  |  Всего подписок: <b>{stats['subscriptions_total']}</b>",
        "",
        "⚡ <b>Флеш-подписки</b>",
        f"  Активных: <b>{stats['flash_total']}</b>  (у {stats['flash_users']} польз.)",
        f"  Новых сегодня: <b>{stats['flash_new_today']}</b>  |  За 30 дн: <b>{stats['flash_new_30d']}</b>",
        f"  Польз. получили уведомление (30 дн): <b>{stats['flash_notified_users_30d']}</b>",
        "",
    ]

    # ── Активность за 14 дней ─────────────────────────────────────
    lines.append("📅 <b>Активность (14 дней):</b>")
    for row in list(stats["daily_activity"])[:14]:
        day   = row["day"]
        cnt   = row["cnt"]
        users = row["users"]
        new_u = row["new_users"] if "new_users" in row.keys() else 0
        wa_u  = stats["webapp_by_day"].get(day, 0)
        new_str = f"  +{new_u}↑" if new_u else ""
        wa_str  = f"  🌐{wa_u}" if wa_u else ""
        lines.append(f"  {day}  DAU <b>{users}</b>{wa_str}  {cnt} запр{new_str}")

    # ── По месяцам ────────────────────────────────────────────────
    if stats.get("monthly_activity"):
        lines += ["", "📆 <b>По месяцам:</b>"]
        for row in stats["monthly_activity"]:
            ym    = row["month"]
            cnt   = row["cnt"]
            users = row["users"]
            new_u = row["new_users"] if "new_users" in row.keys() else 0
            year, mon = ym.split("-")
            label   = f"{MN.get(mon, mon)} {year}"
            wa_u    = stats["webapp_by_month"].get(ym, 0)
            new_str = f"  +{new_u}↑" if new_u else ""
            wa_str  = f"  🌐{wa_u}" if wa_u else ""
            lines.append(f"  {label}  MAU <b>{users}</b>{wa_str}  {cnt} запр{new_str}")

    # ── Топ действий ─────────────────────────────────────────────
    lines += ["", "🔝 <b>Топ действий:</b>"]
    for row in stats["top_actions"]:
        lines.append(f"  {_html.escape(str(row['action']))} — {row['cnt']}")

    return "\n".join(lines)


async def show_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """📊 /stats — статистика со мной (для проверки)."""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Нет доступа.")
        return
    stats = get_stats_data(exclude_admin=False)
    await update.message.reply_text(_format_stats(stats, "📊 СТАТИСТИКА (все)"), parse_mode="HTML")


async def show_ustats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """📊 /ustats — статистика только пользователей (без меня)."""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Нет доступа.")
        return
    stats = get_stats_data(exclude_admin=True)
    await update.message.reply_text(_format_stats(stats, "📊 СТАТИСТИКА ПОЛЬЗОВАТЕЛЕЙ"), parse_mode="HTML")


# ---------------------- Чистка дубликатов ----------------------


def _find_duplicates(conn) -> list[dict]:
    """Возвращает группы дубликатов: (keep_id, delete_ids, title, event_date, place, cnt)."""
    cursor = conn.execute("""
        SELECT MIN(id) AS keep_id,
               GROUP_CONCAT(id) AS all_ids,
               title, event_date,
               COALESCE(place, '') AS place,
               COUNT(*) AS cnt
        FROM events
        GROUP BY LOWER(title), event_date,
                 LOWER(COALESCE(show_time, '')),
                 LOWER(COALESCE(place, ''))
        HAVING COUNT(*) > 1
        ORDER BY cnt DESC
    """)
    result = []
    for row in cursor.fetchall():
        all_ids = [int(x) for x in str(row[1]).split(',')]
        keep_id = int(row[0])
        delete_ids = [i for i in all_ids if i != keep_id]
        result.append({
            'keep_id': keep_id,
            'delete_ids': delete_ids,
            'title': row[2],
            'event_date': row[3],
            'place': row[4],
            'cnt': int(row[5]),
        })
    return result


async def delete_event_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/delete_event <id>            — удалить одно событие
    /delete_event <id_from> <id_to>  — удалить диапазон id (включительно)
    Только для администратора."""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Нет доступа.")
        return

    args = context.args
    if not args:
        await update.message.reply_text(
            "Использование:\n"
            "<code>/delete_event 123</code> — удалить событие #123\n"
            "<code>/delete_event 100 200</code> — удалить события с id 100 по 200",
            parse_mode="HTML"
        )
        return

    try:
        if len(args) == 1:
            id_from = id_to = int(args[0])
        else:
            id_from, id_to = int(args[0]), int(args[1])
            if id_from > id_to:
                id_from, id_to = id_to, id_from
    except ValueError:
        await update.message.reply_text("❌ ID должны быть целыми числами.")
        return

    if id_to - id_from > 1000:
        await update.message.reply_text("❌ Диапазон не должен превышать 1000 событий за раз.")
        return

    with get_db_connection() as conn:
        # Покажем что будет удалено
        rows = conn.execute(
            "SELECT id, title, event_date FROM events WHERE id BETWEEN ? AND ? ORDER BY id LIMIT 20",
            (id_from, id_to)
        ).fetchall()
        total = conn.execute(
            "SELECT COUNT(*) FROM events WHERE id BETWEEN ? AND ?",
            (id_from, id_to)
        ).fetchone()[0]

    if not total:
        await update.message.reply_text(f"ℹ️ События с id {id_from}–{id_to} не найдены.")
        return

    preview = "\n".join(
        f"  #{r['id']} {r['title'][:40]} ({r['event_date']})" for r in rows
    )
    if total > 20:
        preview += f"\n  ... и ещё {total - 20} событий"

    # Сохраняем диапазон для подтверждения
    context.user_data["del_range"] = (id_from, id_to)

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton(f"✅ Да, удалить {total} событий", callback_data=f"adm_del_confirm_{id_from}_{id_to}"),
        InlineKeyboardButton("❌ Отмена", callback_data="adm_del_cancel"),
    ]])
    await update.message.reply_text(
        f"🗑 <b>Будет удалено {total} событий (id {id_from}–{id_to}):</b>\n\n{preview}",
        reply_markup=keyboard,
        parse_mode="HTML"
    )


async def dedup_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/dedup — поиск и удаление дубликатов в БД (только для админа).
    /dedup         — показать статистику без удаления
    /dedup confirm — удалить дубликаты
    """
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Нет доступа.")
        return

    confirm = bool(context.args and context.args[0].lower() == 'confirm')

    with get_db_connection() as conn:
        conn.row_factory = sqlite3.Row
        groups = _find_duplicates(conn)

    total_groups = len(groups)
    total_delete = sum(len(g['delete_ids']) for g in groups)

    if total_groups == 0:
        await update.message.reply_text("✅ Дубликатов не найдено — база чистая.")
        return

    if not confirm:
        # Показываем превью первых 10 групп
        lines = [f"🔍 <b>Найдено дубликатов:</b> {total_groups} групп → {total_delete} лишних записей\n"]
        for g in groups[:10]:
            lines.append(
                f"• <b>{g['title'][:50]}</b>\n"
                f"  📅 {g['event_date']} · 📍 {g['place'][:30] or '—'}\n"
                f"  Копий: {g['cnt']} → оставим id={g['keep_id']}, удалим {len(g['delete_ids'])} шт."
            )
        if total_groups > 10:
            lines.append(f"\n<i>...и ещё {total_groups - 10} групп</i>")
        lines.append(f"\n⚠️ Для удаления: /dedup confirm")
        await update.message.reply_text('\n'.join(lines), parse_mode="HTML")
        return

    # Выполняем удаление
    all_delete_ids = [i for g in groups for i in g['delete_ids']]
    with get_db_connection() as conn:
        placeholders = ','.join('?' * len(all_delete_ids))
        conn.execute(f"DELETE FROM events WHERE id IN ({placeholders})", all_delete_ids)
        conn.commit()

    await update.message.reply_text(
        f"🧹 <b>Удалено {total_delete} дубликатов</b> из {total_groups} групп.\n"
        f"База очищена.",
        parse_mode="HTML"
    )


# ---------------------- Планировщик парсеров ----------------------


async def update_parsers(update_or_query, context: ContextTypes.DEFAULT_TYPE):
    """Ручной запуск парсеров (только для администратора)."""
    # Определяем, откуда пришел вызов
    if isinstance(update_or_query, Update):
        user_id = update_or_query.effective_user.id
        message = update_or_query.message
    else:
        user_id = update_or_query.from_user.id
        message = update_or_query.message
    
    if user_id != ADMIN_ID:
        if isinstance(update_or_query, Update):
            await update_or_query.message.reply_text("⛔ Нет доступа.")
        else:
            await update_or_query.answer("⛔ Нет доступа", show_alert=True)
        return
    
    # Отправляем сообщение о начале обновления
    await message.reply_text("🔄 **Обновление афиши...**\nЗапускаю парсеры, ~1-2 минуты.", parse_mode="Markdown")
    
    try:
        process = await asyncio.create_subprocess_exec(
            sys.executable, "run_all_parsers.py",
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        stdout, stderr = await asyncio.wait_for(process.communicate(), timeout=300)
        elapsed = (datetime.now(MINSK_TZ) - message.date.astimezone(MINSK_TZ)).total_seconds()
        
        if process.returncode == 0:
            output = stdout.decode("utf-8", errors="replace")
            report = _parse_parser_report(output)
            if report:
                text = _format_parser_report(report, elapsed)
            else:
                text = f"✅ Обновление завершено за {elapsed:.0f} сек\n\nℹ️ Детальный отчёт недоступен"
            await message.reply_text(text, parse_mode="Markdown")
            
        else:
            err = stderr.decode("utf-8", errors="replace").strip() if stderr else ""
            out = stdout.decode("utf-8", errors="replace").strip() if stdout else ""
            debug = err or out or "нет вывода"
            await message.reply_text(
                f"❌ **Ошибка парсеров** (код {process.returncode})\n\n```\n{debug[:800]}\n```",
                parse_mode="Markdown"
            )
            
    except asyncio.TimeoutError:
        await message.reply_text("⏰ Превышено время ожидания (5 мин).", parse_mode="Markdown")
    except Exception as e:
        await message.reply_text(f"💥 **Ошибка**: `{e}`", parse_mode="Markdown")


async def run_parsers_job(bot=None):
    """Запускает парсеры по расписанию, отправляет отчёт и рассылает дайджест."""
    logger.info("⏰ Запуск парсеров по расписанию...")
    start_time = datetime.now(MINSK_TZ)
    try:
        process = await asyncio.create_subprocess_exec(
            sys.executable, "run_all_parsers.py",
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        stdout, stderr = await asyncio.wait_for(process.communicate(), timeout=600)
        elapsed = (datetime.now(MINSK_TZ) - start_time).total_seconds()
        if process.returncode == 0:
            output = stdout.decode()
            logger.info(f"✅ Парсеры завершены за {elapsed:.0f} сек")
            if bot:
                report = _parse_parser_report(output)
                await _send_parser_report(bot, report or [], elapsed)
                # Проверяем флеш-подписки после обновления базы
                flash_sent = await check_flash_subscriptions(bot)
                if flash_sent:
                    logger.info(f"⚡ Флеш-подписки: отправлено {flash_sent} уведомлений")
        else:
            error_msg = stderr.decode()[:300] if stderr else "неизвестная ошибка"
            logger.error(f"❌ Парсеры упали: {error_msg}")
            if bot:
                await bot.send_message(chat_id=ADMIN_ID, text=f"❌ **Ошибка парсеров**\n\n```\n{error_msg}\n```", parse_mode="Markdown")
    except asyncio.TimeoutError:
        logger.error("⏰ Таймаут парсеров (10 мин)")
        if bot:
            await bot.send_message(chat_id=ADMIN_ID, text="⏰ **Таймаут** парсеров (>10 мин)", parse_mode="Markdown")
    except Exception as e:
        logger.error(f"💥 Ошибка: {e}")
        if bot:
            await bot.send_message(chat_id=ADMIN_ID, text=f"💥 **Критическая ошибка**: {e}", parse_mode="Markdown")


def _parse_parser_report(output: str) -> dict | None:
    """Извлекает PARSER_REPORT:json из stdout парсеров."""
    for line in output.split("\n"):
        line = line.strip()
        if line.startswith("PARSER_REPORT:"):
            try:
                return json.loads(line[len("PARSER_REPORT:"):])
            except Exception:
                pass
    return None


def _format_parser_report(report: dict, elapsed: float | None = None) -> str:
    """Форматирует отчёт парсеров в читаемый текст для Telegram.
    Relax-парсеры группируются под одной шапкой."""
    dur = elapsed or report.get("duration", 0)
    now = datetime.now(MINSK_TZ).strftime("%d.%m.%Y %H:%M")
    lines = [
        "🤖 *Отчёт парсеров*",
        f"🕐 {now}  ⏱ {dur:.0f} сек",
        "",
    ]

    # Каждый парсер — отдельный блок
    for p in report.get("parsers", []):
        status = "✅" if p["ok"] else "❌"
        lines.append(f"{status} *{p['name']}*")
        if p["ok"]:
            for r in p.get("results", []):
                parts = r.split(":")
                if len(parts) == 4:
                    label, found, saved = parts[1], parts[2], parts[3]
                    lines.append(f"   └ {label}: найдено {found}, добавлено {saved}")
        else:
            lines.append("   └ завершился с ошибкой")

    # Итог
    total_found = sum(
        int(r.split(":")[2])
        for p in report.get("parsers", [])
        for r in p.get("results", [])
        if len(r.split(":")) == 4
    )
    s = report.get("success", 0)
    f = report.get("failed", 0)
    lines += ["", f"📦 Всего найдено: {total_found}", f"📊 Итого: ✅ {s} успешно  ❌ {f} с ошибкой"]
    return "\n".join(lines)


async def _send_parser_report(bot, report_or_results, elapsed: float):
    """Отправляет отчёт админу. Принимает dict (новый формат) или list (старый)."""
    if isinstance(report_or_results, dict):
        text = _format_parser_report(report_or_results, elapsed)
    else:
        # fallback: старый формат — просто список строк
        lines = ["🤖 *Отчёт парсеров*",
                 f"🕐 {datetime.now(MINSK_TZ).strftime('%d.%m.%Y %H:%M')} | ⏱ {elapsed:.0f} сек", ""]
        lines.extend(report_or_results or ["ℹ️ Нет данных"])
        text = "\n".join(lines)
    try:
        await bot.send_message(chat_id=ADMIN_ID, text=text, parse_mode="Markdown")
    except Exception as e:
        logger.error(f"Не удалось отправить отчёт: {e}")


async def send_digest_job(bot=None):
    """Рассылает дайджест подписчикам в 8:00 (после обновления парсеров в 6:00)."""
    if bot:
        for date_type in ("today", "tomorrow", "upcoming", "weekend"):
            sent, errors = await send_subscriptions_digest(bot, date_type)
            logger.info(f"📬 Дайджест [{date_type}]: отправлено {sent} польз., {errors} ошибок")


# ---------------------- Добавление событий (модерация) ----------------------


# Метаданные полей формы — используется везде (добавление, редактирование, модерация)
FIELD_LABELS = {
    "title":       ("📝", "Название",   True),   # обязательное
    "details":     ("📖", "Формат",     True),   # обязательное: краткий формат события
    "category":    ("🎯", "Категория",  True),   # обязательное
    "event_date":  ("📅", "Дата",       True),   # обязательное
    "show_time":   ("⏰", "Время",      True),   # обязательное (формат ЧЧ:ММ или ЧЧ:ММ-ЧЧ:ММ)
    "place":       ("🏢", "Место",      True),   # обязательное
    "address":     ("📍", "Адрес",      False),
    "price":       ("💰", "Цена",       False),
    "description": ("📋", "Описание",   False),  # необязательное: подробное описание
    "source_url":  ("🔗", "Ссылка",     False),
}

FIELD_PROMPTS = {
    "title":       "📝 Введите <b>название</b> события:",
    "event_date":  ("📅 Введите <b>дату</b> в формате ДД.ММ.ГГГГ\n"
                    "Или <b>период</b>: ДД.ММ.ГГГГ-ДД.ММ.ГГГГ (например: 15.04.2026-20.04.2026)\n"
                    "Тогда событие появится на каждый день периода"),
    "show_time":   "⏰ Введите <b>время</b> в формате ЧЧ:ММ или диапазон ЧЧ:ММ-ЧЧ:ММ (например: 10:00-18:00):",
    "place":       "🏢 Введите <b>место проведения</b> (название площадки):",
    "address":     "📍 Введите <b>адрес</b> (например: ул. Притыцкого, 62)\nИли /skip чтобы пропустить:",
    "category":    "🎯 Выберите <b>категорию</b>:",
    "price":       "💰 Введите <b>цену</b> (например: от 20 BYN, Бесплатно)\nИли /skip чтобы пропустить:",
    "details":     "📖 Введите <b>формат события</b> — коротко что это и для кого (до 300 символов):",
    "description": "📋 Введите <b>подробное описание</b> (программа, спикеры и т.д.)\nИли /skip чтобы пропустить:",
    "source_url":  "🔗 Введите <b>ссылку</b> на событие\nИли /skip чтобы пропустить:",
}

def get_prompt(field: str, extra: str = "") -> str:
    """Возвращает строку промпта для поля (корректно обрабатывает tuple)."""
    p = FIELD_PROMPTS.get(field, "Введите значение:")
    if isinstance(p, tuple):
        p = "".join(p)
    if extra:
        p = p + "\n" + extra
    return p


CATEGORY_KEYBOARD = InlineKeyboardMarkup([
    [InlineKeyboardButton("🎬 Кино", callback_data="sc_cinema"),
     InlineKeyboardButton("🎵 Концерт", callback_data="sc_concert")],
    [InlineKeyboardButton("🎭 Театр", callback_data="sc_theater"),
     InlineKeyboardButton("🖼️ Выставка", callback_data="sc_exhibition")],
    [InlineKeyboardButton("🧸 Детям", callback_data="sc_kids"),
     InlineKeyboardButton("⚽ Спорт", callback_data="sc_sport")],
    [InlineKeyboardButton("🌟 Движ", callback_data="sc_party"),
     InlineKeyboardButton("🆓 Бесплатно", callback_data="sc_free")],
    [InlineKeyboardButton("🗺️ Экскурсия", callback_data="sc_excursion"),
     InlineKeyboardButton("🛍️ Маркет", callback_data="sc_market")],
    [InlineKeyboardButton("🎨 Мастер-класс", callback_data="sc_masterclass"),
     InlineKeyboardButton("🎲 Настолки", callback_data="sc_boardgames")],
    [InlineKeyboardButton("📺 Трансляция", callback_data="sc_broadcast"),
     InlineKeyboardButton("📚 Обучение", callback_data="sc_education")],
    [InlineKeyboardButton("❓ Квизы", callback_data="sc_quiz"),
     InlineKeyboardButton("📌 Другое", callback_data="sc_other")],
])



def get_pending_event(pending_id: int) -> dict | None:
    with get_db_connection() as conn:
        row = conn.execute("SELECT * FROM pending_events WHERE id=?", (pending_id,)).fetchone()
        return dict(row) if row else None


def update_pending_event(pending_id: int, data: dict):
    with get_db_connection() as conn:
        # Парсим show_time при сохранении модератором
        raw_t = data.get("show_time", "") or ""
        if "-" in raw_t and raw_t.count(":") == 2:
            _tp = raw_t.split("-", 1)
            _st, _et = _tp[0].strip(), _tp[1].strip()
        else:
            _st, _et = raw_t, data.get("end_time", "") or ""
        conn.execute("""
            UPDATE pending_events
            SET title=?, event_date=?, show_time=?, end_time=?, place=?, address=?,
                category=?, price=?, details=?, description=?, source_url=?, status='edited'
            WHERE id=?
        """, (
            data.get("title", ""), data.get("event_date", ""),
            _st, _et,
            data.get("place", ""),
            data.get("address", ""), data.get("category", "other"),
            data.get("price", ""), data.get("details", ""),
            data.get("description", ""),
            data.get("source_url", ""), pending_id,
        ))
        conn.commit()


def build_fields_keyboard(data: dict, mode: str = "submit") -> InlineKeyboardMarkup:
    """Строит меню полей с галочками у заполненных. mode=submit|mod_edit"""
    rows = []
    row = []
    for field, (emoji, label, required) in FIELD_LABELS.items():
        val = data.get(field, "")
        if required:
            mark = " ✅" if val else " ❗"
        else:
            mark = " ✅" if val else ""
        btn = InlineKeyboardButton(f"{emoji} {label}{mark}",
                                   callback_data=f"{mode}_field_{field}")
        row.append(btn)
        if len(row) == 2:
            rows.append(row); row = []
    if row:
        rows.append(row)
    if mode == "submit":
        # Кнопка промо-публикации (тоггл)
        is_promo = data.get("is_promo", False)
        promo_label = "📣 Промо: ВКЛ ✅" if is_promo else "📣 Промо: ВЫКЛ"
        rows.append([InlineKeyboardButton(promo_label, callback_data="submit_toggle_promo")])
        rows.append([InlineKeyboardButton("👁 Предпросмотр", callback_data="submit_preview")])
        rows.append([
            InlineKeyboardButton("✅ Отправить на модерацию", callback_data="submit_confirm"),
            InlineKeyboardButton("❌ Отмена", callback_data="submit_cancel"),
        ])
    else:  # mod_edit
        pending_id = data.get("_pending_id", "")
        rows.append([InlineKeyboardButton("👁 Предпросмотр", callback_data=f"mod_preview_{pending_id}")])
        rows.append([
            InlineKeyboardButton("📤 Отправить пользователю", callback_data=f"mod_send_edit_{pending_id}"),
            InlineKeyboardButton("❌ Отмена", callback_data=f"mod_edit_cancel_{pending_id}"),
        ])
    return InlineKeyboardMarkup(rows)


def format_promo_post(data: dict) -> str:
    """Форматирует промо-анонс события для публикации в канал."""
    import html as _html
    from normalizer import normalize_price

    cat_emoji = CATEGORY_EMOJI.get(data.get("category", "other"), "🎉")
    title = _html.escape(data.get("title", ""))

    # Дата
    ed = data.get("event_date", "")
    if "|" in str(ed):
        parts = ed.split("|", 1)
        try:
            d1 = datetime.strptime(parts[0].strip(), "%Y-%m-%d").strftime("%d.%m.%Y")
            d2 = datetime.strptime(parts[1].strip(), "%Y-%m-%d").strftime("%d.%m.%Y")
            date_str = f"{d1} — {d2}"
        except Exception:
            date_str = ed
    else:
        try:
            date_str = datetime.strptime(ed, "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            date_str = ed

    # Нормализуем цену (на всякий случай, если вдруг пришло сырое)
    price_raw = data.get("price", "")
    price_normalized = normalize_price(price_raw) if price_raw else ""

    lines = [
        f"🔥 <b>Новое событие в Минске!</b>",
        "",
        f"{cat_emoji} <b>{title}</b>",
    ]
    if data.get("details"):
        lines.append(f"📖 {_html.escape(data['details'][:200])}")
    lines.append(f"📅 {date_str}" + (f" ⏰ {data['show_time']}" if data.get("show_time") else ""))
    if data.get("place"):
        lines.append(f"🏢 {_html.escape(data['place'])}")
    if data.get("address"):
        lines.append(f"📍 {_html.escape(data['address'])}")
    if data.get("price"):
        lines.append(f"💰 {_html.escape(data['price'])}")
    if data.get("description"):
        lines.append(f"\n{_html.escape(data['description'][:500])}")
    if data.get("source_url"):
        lines.append(f"\n🔗 <a href=\"{_html.escape(data['source_url'])}\">Подробнее</a>")
    lines.append("\n👉 @Minskdvizh_bot | #афишаминск")
    return "\n".join(lines)


async def send_promo_to_subscribers(bot, row_data: dict) -> int:
    """Рассылает промо-анонс подписчикам категории события.
    Возвращает количество успешно отправленных сообщений."""
    from normalizer import normalize_price

    category = row_data.get("category", "")
    if not category:
        return 0

    # Берём подписчиков всех date_type этой категории — дедупликация по user_id
    subscribers = get_all_subscribers()
    user_ids = set()
    for (cat, _dt), ids in subscribers.items():
        if cat == category:
            user_ids.update(ids)

    if not user_ids:
        logger.info(f"Промо: нет подписчиков категории {category}")
        return 0

    # Копируем данные и нормализуем цену
    promo_data = dict(row_data)
    if promo_data.get("price"):
        promo_data["price"] = normalize_price(promo_data["price"])

    text = format_promo_post(promo_data)
    sent = 0
    for user_id in user_ids:
        try:
            await bot.send_message(
                chat_id=user_id,
                text=text,
                parse_mode="HTML",
                disable_web_page_preview=False,
            )
            sent += 1
            await asyncio.sleep(0.1)  # вежливая пауза
        except Exception as e:
            logger.warning(f"Промо подписчику {user_id}: {e}")

    logger.info(f"📣 Промо разослано {sent} подписчикам категории {category}")
    return sent


def validate_field(field: str, text: str):
    """Валидирует значение поля. Возвращает (ok, value_or_error).
    Для event_date поддерживает период ДД.ММ.ГГГГ-ДД.ММ.ГГГГ.
    Период кодируется как "YYYY-MM-DD|YYYY-MM-DD" и разворачивается при approve.
    """
    import re as _re
    from datetime import date as _date, timedelta as _td
    if field == "event_date":
        # Уже сохранённый период в ISO: "2026-04-15|2026-04-20" — принимаем как есть
        if _re.match(r"^\d{4}-\d{2}-\d{2}\|\d{4}-\d{2}-\d{2}$", text.strip()):
            return True, text.strip()
        # Уже сохранённая одиночная дата ISO: "2026-04-15" — принимаем как есть
        if _re.match(r"^\d{4}-\d{2}-\d{2}$", text.strip()):
            return True, text.strip()
        # Период: ДД.ММ.ГГГГ-ДД.ММ.ГГГГ
        pm = _re.match(r"^(\d{1,2}\.\d{1,2}\.\d{4})-(\d{1,2}\.\d{1,2}\.\d{4})$", text.strip())
        if pm:
            def _pd(s):
                d, mo, y = s.split(".")
                return _date(int(y), int(mo), int(d))
            try:
                d_from = _pd(pm.group(1))
                d_to   = _pd(pm.group(2))
            except ValueError:
                return False, "❌ Одна из дат периода не существует."
            today_d = _date.today()
            if d_from < today_d:
                return False, f"❌ Дата начала ({pm.group(1)}) уже в прошлом.\nСегодня {today_d.strftime('%d.%m.%Y')}"
            if d_to < today_d:
                return False, f"❌ Дата конца ({pm.group(2)}) уже в прошлом.\nСегодня {today_d.strftime('%d.%m.%Y')}"
            if d_to < d_from:
                return False, f"❌ Дата конца ({pm.group(2)}) раньше даты начала ({pm.group(1)})."
            if d_to == d_from:
                return False, "❌ Для одного дня введите дату без дефиса (ДД.ММ.ГГГГ)."
            if (d_to - d_from).days > 90:
                return False, "❌ Период не может быть больше 90 дней."
            return True, f"{d_from.strftime('%Y-%m-%d')}|{d_to.strftime('%Y-%m-%d')}"
        # Одна дата
        m = _re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", text.strip())
        if not m:
            return False, "❌ Формат: ДД.ММ.ГГГГ или ДД.ММ.ГГГГ-ДД.ММ.ГГГГ"
        day, month, year = m.groups()
        try:
            ev = _date(int(year), int(month), int(day))
        except ValueError:
            return False, "❌ Такой даты не существует."
        today_d = _date.today()
        if ev < today_d:
            return False, f"❌ Дата {ev.strftime('%d.%m.%Y')} уже в прошлом. Сегодня {today_d.strftime('%d.%m.%Y')}"
        return True, f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    elif field == "show_time":
        # Диапазон ЧЧ:ММ-ЧЧ:ММ
        rng = _re.match(r"^(\d{1,2}:\d{2})-(\d{1,2}:\d{2})$", text.strip())
        if rng:
            t1, t2 = rng.group(1), rng.group(2)
            for t in (t1, t2):
                if not _re.match(r"^\d{1,2}:\d{2}$", t):
                    return False, "❌ Формат времени: ЧЧ:ММ-ЧЧ:ММ (например: 10:00-18:00)"
            return True, text.strip()  # сохраняем как "10:00-18:00" в show_time
        # Одиночное время
        m = _re.match(r"^(\d{1,2}):(\d{2})$", text.strip())
        if not m:
            return False, "❌ Формат: ЧЧ:ММ или ЧЧ:ММ-ЧЧ:ММ (например: 19:00 или 10:00-18:00)"
        return True, text.strip()
    elif field == "title" and len(text) < 3:
        return False, "❌ Название слишком короткое."
    elif field == "details":
        return True, text[:300]
    elif field == "description":
        return True, text[:1000]
    else:
        return True, text

def save_pending_event(user_id, username, first_name, data: dict) -> int:
    with get_db_connection() as conn:
        cursor = conn.cursor()
        # Парсим show_time: "10:00-18:00" → show_time="10:00", end_time="18:00"
        raw_time = data.get("show_time", "") or ""
        if "-" in raw_time and raw_time.count(":") == 2:
            _t_parts = raw_time.split("-", 1)
            _show_t, _end_t = _t_parts[0].strip(), _t_parts[1].strip()
        else:
            _show_t, _end_t = raw_time, ""
        cursor.execute("""
            INSERT INTO pending_events
              (user_id, username, first_name, title, event_date, show_time, end_time,
               place, address, category, details, description, price, source_url, is_promo, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            user_id, username, first_name,
            data.get("title", ""),
            data.get("event_date", ""),
            _show_t,
            _end_t,
            data.get("place", ""),
            data.get("address", ""),
            data.get("category", "other"),
            data.get("details", ""),
            data.get("description", ""),
            data.get("price", ""),
            data.get("source_url", ""),
            1 if data.get("is_promo") else 0,
            datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S"),
        ))
        conn.commit()
        return cursor.lastrowid


def check_duplicate_event(title: str, event_date: str, place: str) -> dict | None:
    """Проверяет дубликат в таблицах events и pending_events.

    Стратегия (3 уровня):
      1. Точное совпадение: title + event_date + place (нормализованные)
      2. Мягкое: title + event_date (без учёта места — одно событие на разных площадках редкость)
      3. Только дата + место (ловит переименованные события)

    Возвращает dict с описанием дубликата или None.
    """
    import re as _re
    from normalizer import normalize_title

    def _norm_title(s: str) -> str:
        return normalize_title(s)

    def _norm_place(s: str) -> str:
        # Простая очистка для сравнения (без канонизации)
        return _re.sub(r"[\s\-—–,\.!?]+", " ", (s or "").lower()).strip()

    t_norm  = _norm_title(title)
    p_norm  = _norm_place(place)

    # Для периода берём начальную дату
    if "|" in (event_date or ""):
        dates = [event_date.split("|")[0].strip()]
    else:
        dates = [event_date] if event_date else []

    if not dates or not t_norm:
        return None

    with get_db_connection() as conn:
        for chk_date in dates:
            # ── Уровень 1: title + date + place ──────────────────
            if p_norm:
                rows = conn.execute(
                    "SELECT id, title, event_date, place FROM events "
                    "WHERE event_date = ? AND LOWER(title) LIKE ? AND LOWER(COALESCE(place,'')) LIKE ?",
                    (chk_date, f"%{t_norm[:20]}%", f"%{p_norm[:20]}%")
                ).fetchall()
                for r in rows:
                    if _norm_title(r["title"]) == t_norm and _norm_place(r["place"] or "") == p_norm:
                        return {"level": 1, "source": "events", "id": r["id"],
                                "title": r["title"], "date": r["event_date"], "place": r["place"] or ""}

            # ── Уровень 2: title + date (без места) ──────────────
            rows = conn.execute(
                "SELECT id, title, event_date, place FROM events "
                "WHERE event_date = ? AND LOWER(title) LIKE ?",
                (chk_date, f"%{t_norm[:20]}%")
            ).fetchall()
            for r in rows:
                if _norm_title(r["title"]) == t_norm:
                    return {"level": 2, "source": "events", "id": r["id"],
                            "title": r["title"], "date": r["event_date"], "place": r["place"] or ""}

            # ── Проверяем pending_events (pending/edited, не rejected) ──
            rows_p = conn.execute(
                "SELECT id, title, event_date, place FROM pending_events "
                "WHERE event_date LIKE ? AND status NOT IN ('rejected','approved') AND LOWER(title) LIKE ?",
                (f"%{chk_date}%", f"%{t_norm[:20]}%")
            ).fetchall()
            for r in rows_p:
                if _norm_title(r["title"]) == t_norm:
                    return {"level": 2, "source": "pending", "id": r["id"],
                            "title": r["title"], "date": r["event_date"], "place": r["place"] or ""}

    return None


def _fmt_duplicate_reason(dup: dict) -> str:
    """Формирует текст причины отказа."""
    import html as _html
    src = "афише" if dup["source"] == "events" else "очереди на модерацию"
    try:
        from datetime import datetime as _dt
        d = _dt.strptime(dup["date"], "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        d = dup["date"]
    place_str = f" в «{_html.escape(dup['place'])}»" if dup["place"] else ""
    lvl_hint = {
        1: "точное совпадение названия, даты и места",
        2: "совпадение названия и даты",
        3: "совпадение даты и места",
    }.get(dup["level"], "")
    return (
        f"⚠️ Событие уже есть в {src}:\n"
        f"📌 <b>{_html.escape(dup['title'])}</b>\n"
        f"📅 {d}{place_str}\n"
        f"<i>({lvl_hint})</i>"
    )


def approve_pending_event(pending_id: int) -> tuple[bool, dict | None]:
    """Одобряет событие. Возвращает (success, row_data) — row_data нужен для промо."""
    from datetime import date as _date, timedelta as _td
    from normalizer import normalize_place, normalize_price
    
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM pending_events WHERE id = ?", (pending_id,))
        row = cursor.fetchone()
        if not row:
            return False, None

        _addr = (row["address"] or "") if "address" in row.keys() else ""
        try:
            _details = row["details"] or ""
        except Exception:
            _details = ""
        if not _details:
            _details = row["description"] or ""
        try:
            _description = row["description"] or ""
        except Exception:
            _description = ""
        # НОРМАЛИЗУЕМ МЕСТО И ЦЕНУ
        place_normalized = normalize_place(row["place"] or "")
        price_normalized = normalize_price(row["price"] or "")
            
        event_date_raw = row["event_date"] or ""
        if "|" in event_date_raw:
            parts = event_date_raw.split("|", 1)
            try:
                d_from = _date.fromisoformat(parts[0].strip())
                d_to   = _date.fromisoformat(parts[1].strip())
                dates = [d_from + _td(days=i) for i in range((d_to - d_from).days + 1)]
            except Exception:
                dates = [_date.today()]
        else:
            try:
                dates = [_date.fromisoformat(event_date_raw)]
            except Exception:
                dates = []

        for d in dates:
            try:
                _end_time = row["end_time"] or ""
            except Exception:
                _end_time = ""
            cursor.execute("""
                INSERT INTO events
                  (title, details, description, event_date, show_time, end_time,
                   place, location, price, category, source_name, source_url)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                row["title"],
                _details,
                _description,
                d.strftime("%Y-%m-%d"),
                row["show_time"] or "",
                _end_time,
                place_normalized,  #нормализованное место
                _addr or "Минск",
                price_normalized,
                row["category"] or "other",
                "user_submitted",
                row["source_url"] or "",
            ))

        cursor.execute("UPDATE pending_events SET status = 'approved' WHERE id = ?", (pending_id,))
        conn.commit()

        # Возвращаем данные для промо-публикации
        row_data = dict(row)
        row_data["place"] = place_normalized
        row_data["price"] = price_normalized
        return True, row_data


def reject_pending_event(pending_id: int):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE pending_events SET status = 'rejected' WHERE id = ?", (pending_id,))
        conn.commit()


def format_pending_preview(data: dict, user=None) -> str:
    import html as _html
    cat_emoji = CATEGORY_EMOJI.get(data.get("category", "other"), "📌")
    lines = ["📋 <b>Новое событие на модерацию</b>"]
    if user:
        uname = _html.escape(user.username or "нет")
        fname = _html.escape(user.first_name or "")
        lines.append(f"👤 От: {fname} (@{uname}, ID: {user.id})")
    lines.append("")
    lines.append(f"{cat_emoji} <b>{_html.escape(data.get('title', '—'))}</b>")
    if data.get("event_date"):
        ed = data["event_date"]
        if "|" in str(ed):
            parts = ed.split("|", 1)
            try:
                d1 = datetime.strptime(parts[0].strip(), "%Y-%m-%d").strftime("%d.%m.%Y")
                d2 = datetime.strptime(parts[1].strip(), "%Y-%m-%d").strftime("%d.%m.%Y")
                d = f"{d1} — {d2}"
            except Exception:
                d = ed
        else:
            try:
                d = datetime.strptime(ed, "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                d = ed
        line = f"📅 {d}"
        if data.get("show_time"):
            line += f" ⏰ {data['show_time']}"
        lines.append(line)
    if data.get("place"):
        lines.append(f"🏢 {_html.escape(data['place'])}")
    if data.get("price"):
        lines.append(f"💰 {_html.escape(data['price'])}")
    if data.get("details"):
        lines.append(f"📖 <b>Формат:</b> {_html.escape(data['details'][:300])}")
    if data.get("description"):
        lines.append(f"📋 <b>Описание:</b> {_html.escape(data['description'][:300])}")
    if data.get("address"):
        lines.append(f"📍 {_html.escape(data['address'])}")
    if data.get("source_url"):
        lines.append(f"🔗 {_html.escape(data['source_url'])}")
    return "\n".join(lines)


async def start_submit(update_or_query, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["submit"] = {}
    context.user_data["in_submit"] = True
    context.user_data.pop("submit_field", None)
    text = (
        "➕ <b>Добавление события</b>\n\n"
        "Выберите поле для заполнения.\n"
        "❗ — обязательное поле, ✅ — заполнено.\n"
        "Когда всё готово — нажмите <b>Отправить на модерацию</b>.\n"
        "Для отмены введите /cancel\n\n"
        "📦 Чтобы загрузить <b>несколько событий сразу</b> — получите шаблон: /template"
    )
    kbd = build_fields_keyboard({}, mode="submit")
    if isinstance(update_or_query, Update):
        await update_or_query.message.reply_text(text, reply_markup=kbd, parse_mode="HTML")
    else:
        await update_or_query.answer()
        await update_or_query.message.reply_text(text, reply_markup=kbd, parse_mode="HTML")


async def handle_submit_step(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    # Редактирование поля модератором
    if context.user_data.get("mod_edit_id") and update.effective_user.id == ADMIN_ID:
        field = context.user_data.get("mod_edit_field")
        if not field:
            return False
        text = update.message.text.strip()
        edit_data = context.user_data.get("mod_edit_data", {})
        pending_id = context.user_data["mod_edit_id"]
        if text == "/skip":
            edit_data[field] = ""
        elif text == "/cancel":
            context.user_data.pop("mod_edit_field", None)
            await update.message.reply_text("↩️ Изменение поля отменено.")
            edit_data["_pending_id"] = pending_id
            await update.message.reply_text(
                "✏️ Продолжайте редактирование или отправьте пользователю:",
                reply_markup=build_fields_keyboard(edit_data, mode="mod_edit")
            )
            return True
        else:
            ok, val = validate_field(field, text)
            if not ok:
                await update.message.reply_text(val)
                return True
            edit_data[field] = val
        context.user_data.pop("mod_edit_field", None)
        edit_data["_pending_id"] = pending_id
        await update.message.reply_text(
            f"✅ <b>{FIELD_LABELS[field][1]}</b> обновлено.\n\nПродолжайте редактирование или отправьте пользователю:",
            reply_markup=build_fields_keyboard(edit_data, mode="mod_edit"),
            parse_mode="HTML"
        )
        return True

    # Ввод поля пользователем
    if not context.user_data.get("in_submit"):
        return False

    field = context.user_data.get("submit_field")
    if not field:
        return False

    text = update.message.text.strip()
    data = context.user_data["submit"]

    if text == "/cancel":
        context.user_data.pop("in_submit", None)
        context.user_data.pop("submit", None)
        context.user_data.pop("submit_field", None)
        await update.message.reply_text("❌ Добавление отменено.", reply_markup=get_reply_main_menu())
        return True

    if text == "/skip":
        data[field] = ""
    else:
        ok, val = validate_field(field, text)
        if not ok:
            await update.message.reply_text(val)
            return True
        data[field] = val

    context.user_data.pop("submit_field", None)
    await update.message.reply_text(
        f"✅ <b>{FIELD_LABELS[field][1]}</b> сохранено.\n\nВыберите следующее поле или отправьте на модерацию:",
        reply_markup=build_fields_keyboard(data, mode="submit"),
        parse_mode="HTML"
    )
    return True


# _send_next_submit_prompt удалён — заменён на build_fields_keyboard

# ---------------------- Публикация в канал ----------------------

async def post_to_channel(bot, post_type: str = "today"):
    """Публикует подборку событий в Telegram канал."""
    if not CHANNEL_ID:
        logger.warning("CHANNEL_ID не задан — пропускаем публикацию в канал")
        return

    now = datetime.now(MINSK_TZ)
    DAY_NAMES = ["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье"]
    MONTH_NAMES = ["января","февраля","марта","апреля","мая","июня",
                   "июля","августа","сентября","октября","ноября","декабря"]
    # Эмодзи для категорий в постах (заглавные)
    CAT_POST_EMOJI = {
        "cinema": "🎬", "concert": "🎵", "theater": "🎭", "exhibition": "🖼",
        "kids": "🧸", "sport": "⚽", "party": "🌟", "free": "🆓",
        "excursion": "🗺", "market": "🛍", "masterclass": "🎨",
        "boardgames": "🎲", "broadcast": "📺", "education": "📚",
        "quiz": "❓",
    }
    CAT_POST_NAMES = {
        "cinema": "КИНО", "concert": "КОНЦЕРТЫ", "theater": "ТЕАТР",
        "exhibition": "ВЫСТАВКИ", "kids": "ДЕТЯМ", "sport": "СПОРТ",
        "party": "ДВИЖ", "free": "БЕСПЛАТНО", "excursion": "ЭКСКУРСИИ",
        "market": "МАРКЕТЫ", "masterclass": "МАСТЕР-КЛАССЫ",
        "boardgames": "НАСТОЛКИ", "broadcast": "ТРАНСЛЯЦИИ", "education": "ОБУЧЕНИЕ",
        "quiz": "КВИЗЫ",
    }

    def _fmt_price(price: str) -> str:
        if not price:
            return ""
        p = price.strip()
        if p.lower() in ("бесплатно", "free", "0", "0 byn", "0byn"):
            return "🆓"
        return p

    def _fmt_event_line(e) -> str:
        title = e.get("title") or ""
        price = _fmt_price(e.get("price", "") or "")
        url = e.get("source_url") or ""
        title_part = f"<a href=\"{url}\">{title}</a>" if url else title
        return f"→ {title_part}" + (f" | {price}" if price else "")

    if post_type == "today":
        events_raw = get_events_by_date_and_category(now)
        events = [dict(e) for e in events_raw] if events_raw else []
        if not events:
            return
        day_name = DAY_NAMES[now.weekday()].lower()
        day_num = now.day
        month_name = MONTH_NAMES[now.month - 1]
        lines = [
            f"👹 Сегодня {day_num} {month_name} ({day_name}). #афишаминск #дайджест",
            f"😎 Всем доброго утра и продуктивного дня!",
            f"✨ Куда пойти сегодня в Минске?\n",
            f"Планируй когда удобно — всё открыто для тебя.\n",
        ]
        from collections import defaultdict as _dd
        by_cat = _dd(list)
        # Дедупликация между категориями — глобальный set по (title, place)
        seen_global: set = set()
        for e in events[:60]:
            key = (e.get("title", ""), e.get("place", ""))
            if key in seen_global:
                continue
            seen_global.add(key)
            by_cat[e.get("category")].append(e)
        for cat, evs in by_cat.items():
            emoji = CAT_POST_EMOJI.get(cat, "📌")
            cat_name = CAT_POST_NAMES.get(cat, cat.upper())
            # Дедупликация внутри категории по title
            seen_cat: set = set()
            unique_evs = []
            for e in evs:
                t = e.get("title", "")
                if t not in seen_cat:
                    seen_cat.add(t)
                    unique_evs.append(e)
            lines.append(f"\n{emoji} {cat_name}")
            for e in unique_evs[:5]:
                lines.append(_fmt_event_line(e))
        lines.append(f"\nА вечером тебя ждет ДВИЖ 🌟")
        lines.append(f"\n👉 Ищи все события: @Minskdvizh_bot")
        lines.append("#афишаминск #мероприятияминск #концертыминск #выставкиминск #движ")

    elif post_type == "weekend":
        saturday = now + timedelta(days=(5 - now.weekday()) % 7 or 7)
        sunday = saturday + timedelta(days=1)
        events_sat_raw = get_events_by_date_and_category(saturday)
        events_sun_raw = get_events_by_date_and_category(sunday)
        events_sat = [dict(e) for e in events_sat_raw] if events_sat_raw else []
        events_sun = [dict(e) for e in events_sun_raw] if events_sun_raw else []        
        all_events = list(events_sat)[:15] + list(events_sun)[:15]
        if not all_events:
            return
        sat_d = saturday.day
        sun_d = sunday.day
        mon = MONTH_NAMES[saturday.month - 1]
        lines = [
            f"🎉 Выходные {sat_d}-{sun_d} {mon} (сб-вск). #минск #выходные",
            f"😎 Планируем яркие выходные в Минске!\n",
        ]
        from collections import defaultdict as _dd
        by_cat = _dd(list)
        # Дедупликация между категориями
        seen_global: set = set()
        for e in all_events:
            key = (e.get("title", ""), e.get("place", ""))
            if key in seen_global:
                continue
            seen_global.add(key)
            by_cat[e.get("category")].append(e)
        for cat, evs in by_cat.items():
            emoji = CAT_POST_EMOJI.get(cat, "📌")
            cat_name = CAT_POST_NAMES.get(cat, cat.upper())
            # Дедупликация внутри категории по title
            seen_cat: set = set()
            unique_evs = []
            for e in evs:
                t = e.get("title", "")
                if t not in seen_cat:
                    seen_cat.add(t)
                    unique_evs.append(e)
            lines.append(f"\n{emoji} {cat_name}")
            for e in unique_evs[:4]:
                try:
                    date_str = datetime.strptime(e["event_date"], "%Y-%m-%d").strftime("%d.%m")
                except Exception:
                    date_str = e.get("event_date", "")[:5]  
                _SHORT_DAYS = ["пн","вт","ср","чт","пт","сб","вск"]
                day_n = _SHORT_DAYS[datetime.strptime(e["event_date"], "%Y-%m-%d").weekday()] if e.get("event_date") else ""
                time_str = f" {e.get('show_time')}" if e.get("show_time") else ""
                price = _fmt_price(e.get("price", "") or "")
                url = e.get("source_url") or ""
                title = e["title"] or ""
                title_part = f"<a href=\"{url}\">{title}</a>" if url else title
                lines.append(f"→ {title_part} ({date_str} {day_n}{time_str})" + (f" | {price}" if price else ""))
        lines.append(f"\n👉 Ищи все события: @Minskdvizh_bot")
        lines.append("#афишаминск #выходныеминск #движ")
    else:
        return

    text = "\n".join(lines)
    if len(text) > 4096:
        text = text[:4040] + "...\n\n👉 @Minskdvizh_bot"

    from telegram.error import RetryAfter
    try:
        await bot.send_message(
            chat_id=CHANNEL_ID,
            text=text,
            parse_mode="HTML",
            disable_web_page_preview=True,
        )
        logger.info(f"📢 Пост в канал ({post_type}) опубликован")
    except RetryAfter as e:
        logger.warning(f"post_to_channel RetryAfter {e.retry_after}с, повтор...")
        await asyncio.sleep(e.retry_after + 1)
        try:
            await bot.send_message(chat_id=CHANNEL_ID, text=text,
                                   parse_mode="HTML", disable_web_page_preview=True)
            logger.info(f"📢 Пост в канал ({post_type}) опубликован (повтор)")
        except Exception as e2:
            logger.error(f"Ошибка публикации в канал (повтор): {e2}")
    except Exception as e:
        logger.error(f"Ошибка публикации в канал: {e}")


def setup_scheduler(application):
    scheduler = AsyncIOScheduler()
    scheduler.add_job(
        run_parsers_job,
        trigger=CronTrigger(hour=3, minute=0, timezone="UTC"),  # UTC = 6:00 Минск
        kwargs={"bot": application.bot},
        id="daily_parsers",
        replace_existing=True,
    )
    scheduler.add_job(
        send_digest_job,
        trigger=CronTrigger(hour=5, minute=0, timezone="UTC"),  # UTC = 8:00 Минск
        kwargs={"bot": application.bot},
        id="daily_digest",
        replace_existing=True,
    )
    async def channel_today_job(bot=None):
        if bot:
            await post_to_channel(bot, "today")
    scheduler.add_job(
        channel_today_job,
        trigger=CronTrigger(hour=5, minute=5, timezone="UTC"),  # UTC = 8:05 Минск
        kwargs={"bot": application.bot},
        id="channel_today", replace_existing=True,
    )
    async def channel_weekend_job(bot=None):
        if bot:
            await post_to_channel(bot, "weekend")
    scheduler.add_job(
        channel_weekend_job,
        trigger=CronTrigger(day_of_week="fri", hour=8, minute=0, timezone="UTC"),  # UTC = 11:00 Минск
        kwargs={"bot": application.bot},
        id="channel_weekend", replace_existing=True,
    )
    scheduler.start()
    logger.info("⏰ Планировщик: парсеры 6:00, дайджест 8:00, канал 8:05, выходные пятница 11:00 (Минск)")


# ---------------------- Донат ----------------------


def _build_donate_keyboard():
    keyboard, row = [], []
    for amount in DONATION_SUGGESTIONS:
        row.append(InlineKeyboardButton(f"⭐ {amount} Stars", callback_data=f"donate_{amount}"))
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    return keyboard


DONATE_TEXT = (
    "🌟 **Поддержать проект**\n\n"
    "Если вам нравится бот и вы хотите поддержать его развитие, "
    "вы можете отправить донат в Telegram Stars.\n\n"
    "Выберите сумму ниже или отправьте команду\n"
    "`/donate <сумма>` (например, `/donate 150`)"
)


async def donate_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    log_user_action(user.id, user.username, user.first_name, "donate_menu")
    await update.message.reply_text(DONATE_TEXT, reply_markup=InlineKeyboardMarkup(_build_donate_keyboard()), parse_mode=ParseMode.MARKDOWN)


async def custom_donate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    try:
        if not context.args or len(context.args) != 1:
            await update.message.reply_text("❌ Используйте: `/donate <сумма>`\nНапример: `/donate 150`", parse_mode=ParseMode.MARKDOWN)
            return
        amount = int(context.args[0])
        if amount < 1:
            await update.message.reply_text("❌ Минимальная сумма — 1 Star")
            return
        if amount > 2500:
            await update.message.reply_text("❌ Максимальная сумма — 2500 Stars")
            return
        log_user_action(user.id, user.username, user.first_name, "donate_custom", str(amount))
        await send_star_invoice(update, context, amount)
    except ValueError:
        await update.message.reply_text("❌ Введите число. Пример: `/donate 150`", parse_mode=ParseMode.MARKDOWN)


async def send_star_invoice(update: Update, context: ContextTypes.DEFAULT_TYPE, amount: int):
    chat_id = update.callback_query.message.chat_id if update.callback_query else update.message.chat_id
    await context.bot.send_invoice(
        chat_id=chat_id,
        title="Поддержка бота",
        description=f"Благодарим за поддержку! Вы отправляете {amount} Telegram Stars.",
        payload=f"donation_{amount}_{datetime.now(MINSK_TZ).timestamp()}",
        provider_token="",
        currency=DONATION_CURRENCY,
        prices=[LabeledPrice(label=f"Stars {amount}", amount=amount)],
        need_name=False, need_phone_number=False, need_email=False,
        need_shipping_address=False, is_flexible=False,
    )


async def precheckout_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.pre_checkout_query
    if query.invoice_payload.startswith("donation_"):
        await query.answer(ok=True)
    else:
        await query.answer(ok=False, error_message="Что-то пошло не так")


async def successful_payment_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    amount = update.message.successful_payment.total_amount
    log_user_action(user.id, user.username, user.first_name, "donate_success", str(amount))
    await update.message.reply_text(
        f"✅ **Спасибо за поддержку!**\n\nВы отправили {amount} ⭐ Stars.\nВаша помощь очень ценится! 🙏",
        parse_mode=ParseMode.MARKDOWN,
    )
    await context.bot.send_message(
        chat_id=ADMIN_ID,
        text=(f"💰 **Получен донат!**\n\nОт: {user.first_name}\n"
              f"Username: @{user.username or 'нет'}\nID: `{user.id}`\nСумма: {amount} ⭐ Stars"),
        parse_mode=ParseMode.MARKDOWN,
    )


# ---------------------- Хендлеры сообщений ----------------------


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    log_user_action(user.id, user.username, user.first_name, "start")
    await update.message.reply_text(
        f"🎉 Привет, {user.first_name}!\n\n"
        "Я — 🌟**MinskDvizh**, твой персональный гид по событиям Минска.\n\n"
        "🔍 **Я еще учусь, но уже вот что умею:**\n"
        "• Искать по **названию**|**дате**|**описанию**\n"
        "• Показывать **сегодня**|**завтра**|**выходные** и **ближайшие** события\n"
        "• 🔔 Подписываться на **новые события** в выбранной категории\n"
        "• ✍️ **Добавлять твои события**, чтобы рассказать всем о мероприятии\n"
        "• ⚡ **Флеш-подписка** — узнавай о долгожданных событиях первым\n\n"
        "Используй кнопки меню 👇 или открой приложение по 🌟\n",
        reply_markup=get_reply_main_menu(), 
        parse_mode="Markdown",
    )

async def today_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    log_user_action(user.id, user.username, user.first_name, "cmd_today")
    today = datetime.now(MINSK_TZ)
    events = get_events_by_date_and_category(today)
    set_pagination(context, events, f"<b>События на {today.strftime('%d.%m.%Y')}:</b>",
                   share_query=f"date:{today.strftime('%Y-%m-%d')}")
    await show_page(update, context)


async def app_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    log_user_action(user.id, user.username, user.first_name, "open_webapp")
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("🌐 Открыть MinskDvizh", web_app=WebAppInfo(url=WEB_APP_URL))
    ]])
    await update.message.reply_text(
        "🌐 <b>MinskDvizh — веб-версия</b>\n\nОткрой полную афишу Минска прямо в Telegram:",
        reply_markup=keyboard, parse_mode="HTML",
    )


async def post_channel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ручная публикация в канал (только для админа)."""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Нет доступа.")
        return
    args = context.args
    post_type = args[0] if args else "today"
    if post_type not in ("today", "weekend"):
        await update.message.reply_text("Использование: /post_channel [today|weekend]")
        return
    await update.message.reply_text(f"📢 Публикую в канал ({post_type})...")
    await post_to_channel(context.bot, post_type)
    await update.message.reply_text("✅ Готово!")


async def inline_query_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Inline режим — для кнопки Поделиться.
    share_query формат: "cat:concert date:2026-03-10" или свободный текст.
    """
    from telegram import InlineQueryResultArticle, InputTextMessageContent
    import uuid as _uuid
    try:
        now = datetime.now(MINSK_TZ)
        today = now.strftime("%Y-%m-%d")
        query_text = (update.inline_query.query or "").strip()
        logger.info(f"[inline] from={update.inline_query.from_user.id} query='{query_text}'")

        # Пустой запрос — показываем подсказку вместо всех событий
        if not query_text:
            from telegram import InlineQueryResultArticle, InputTextMessageContent
            import uuid as _uuid
            hint = InlineQueryResultArticle(
                id=str(_uuid.uuid4()),
                title="🔍 Введите запрос для поиска событий",
                description="Например: концерт, выставка, 15.04.2026",
                input_message_content=InputTextMessageContent(
                    "🔍 Ищи события в @Minskdvizh_bot",
                    parse_mode="HTML"
                ),
            )
            await update.inline_query.answer([hint], cache_time=5)
            return

        # Парсим cat: и date: теги
        cat_filter = None
        date_filter = None
        date_from_filter = None
        date_to_filter = None
        text_parts = []
        for part in query_text.split():
            if part.startswith("cat:"):
                cat_filter = part[4:]
            elif part.startswith("date_from:"):
                date_from_filter = part[10:]
            elif part.startswith("date_to:"):
                date_to_filter = part[8:]
            elif part.startswith("date:"):
                date_filter = part[5:]
            else:
                text_parts.append(part)
        text_filter = " ".join(text_parts) if text_parts else None

        now_time = now.strftime("%H:%M")
        with get_db_connection() as conn:
            where = []
            params = []
            
            # 🔧 ОСОБЫЙ СЛУЧАЙ: категория "free" — показываем ВСЕ бесплатные события
            if cat_filter == "free":
                where.append("price = 'Бесплатно'")
                # Даты
                if date_filter:
                    where.append("event_date = ?")
                    params.append(date_filter)
                    if date_filter == today:
                        time_filter, time_params = _build_time_filter(date_filter, today, now_time)
                        where.append(time_filter)
                        params.extend(time_params)
                elif date_from_filter and date_to_filter:
                    where.append("event_date BETWEEN ? AND ?")
                    params += [date_from_filter, date_to_filter]
                    if date_from_filter == today:
                        where.append("(event_date > ? OR show_time = '' OR show_time IS NULL OR show_time > ?)")
                        params += [today, now_time]
                elif date_from_filter:
                    where.append("event_date >= ?")
                    params.append(date_from_filter)
                    if date_from_filter == today:
                        where.append("(event_date > ? OR show_time = '' OR show_time IS NULL OR show_time > ?)")
                        params += [today, now_time]
                else:
                    where.append("event_date >= ?")
                    params.append(today)
                    where.append("(event_date > ? OR show_time = '' OR show_time IS NULL OR show_time > ?)")
                    params += [today, now_time]
            else:
                # Обычная категория
                if date_filter:
                    where.append("event_date = ?")
                    params.append(date_filter)
                    # Для сегодня — исключаем прошедшие сеансы
                    if date_filter == today:
                        time_filter, time_params = _build_time_filter(date_filter, today, now_time)
                        where.append(time_filter)
                        params.extend(time_params)
                elif date_from_filter and date_to_filter:
                    where.append("event_date BETWEEN ? AND ?")
                    params += [date_from_filter, date_to_filter]
                    # Если начало диапазона — сегодня, фильтруем время
                    if date_from_filter == today:
                        where.append("(event_date > ? OR show_time = '' OR show_time IS NULL OR show_time > ?)")
                        params += [today, now_time]
                elif date_from_filter:
                    where.append("event_date >= ?")
                    params.append(date_from_filter)
                    if date_from_filter == today:
                        where.append("(event_date > ? OR show_time = '' OR show_time IS NULL OR show_time > ?)")
                        params += [today, now_time]
                else:
                    where.append("event_date >= ?")
                    params.append(today)
                    where.append("(event_date > ? OR show_time = '' OR show_time IS NULL OR show_time > ?)")
                    params += [today, now_time]
                if cat_filter:
                    where.append("category = ?")
                    params.append(cat_filter)
            if text_filter:
                where.append("(title LIKE ? OR place LIKE ?)")
                params += [f"%{text_filter}%", f"%{text_filter}%"]
            sql = f"""
                SELECT DISTINCT title, event_date, show_time, place, price, category, source_url
                FROM events WHERE {" AND ".join(where)}
                ORDER BY event_date, show_time LIMIT 10
            """
            rows = conn.execute(sql, params).fetchall()

        results = []
        for row in rows:
            cat_emoji = CATEGORY_EMOJI.get(row["category"] or "", "🎉")
            title = row["title"] or "Событие"
            try:
                date_str = datetime.strptime(row["event_date"], "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                date_str = row["event_date"] or ""
            time_str = f" ⏰ {row['show_time']}" if row["show_time"] else ""
            place_str = f"\n🏢 {row['place']}" if row["place"] else ""
            price_str = f"\n💰 {row['price']}" if row["price"] else ""
            url_str = f"\n🔗 {row['source_url']}" if row["source_url"] else ""
            msg = (f"{cat_emoji} <b>{title}</b>\n"
                   f"📅 {date_str}{time_str}{place_str}{price_str}{url_str}\n\n"
                   f"👉 @Minskdvizh_bot")
            results.append(InlineQueryResultArticle(
                id=str(_uuid.uuid4()),
                title=f"{cat_emoji} {title}",
                description=f"📅 {date_str}{time_str}" + (f" | {row['place']}" if row["place"] else ""),
                input_message_content=InputTextMessageContent(message_text=msg, parse_mode="HTML"),
            ))
        await update.inline_query.answer(results, cache_time=0)
    except Exception as e:
        logger.error(f"[inline] ошибка: {e}", exc_info=True)
        await update.inline_query.answer([], cache_time=5)


async def about(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    log_user_action(user.id, user.username, user.first_name, "about")
    counts = get_raw_events_count_by_category()
    total_events = sum(counts.values())
    cat_lines = [
        f"  {CATEGORY_NAMES[cat]} — {cnt}"
        for cat, cnt in counts.items()
        if cat in CATEGORY_NAMES and cat != "free" and cnt > 0
    ]
    # Бесплатные события — все с price='Бесплатно', независимо от категории
    today = datetime.now(MINSK_TZ).strftime("%Y-%m-%d")
    with get_db_connection() as conn:
        free_cnt = conn.execute(
            "SELECT COUNT(DISTINCT title || '|' || COALESCE(place,'') || '|' || event_date) "
            "FROM events WHERE event_date >= ? AND price = 'Бесплатно'",
            (today,)
        ).fetchone()[0]
    if free_cnt > 0:
        cat_lines.append(f"  🆓 Бесплатно — {free_cnt}")
    text = (
        "🌟 **MinskDvizh** — твой гид по событиям Минска!\n\n"
        "📅 **О проекте:**\n"
        "Собираем данные из разных источников и обновляем афишу каждое утро.\n\n"
        "🎯 **Актуальные события:**\n"
        + "\n".join(cat_lines) + "\n"
        f"\n📊 Всего событий: **{total_events}**\n\n"
        "🔍 **Как пользоваться:**\n"
        "• Отправь **название** события или **дату** (ДД.ММ)\n"
        "• Или используй кнопки меню\n\n"
        "💼 **Добавить мероприятие / сотрудничество:**\n"
        "📱 @i354444\n\n"
        "⭐ Если бот полезен — поддержи проект!\n\n"
        "#minskdvizh #афишаминск #минск"
    )
    await update.message.reply_text(
        text,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🌐 Открыть сайт", web_app=WebAppInfo(url=WEB_APP_URL))],
            [
                InlineKeyboardButton("➕ Добавить событие", callback_data="show_submit"),
                InlineKeyboardButton("📦 Загрузить список", callback_data="show_batch_template"),
            ],
            [
                InlineKeyboardButton("⭐ Поддержать", callback_data="show_donate"),
                InlineKeyboardButton("🔔 Мои подписки", callback_data="show_subs"),
            ],
        ]),
        parse_mode=ParseMode.MARKDOWN,
        disable_web_page_preview=True,
    )


async def search_by_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.message.text.strip()
    user = update.effective_user
    if len(query) < 3:
        await update.message.reply_text("🔍 Введите минимум 3 символа.")
        return
    log_user_action(user.id, user.username, user.first_name, "search_title", query)
    await update.message.chat.send_action(action="typing")
    events = search_events_by_title(query)
    if events:
        set_pagination(context, events, f"<b>Результаты: «{query}»</b>",
                       share_query=query)
        await show_page(update, context)
        # Предлагаем флеш-подписку
        await update.message.reply_text(
            f"⚡ <b>Флеш-подписка</b>\n\nХотите получить уведомление когда появятся новые события по запросу «{query}»?",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(
                    f"🔔 Подписаться на «{query[:30]}»",
                    callback_data=f"flash_sub_{query[:50]}"
                )
            ]]),
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(
            f"По запросу «{query}» ничего не найдено.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(
                    f"🔔 Уведомить когда появится",
                    callback_data=f"flash_sub_{query[:50]}"
                )
            ]]),
            parse_mode="Markdown"
        )


async def search_by_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    date_text = update.message.text.strip()
    user = update.effective_user
    log_user_action(user.id, user.username, user.first_name, "search_date", date_text)
    result, formatted_date, status = search_events_by_date_raw(date_text)
    if status == "неверный_формат":
        await update.message.reply_text(
            f"📅 Не удалось распознать дату «{date_text}».\nФормат: ДД.ММ или ДД.ММ.ГГГГ",
            parse_mode="Markdown",
        )
    elif status == "нет_событий":
        await update.message.reply_text(f"📅 Событий на {formatted_date} не найдено.")
    elif status == "найдены":
        set_pagination(context, result, f"📅 <b>События на {formatted_date}:</b>")
        await show_page(update, context)
    else:
        await update.message.reply_text("❌ Ошибка при поиске. Попробуйте позже.")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    user = update.effective_user

    if text == "⭐ Поддержать":
        log_user_action(user.id, user.username, user.first_name, "donate_menu_button")
        await donate_command(update, context)
        return
    if text == "ℹ️ О проекте":
        log_user_action(user.id, user.username, user.first_name, "about_button")
        await about(update, context)
        return
    if text == "📅 Сегодня":
        log_user_action(user.id, user.username, user.first_name, "menu_today")
        today = datetime.now(MINSK_TZ)
        events = get_events_by_date_and_category(today)
        set_pagination(context, events, f"<b>События на {today.strftime('%d.%m.%Y')}:</b>",
                       share_query=f"date:{today.strftime('%Y-%m-%d')}")
        await show_page(update, context)
        return
    if text == "📆 Завтра":
        log_user_action(user.id, user.username, user.first_name, "menu_tomorrow")
        tomorrow = datetime.now(MINSK_TZ) + timedelta(days=1)
        events = get_events_by_date_and_category(tomorrow)
        set_pagination(context, events, f"<b>События на {tomorrow.strftime('%d.%m.%Y')}:</b>",
                       share_query=f"date:{tomorrow.strftime('%Y-%m-%d')}")
        await show_page(update, context)
        return
    if text == "🎉 Выходные":
        log_user_action(user.id, user.username, user.first_name, "menu_weekend")
        events, saturday, sunday = get_weekend_events()
        set_pagination(context, events, f"<b>Выходные ({saturday.strftime('%d.%m')}–{sunday.strftime('%d.%m')}):</b>",
                       share_query=f"date_from:{saturday.strftime('%Y-%m-%d')} date_to:{sunday.strftime('%Y-%m-%d')}")
        await show_page(update, context)
        return
    if text == "⏰ Ближайшие":
        log_user_action(user.id, user.username, user.first_name, "menu_upcoming")
        events = get_upcoming_events(limit=100)
        if events:
            set_pagination(context, events, "⏰ <b>Ближайшие события:</b>", share_query="")
            await show_page(update, context)
        else:
            await update.message.reply_text("😕 Ближайших событий не найдено.")
        return
    if text == "🗓 Календарь":
        log_user_action(user.id, user.username, user.first_name, "menu_calendar")
        await show_calendar(update, context)
        return
    if text == "🎯 Категории":
        log_user_action(user.id, user.username, user.first_name, "menu_categories")
        counts = get_events_count_by_category()
        # Сначала известные категории в нужном порядке,
        # потом любые новые из БД — появляются автоматически
        ordered = {}
        for cat in CATEGORY_NAMES:
            if counts.get(cat, 0) > 0:
                ordered[cat] = CATEGORY_NAMES[cat]
        for cat in counts:
            if cat not in ordered and counts[cat] > 0:
                emoji = CATEGORY_EMOJI.get(cat, "📌")
                ordered[cat] = f"{emoji} {cat.replace('_', ' ').capitalize()}"
        keyboard = []
        row = []
        for cat, name in ordered.items():
            n = counts[cat]
            label = f"{name} ({n})"
            row.append(InlineKeyboardButton(label, callback_data=f"cat_{cat}"))
            if len(row) == 2:
                keyboard.append(row)
                row = []
        if row:
            keyboard.append(row)
        if not keyboard:
            await update.message.reply_text("😔 Пока нет доступных событий.")
            return
        await update.message.reply_text(
            "🎯 **Выберите категорию:**",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="Markdown",
        )
        return
    # Если пользователь заполняет форму — перехватываем ввод
    if await handle_submit_step(update, context):
        return

    if re.match(r"^\d{1,2}\.\d{1,2}(\.\d{2,4})?$", text):
        await search_by_date(update, context)
    else:
        await search_by_title(update, context)


# ---------------------- Хендлер кнопок ----------------------


async def handle_filter_buttons(query, context: ContextTypes.DEFAULT_TYPE, category: str):
    data = context.user_data.get("pagination")
    if not data:
        await query.answer("Устарело. Попробуйте снова.")
        return
    
    user = query.from_user
    log_user_action(user.id, user.username, user.first_name, "filter_category", category)
    
    # Удаляем клавиатуру с категориями из текущего сообщения
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception as e:
        logger.debug(f"Не удалось убрать клавиатуру: {e}")
    
    filtered = data["events"] if category == "all" else filter_events_by_category(data["events"], category)
    
    # Сбрасываем кеш группировки
    data.pop("_grouped", None)
    data.pop("_grouped_key", None)
    
    # Обновляем share_query
    old_sq = data.get("share_query") or ""
    sq_parts = [p for p in old_sq.split() if not p.startswith("cat:")]
    if category and category != "all":
        sq_parts.append(f"cat:{category}")
    new_sq = " ".join(sq_parts)
    
    set_pagination(context, filtered, data["title"], date_info=data["date_info"], share_query=new_sq)
    
    # Показываем отфильтрованные события в новом сообщении (как и было)
    await show_page(query, context)


async def handle_date_category_buttons(query, context: ContextTypes.DEFAULT_TYPE, date_type: str, category: str):
    user = query.from_user
    log_user_action(user.id, user.username, user.first_name, f"cat_{category}_{date_type}")
    display_name = CATEGORY_NAMES.get(category, category)
    if date_type == "today":
        today = datetime.now(MINSK_TZ)
        events = get_events_by_date_and_category(today, category)
        set_pagination(context, events, f"<b>{display_name} на {today.strftime('%d.%m.%Y')}:</b>",
                       share_query=f"cat:{category} date:{today.strftime('%Y-%m-%d')}")
        await show_page(query, context)
        await send_subscription_prompt(query, category, "today")
    elif date_type == "tomorrow":
        tomorrow = datetime.now(MINSK_TZ) + timedelta(days=1)
        events = get_events_by_date_and_category(tomorrow, category)
        set_pagination(context, events, f"<b>{display_name} на {tomorrow.strftime('%d.%m.%Y')}:</b>",
                       share_query=f"cat:{category} date:{tomorrow.strftime('%Y-%m-%d')}")
        await show_page(query, context)
        await send_subscription_prompt(query, category, "tomorrow")
    elif date_type == "upcoming":
        events = get_upcoming_events(limit=100, category=category)
        if events:
            set_pagination(context, events, f"<b>Ближайшие {display_name}:</b>",
                           share_query=f"cat:{category}")
            await show_page(query, context)
            await send_subscription_prompt(query, category, "upcoming")
        else:
            await query.edit_message_text(f"😕 Ближайших событий в категории {display_name} не найдено.", parse_mode="Markdown")
    elif date_type == "weekend":
        events, saturday, sunday = get_weekend_events(category=category)
        set_pagination(context, events, f"<b>{display_name} на выходные ({saturday.strftime('%d.%m')}–{sunday.strftime('%d.%m')}):</b>",
                       share_query=f"cat:{category} date_from:{saturday.strftime('%Y-%m-%d')} date_to:{sunday.strftime('%Y-%m-%d')}")
        await show_page(query, context)
        await send_subscription_prompt(query, category, "weekend")


async def handle_simple_buttons(query, context: ContextTypes.DEFAULT_TYPE, data: str):
    chat_id = query.message.chat_id
    user = query.from_user
    if data == "today":
        log_user_action(user.id, user.username, user.first_name, "btn_today")
        today = datetime.now(MINSK_TZ)
        events = get_events_by_date_and_category(today)
        set_pagination(context, events, f"<b>События на {today.strftime('%d.%m.%Y')}:</b>",
                       share_query=f"date:{today.strftime('%Y-%m-%d')}")
        await show_page(query, context)
    elif data == "tomorrow":
        log_user_action(user.id, user.username, user.first_name, "btn_tomorrow")
        tomorrow = datetime.now(MINSK_TZ) + timedelta(days=1)
        events = get_events_by_date_and_category(tomorrow)
        set_pagination(context, events, f"<b>События на {tomorrow.strftime('%d.%m.%Y')}:</b>",
                       share_query=f"date:{tomorrow.strftime('%Y-%m-%d')}")
        await show_page(query, context)
    elif data == "weekend":
        log_user_action(user.id, user.username, user.first_name, "btn_weekend")
        events, saturday, sunday = get_weekend_events()
        set_pagination(context, events, f"<b>Выходные ({saturday.strftime('%d.%m')}–{sunday.strftime('%d.%m')}):</b>",
                       share_query=f"date_from:{saturday.strftime('%Y-%m-%d')} date_to:{sunday.strftime('%Y-%m-%d')}")
        await show_page(query, context)
    elif data == "soon":
        log_user_action(user.id, user.username, user.first_name, "btn_upcoming")
        events = get_upcoming_events(limit=100)
        if events:
            set_pagination(context, events, "⏰ <b>Ближайшие события:</b>")
            await show_page(query, context)
        else:
            await query.edit_message_text("😕 Ближайших событий не найдено.", parse_mode="Markdown")
    elif data == "calendar":
        await show_calendar(query, context)
    elif data == "show_categories":
        await show_categories_menu(query, context)
    elif data == "back_to_main":
        await show_main_menu(chat_id, context, query.message.reply_text)
    elif data.startswith("cat_"):
        category = data.replace("cat_", "")
        log_user_action(user.id, user.username, user.first_name, "open_category", category)
        await show_date_options(query, category)


async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        query = update.callback_query
        data = query.data

        # ── Кнопки /admin панели ─────────────────────────────────
        if data.startswith("adm_") and query.from_user.id == ADMIN_ID:
            await query.answer()
            cmd = data[4:]
            if cmd == "stats":
                stats = get_stats_data(exclude_admin=False)
                await query.message.reply_text(_format_stats(stats, "📊 СТАТИСТИКА (все)"), parse_mode="HTML")
            elif cmd == "ustats":
                stats = get_stats_data(exclude_admin=True)
                await query.message.reply_text(_format_stats(stats, "📊 СТАТИСТИКА ПОЛЬЗОВАТЕЛЕЙ"), parse_mode="HTML")
            elif cmd == "update":
                await update_parsers(query, context)  # передаем query, а не update
            elif cmd == "download":
                try:
                    await query.message.reply_document(
                        document=open(DB_NAME, "rb"),
                        filename="events_final.db",
                        caption=f"🗄 База данных\n📅 {datetime.now(MINSK_TZ).strftime('%d.%m.%Y %H:%M')}",
                    )
                except Exception as e:
                    await query.message.reply_text(f"❌ Ошибка: {e}")
            elif cmd == "post_today":
                await query.message.reply_text("📢 Публикую подборку на сегодня...")
                await post_to_channel(context.bot, "today")
                await query.message.reply_text("✅ Готово!")
            elif cmd == "post_weekend":
                await query.message.reply_text("📢 Публикую подборку на выходные...")
                await post_to_channel(context.bot, "weekend")
                await query.message.reply_text("✅ Готово!")
            elif cmd == "dedup":
                with get_db_connection() as conn:
                    conn.row_factory = sqlite3.Row
                    groups = _find_duplicates(conn)
                total_groups = len(groups)
                total_delete = sum(len(g['delete_ids']) for g in groups)
                if total_groups == 0:
                    await query.message.reply_text("✅ Дубликатов не найдено — база чистая.")
                else:
                    lines = [f"🔍 <b>Найдено дубликатов:</b> {total_groups} групп → {total_delete} лишних записей\n"]
                    for g in groups[:10]:
                        lines.append(
                            f"• <b>{g['title'][:50]}</b>\n"
                            f"  📅 {g['event_date']} · 📍 {g['place'][:30] or '—'}\n"
                            f"  Копий: {g['cnt']} → оставим id={g['keep_id']}, удалим {len(g['delete_ids'])} шт."
                        )
                    if total_groups > 10:
                        lines.append(f"\n<i>...и ещё {total_groups - 10} групп</i>")
                    confirm_kb = InlineKeyboardMarkup([[
                        InlineKeyboardButton("✅ Удалить дубликаты", callback_data="adm_dedup_confirm"),
                        InlineKeyboardButton("❌ Отмена", callback_data="adm_dedup_cancel"),
                    ]])
                    await query.message.reply_text('\n'.join(lines), parse_mode="HTML", reply_markup=confirm_kb)
            elif cmd == "dedup_confirm":
                with get_db_connection() as conn:
                    conn.row_factory = sqlite3.Row
                    groups = _find_duplicates(conn)
                total_delete = sum(len(g['delete_ids']) for g in groups)
                all_delete_ids = [i for g in groups for i in g['delete_ids']]
                if all_delete_ids:
                    with get_db_connection() as conn:
                        placeholders = ','.join('?' * len(all_delete_ids))
                        conn.execute(f"DELETE FROM events WHERE id IN ({placeholders})", all_delete_ids)
                        conn.commit()
                await query.message.reply_text(
                    f"🧹 <b>Удалено {total_delete} дубликатов</b> из {len(groups)} групп. База очищена.",
                    parse_mode="HTML"
                )
            elif cmd == "dedup_cancel":
                await query.message.reply_text("Отменено.")
            elif cmd == "del_prompt":
                await query.message.reply_text(
                    "🗑 <b>Удаление событий</b>\n\n"
                    "Используйте команду:\n"
                    "<code>/delete_event 123</code> — удалить событие #123\n"
                    "<code>/delete_event 100 200</code> — удалить события с id 100 по 200\n\n"
                    "Перед удалением будет показан список и запрошено подтверждение.",
                    parse_mode="HTML"
                )
            elif cmd.startswith("del_confirm_"):
                try:
                    _, _, id_from_s, id_to_s = cmd.split("_", 3)
                    id_from, id_to = int(id_from_s), int(id_to_s)
                except ValueError:
                    await query.message.reply_text("❌ Ошибка разбора диапазона.")
                    return
                with get_db_connection() as conn:
                    deleted = conn.execute(
                        "DELETE FROM events WHERE id BETWEEN ? AND ?", (id_from, id_to)
                    ).rowcount
                    conn.commit()
                await query.message.reply_text(
                    f"🗑 <b>Удалено {deleted} событий</b> (id {id_from}–{id_to}).",
                    parse_mode="HTML"
                )
                try:
                    await query.edit_message_reply_markup(reply_markup=None)
                except Exception:
                    pass
            elif cmd == "del_cancel":
                await query.answer("Отменено.", show_alert=False)
                try:
                    await query.edit_message_reply_markup(reply_markup=None)
                except Exception:
                    pass
            elif cmd == "pending":
                await show_pending_list(query, context)
            elif cmd == "approve_all":
                with get_db_connection() as conn:
                    ids = [r["id"] for r in conn.execute(
                        "SELECT id FROM pending_events WHERE status IN ('pending','edited')"
                    ).fetchall()]
                cnt = 0
                for pid in ids:
                    ok, row_data = approve_pending_event(pid)
                    if ok:
                        cnt += 1
                        if row_data:
                            try:
                                await context.bot.send_message(
                                    chat_id=row_data["user_id"],
                                    text=f"✅ Ваше событие <b>{row_data['title']}</b> одобрено и добавлено в афишу! 🎉",
                                    parse_mode="HTML"
                                )
                            except Exception:
                                pass
                            # Промо-публикация при массовом одобрении
                            if row_data.get("is_promo"):
                                promo_text = format_promo_post(row_data)
                                if CHANNEL_ID:
                                    try:
                                        await context.bot.send_message(
                                            chat_id=CHANNEL_ID,
                                            text=promo_text,
                                            parse_mode="HTML",
                                            disable_web_page_preview=False,
                                        )
                                    except Exception as e:
                                        logger.error(f"Промо в канал (approve_all): {e}")
                                await send_promo_to_subscribers(context.bot, row_data)
                await query.message.reply_text(f"✅ Одобрено событий: <b>{cnt}</b>", parse_mode="HTML")
                await show_pending_list(query, context)
            elif cmd == "reject_all":
                with get_db_connection() as conn:
                    rows = conn.execute(
                        "SELECT id, user_id, title FROM pending_events WHERE status IN ('pending','edited')"
                    ).fetchall()
                for r in rows:
                    reject_pending_event(r["id"])
                    try:
                        await context.bot.send_message(
                            chat_id=r["user_id"],
                            text=f"❌ Ваше событие <b>{r['title']}</b> не прошло модерацию.",
                            parse_mode="HTML"
                        )
                    except Exception:
                        pass
                await query.message.reply_text(f"❌ Отклонено событий: <b>{len(rows)}</b>", parse_mode="HTML")
                await show_pending_list(query, context)
            return

        # Выбор категории в форме добавления события
        if data.startswith("sc_"):
            category = data[3:]
            cat_name = CATEGORY_NAMES.get(category, category)
            if context.user_data.get("mod_edit_field") == "category" and context.user_data.get("mod_edit_id"):
                # Модератор меняет категорию
                edit_data = context.user_data.get("mod_edit_data", {})
                edit_data["category"] = category
                context.user_data.pop("mod_edit_field", None)
                pending_id = context.user_data["mod_edit_id"]
                edit_data["_pending_id"] = pending_id
                await query.answer(f"Выбрано: {cat_name}")
                await query.message.reply_text(
                    f"✅ <b>Категория</b> обновлена: {cat_name}\n\nПродолжайте редактирование:",
                    reply_markup=build_fields_keyboard(edit_data, mode="mod_edit"),
                    parse_mode="HTML"
                )
            elif context.user_data.get("submit_field") == "category" and context.user_data.get("in_submit"):
                # Пользователь выбирает категорию
                context.user_data["submit"]["category"] = category
                context.user_data.pop("submit_field", None)
                submit_data = context.user_data["submit"]
                await query.answer(f"Выбрано: {cat_name}")
                await query.message.reply_text(
                    f"✅ <b>Категория</b> сохранена: {cat_name}\n\nВыберите следующее поле:",
                    reply_markup=build_fields_keyboard(submit_data, mode="submit"),
                    parse_mode="HTML"
                )
            return

        # Подтверждение/отмена отправки события
        if data == "submit_toggle_promo":
            submit_data = context.user_data.get("submit", {})
            submit_data["is_promo"] = not submit_data.get("is_promo", False)
            context.user_data["submit"] = submit_data
            await query.answer(
                "📣 Промо включено!" if submit_data["is_promo"] else "Промо отключено",
                show_alert=False
            )
            await query.edit_message_reply_markup(
                reply_markup=build_fields_keyboard(submit_data, mode="submit")
            )
            return

        if data == "submit_confirm":
            user = query.from_user
            data_form = context.user_data.get("submit", {})
            missing = [FIELD_LABELS[f][1] for f in ["title","event_date","category","place","details","show_time"]
                       if not data_form.get(f)]
            if missing:
                await query.answer(f"Не заполнены обязательные поля: {', '.join(missing)}", show_alert=True)
                return

            # ── Проверка дубликата ────────────────────────────────
            dup = check_duplicate_event(
                title=data_form.get("title", ""),
                event_date=data_form.get("event_date", ""),
                place=data_form.get("place", ""),
            )
            if dup:
                await query.answer()
                reason = _fmt_duplicate_reason(dup)
                for k in ["in_submit", "submit", "submit_field"]:
                    context.user_data.pop(k, None)
                await query.edit_message_text(
                    f"❌ <b>Событие не принято — дубликат</b>\n\n{reason}\n\n"
                    f"Если вы считаете, что это ошибка — свяжитесь с @i354444",
                    parse_mode="HTML"
                )
                log_user_action(user.id, user.username, user.first_name, "submit_duplicate",
                                data_form.get("title"))
                return
            # ─────────────────────────────────────────────────────

            pending_id = save_pending_event(user.id, user.username, user.first_name, data_form)
            for k in ["in_submit", "submit", "submit_field"]:
                context.user_data.pop(k, None)
            await query.answer()
            await query.edit_message_text(
                "✅ <b>Событие отправлено на модерацию!</b>\n\nМы рассмотрим его в ближайшее время.",
                parse_mode="HTML"
            )
            log_user_action(user.id, user.username, user.first_name, "submit_event_sent", data_form.get("title"))
            preview = format_pending_preview(data_form, user)
            if data_form.get("is_promo"):
                preview += "\n\n📣 <b>Пользователь запросил промо-публикацию в канале</b>"
            admin_keyboard = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ Одобрить", callback_data=f"mod_approve_{pending_id}"),
                    InlineKeyboardButton("❌ Отклонить", callback_data=f"mod_reject_{pending_id}"),
                ],
                [InlineKeyboardButton("✏️ Редактировать", callback_data=f"mod_edit_{pending_id}")],
            ])
            try:
                await context.bot.send_message(
                    chat_id=ADMIN_ID, text=preview,
                    reply_markup=admin_keyboard, parse_mode="HTML"
                )
            except Exception as e:
                logger.error(f"Не удалось отправить событие админу: {e}")
            return

        if data == "submit_cancel":
            for k in ["in_submit", "submit", "submit_field"]:
                context.user_data.pop(k, None)
            await query.answer()
            await query.edit_message_text("❌ Добавление отменено.")
            return

        # Модерация (только для админа)
        if data.startswith("mod_approve_"):
            if query.from_user.id != ADMIN_ID:
                await query.answer("⛔ Нет доступа", show_alert=True)
                return
            pending_id = int(data.split("_")[-1])
            # Сохраняем title до approve (статус изменится)
            with get_db_connection() as conn:
                _row = conn.execute("SELECT user_id, title FROM pending_events WHERE id=?", (pending_id,)).fetchone()
            ok, row_data = approve_pending_event(pending_id)
            if ok:
                await query.answer("✅ Одобрено!")
                if _row:
                    try:
                        await context.bot.send_message(
                            chat_id=_row["user_id"],
                            text=f"✅ Ваше событие <b>{_row['title']}</b> одобрено и добавлено в афишу! 🎉",
                            parse_mode="HTML"
                        )
                    except Exception:
                        pass
                # Промо-публикация в канал + рассылка подписчикам категории
                if row_data and row_data.get("is_promo"):
                    promo_text = format_promo_post(row_data)
                    if CHANNEL_ID:
                        try:
                            await context.bot.send_message(
                                chat_id=CHANNEL_ID,
                                text=promo_text,
                                parse_mode="HTML",
                                disable_web_page_preview=False,
                            )
                        except Exception as e:
                            logger.error(f"Промо в канал: {e}")
                    subs_sent = await send_promo_to_subscribers(context.bot, row_data)
                    await query.message.reply_text(
                        f"📣 Промо опубликовано в канале и разослано {subs_sent} подписчикам."
                    )
                import html as _html
                title_escaped = _html.escape(_row["title"] if _row else f"#{pending_id}")
                await query.message.reply_text(
                    f"✅ <b>Одобрено:</b> {title_escaped}",
                    parse_mode="HTML"
                )
                await show_pending_list(query, context)
            else:
                await query.answer("Ошибка при одобрении", show_alert=True)
            return

        if data.startswith("mod_reject_"):
            if query.from_user.id != ADMIN_ID:
                await query.answer("⛔ Нет доступа", show_alert=True)
                return
            pending_id = int(data.split("_")[-1])
            with get_db_connection() as conn:
                _row = conn.execute("SELECT user_id, title FROM pending_events WHERE id=?", (pending_id,)).fetchone()
            reject_pending_event(pending_id)
            await query.answer("❌ Отклонено")
            if _row:
                try:
                    await context.bot.send_message(
                        chat_id=_row["user_id"],
                        text=f"❌ Ваше событие <b>{_row['title']}</b> не прошло модерацию.",
                        parse_mode="HTML"
                    )
                except Exception:
                    pass
            import html as _html
            title_escaped = _html.escape(_row["title"] if _row else f"#{pending_id}")
            await query.message.reply_text(
                f"❌ <b>Отклонено:</b> {title_escaped}",
                parse_mode="HTML"
            )
            await show_pending_list(query, context)
            return

        if data.startswith("submit_field_"):
            field = data.replace("submit_field_", "")
            context.user_data["submit_field"] = field
            await query.answer()
            if field == "category":
                await query.message.reply_text(FIELD_PROMPTS["category"], reply_markup=CATEGORY_KEYBOARD, parse_mode="HTML")
            else:
                await query.message.reply_text(get_prompt(field), parse_mode="HTML")
            return

        if data.startswith("mod_edit_") and not data.startswith("mod_edit_cancel_") and not data.startswith("mod_edit_field_"):
            if query.from_user.id != ADMIN_ID:
                await query.answer("⛔ Нет доступа", show_alert=True)
                return
            pending_id = int(data.split("_")[-1])
            event = get_pending_event(pending_id)
            if not event:
                await query.answer("Событие не найдено", show_alert=True)
                return
            context.user_data["mod_edit_id"] = pending_id
            context.user_data["mod_edit_user_id"] = event["user_id"]
            # Собираем edit_data по всем полям FIELD_LABELS; details фолбек на description
            edit_data = {}
            for k in FIELD_LABELS:
                edit_data[k] = event.get(k) or ""
            if not edit_data.get("details") and event.get("description"):
                edit_data["details"] = event["description"]
            edit_data["_pending_id"] = pending_id
            context.user_data["mod_edit_data"] = edit_data
            context.user_data.pop("mod_edit_field", None)
            await query.answer()
            await query.message.reply_text(
                f"✏️ <b>Редактирование события #{pending_id}</b>\n\n"
                f"Выберите поле для изменения:",
                reply_markup=build_fields_keyboard(edit_data, mode="mod_edit"),
                parse_mode="HTML"
            )
            return

        if data.startswith("mod_edit_field_"):
            if query.from_user.id != ADMIN_ID:
                await query.answer("⛔ Нет доступа", show_alert=True)
                return
            field = data.replace("mod_edit_field_", "")
            context.user_data["mod_edit_field"] = field
            await query.answer()
            if field == "category":
                await query.message.reply_text(FIELD_PROMPTS["category"], reply_markup=CATEGORY_KEYBOARD, parse_mode="HTML")
            else:
                await query.message.reply_text(get_prompt(field, "/skip — очистить поле | /cancel — отмена"), parse_mode="HTML")
            return

        if data.startswith("mod_send_edit_"):
            if query.from_user.id != ADMIN_ID:
                await query.answer("⛔ Нет доступа", show_alert=True)
                return
            pending_id = int(data.split("_")[-1])
            edit_data = context.user_data.get("mod_edit_data", {})
            user_id = context.user_data.get("mod_edit_user_id")
            if not edit_data or not user_id:
                await query.answer("Нет данных", show_alert=True)
                return
            clean = {k: v for k, v in edit_data.items() if k != "_pending_id"}
            update_pending_event(pending_id, clean)
            await query.answer("📤 Отправлено")
            preview = format_pending_preview(clean)
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=f"✏️ <b>Модератор внёс изменения в ваше событие</b>\n\n{preview}\n\n<i>Вы согласны?</i>",
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("✅ Принять", callback_data=f"user_accept_edit_{pending_id}"),
                        InlineKeyboardButton("❌ Отклонить", callback_data=f"user_reject_edit_{pending_id}"),
                    ]]),
                    parse_mode="HTML"
                )
                await query.message.reply_text(f"✅ Событие #{pending_id} отправлено пользователю на согласование.")
            except Exception as e:
                await query.message.reply_text(f"⚠️ Не удалось отправить: {e}")
            for k in ["mod_edit_id", "mod_edit_user_id", "mod_edit_data", "mod_edit_field"]:
                context.user_data.pop(k, None)
            return

        if data.startswith("mod_edit_cancel_"):
            for k in ["mod_edit_id", "mod_edit_user_id", "mod_edit_data", "mod_edit_field"]:
                context.user_data.pop(k, None)
            await query.answer("Редактирование отменено")
            await query.edit_message_reply_markup(reply_markup=None)
            return

        if data.startswith("user_accept_edit_"):
            pending_id = int(data.split("_")[-1])
            ok, row_data = approve_pending_event(pending_id)
            await query.answer()
            if ok:
                await query.edit_message_text(
                    query.message.text + "\n\n✅ <b>Вы приняли изменения. Событие добавлено в афишу!</b>",
                    parse_mode="HTML"
                )
                try:
                    await context.bot.send_message(chat_id=ADMIN_ID,
                        text=f"✅ Пользователь принял правки события #{pending_id} — добавлено в афишу.")
                except Exception:
                    pass
                # Промо при принятии пользователем
                if row_data and row_data.get("is_promo"):
                    promo_text = format_promo_post(row_data)
                    if CHANNEL_ID:
                        try:
                            await context.bot.send_message(
                                chat_id=CHANNEL_ID,
                                text=promo_text,
                                parse_mode="HTML",
                                disable_web_page_preview=False,
                            )
                        except Exception as e:
                            logger.error(f"Промо (user_accept): {e}")
                    await send_promo_to_subscribers(context.bot, row_data)
            else:
                await query.edit_message_text("⚠️ Ошибка при добавлении.", parse_mode="HTML")
            return

        if data.startswith("user_reject_edit_"):
            pending_id = int(data.split("_")[-1])
            reject_pending_event(pending_id)
            await query.answer()
            await query.edit_message_text(
                query.message.text + "\n\n❌ <b>Вы отклонили изменения. Событие не добавлено.</b>",
                parse_mode="HTML"
            )
            try:
                await context.bot.send_message(chat_id=ADMIN_ID,
                    text=f"❌ Пользователь отклонил правки события #{pending_id}.")
            except Exception:
                pass
            return

        if data == "submit_preview":
            data_form = context.user_data.get("submit", {})
            await query.answer()
            await query.message.reply_text(
                "👁 <b>Предпросмотр вашего события:</b>\n\n" + format_pending_preview(data_form),
                parse_mode="HTML"
            )
            return

        if data.startswith("mod_preview_"):
            if query.from_user.id != ADMIN_ID:
                await query.answer("⛔ Нет доступа", show_alert=True)
                return
            edit_data = context.user_data.get("mod_edit_data", {})
            clean = {k: v for k, v in edit_data.items() if k != "_pending_id"}
            await query.answer()
            await query.message.reply_text(
                "👁 <b>Предпросмотр после редактирования:</b>\n\n" + format_pending_preview(clean),
                parse_mode="HTML"
            )
            return

        if data == "show_submit":
            user = query.from_user
            log_user_action(user.id, user.username, user.first_name, "submit_event_start")
            await start_submit(query, context)
            return

        if data == "show_subs":
            await query.answer()
            await show_subscriptions_query(query, context)
            return
        if data == "show_batch_template":
            await query.answer()
            user = query.from_user
            log_user_action(user.id, user.username, user.first_name, "batch_template")
            await send_batch_template(query.message)
            return

        if data == "show_donate":
            await query.answer()
            user = query.from_user
            log_user_action(user.id, user.username, user.first_name, "donate_menu")
            await query.message.reply_text(DONATE_TEXT, reply_markup=InlineKeyboardMarkup(_build_donate_keyboard()), parse_mode=ParseMode.MARKDOWN)
            return
        if data.startswith("donate_"):
            await query.answer()
            await send_star_invoice(update, context, int(data.replace("donate_", "")))
            return
        if data.startswith("filter_"):
            await handle_filter_buttons(query, context, data.replace("filter_", ""))
            return
        if data.startswith("date_"):
            parts = data.split("_")
            await handle_date_category_buttons(query, context, parts[1], parts[2])
            return
        if data == "page_noop":
            await query.answer()
            return
        if data == "page_prev":
            if "pagination" in context.user_data:
                context.user_data["pagination"]["page"] = max(0, context.user_data["pagination"]["page"] - 1)
            await show_page(query, context)
            return
        if data == "page_next":
            if "pagination" in context.user_data:
                context.user_data["pagination"]["page"] += 1
            await show_page(query, context)
            return
        if data.startswith("sub_"):
            _, category, date_type = data.split("_", 2)
            user = query.from_user
            add_subscription(user.id, category, date_type)
            log_user_action(user.id, user.username, user.first_name, "subscribe", f"{category}_{date_type}")
            try:
                await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔕 Отписаться", callback_data=f"unsub_{category}_{date_type}")
                ]]))
            except Exception: pass
            await query.answer("Подписка оформлена 🔔", show_alert=False)
            return
        if data.startswith("flash_sub_"):
            user = query.from_user
            flash_query = data[len("flash_sub_"):]
            added = add_flash_subscription(user.id, flash_query)
            log_user_action(user.id, user.username, user.first_name, "flash_subscribe", flash_query)
            if added:
                await query.answer("⚡ Флеш-подписка оформлена!", show_alert=False)
                try:
                    await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("✅ Подписан", callback_data="page_noop")
                    ]]))
                except Exception:
                    pass
            else:
                await query.answer("Вы уже подписаны на этот запрос", show_alert=True)
            return
        if data.startswith("flash_found_"):
            user = query.from_user
            flash_id = int(data[len("flash_found_"):])
            remove_flash_subscription(flash_id, user.id)
            log_user_action(user.id, user.username, user.first_name, "flash_found", str(flash_id))
            await query.answer("🎉 Отлично! Подписка удалена.", show_alert=False)
            try:
                await query.edit_message_reply_markup(reply_markup=None)
            except Exception:
                pass
            return
        if data.startswith("flash_continue_"):
            flash_id = int(data[len("flash_continue_"):])
            await query.answer("🔄 Продолжаем поиск!", show_alert=False)
            try:
                await query.edit_message_reply_markup(reply_markup=None)
            except Exception:
                pass
            return
        if data.startswith("flash_unsub_"):
            user = query.from_user
            flash_id = int(data[len("flash_unsub_"):])
            remove_flash_subscription(flash_id, user.id)
            log_user_action(user.id, user.username, user.first_name, "flash_unsubscribe", str(flash_id))
            await query.answer("⚡ Флеш-подписка отменена", show_alert=False)
            await show_subscriptions_query(query, context)
            return
        if data.startswith("unsub_"):
            _, category, date_type = data.split("_", 2)
            user = query.from_user
            remove_subscription(user.id, category, date_type)
            log_user_action(user.id, user.username, user.first_name, "unsubscribe", f"{category}_{date_type}")
            await query.answer("Подписка отменена 🔕", show_alert=False)
            # Если открыт экран /subs — обновляем список
            msg_text = (query.message.text or "")
            if "Мои подписки" in msg_text:
                await show_subscriptions_query(query, context)
            else:
                try:
                    await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("🔔 Подписаться", callback_data=f"sub_{category}_{date_type}")
                    ]]))
                except Exception:
                    pass
            return
        if data.startswith("cal_"):
            parts = data.split("_")
            action = parts[1]
            year, month = int(parts[2]), int(parts[3])
            if action == "prev":
                month -= 1
                if month < 1: month = 12; year -= 1
            elif action == "next":
                month += 1
                if month > 12: month = 1; year += 1
            elif action == "day":
                day = int(parts[4])
                date_obj = datetime(year, month, day, tzinfo=MINSK_TZ)
                user = query.from_user
                log_user_action(user.id, user.username, user.first_name, "calendar_day", f"{day:02d}.{month:02d}.{year}")
                events = get_events_by_date_and_category(date_obj)
                if events:
                    set_pagination(context, events, f"📅 События на {day:02d}.{month:02d}.{year}:")
                    await show_page(query, context)
                else:
                    await query.answer(f"На {day:02d}.{month:02d}.{year} событий нет", show_alert=True)
                return
            await show_calendar(query, context, year, month)
            return
        await handle_simple_buttons(query, context, data)

    except Exception as e:
        logger.error(f"button_handler error: {e}", exc_info=True)
        try:
            await query.answer("Произошла ошибка, попробуйте ещё раз", show_alert=True)
        except Exception:
            pass


# ---------------------- Пакетная загрузка событий из файла ----------------------



async def send_batch_template(message):
    """Отправляет xlsx-шаблон для пакетной загрузки событий."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "События"

        header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(BATCH_TEMPLATE_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        example_fill = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
        for col, val in enumerate(BATCH_TEMPLATE_EXAMPLE, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.fill = example_fill

        widths = [35, 35, 15, 12, 8, 25, 25, 15, 40, 30, 10]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        ws2 = wb.create_sheet("Подсказки")
        ws2["A1"] = "Поле"
        ws2["B1"] = "Обязательное"
        ws2["C1"] = "Формат / пример"
        hints = [
            ("title",       "ДА",  "Название события"),
            ("details",     "ДА",  "Краткое описание — для кого и что"),
            ("category",    "ДА",  "concert / theater / cinema / exhibition / kids / sport / party / free / excursion / market / masterclass / boardgames / broadcast / education / quiz"),
            ("event_date",  "ДА",  "ДД.ММ.ГГГГ  или  ДД.ММ.ГГГГ-ДД.ММ.ГГГГ (период)"),
            ("show_time",   "ДА",  "ЧЧ:ММ  или  ЧЧ:ММ-ЧЧ:ММ"),
            ("place",       "ДА",  "Название площадки"),
            ("address",     "нет", "ул. Пример, 1"),
            ("price",       "нет", "от 20 руб  /  Бесплатно"),
            ("description", "нет", "Подробное описание, программа"),
            ("source_url",  "нет", "https://..."),
            ("is_promo",    "нет", "1 = анонсировать в канал и подписчикам, 0 = нет (по умолчанию)"),
        ]
        for row, (f, req, hint) in enumerate(hints, 2):
            ws2.cell(row=row, column=1, value=f)
            ws2.cell(row=row, column=2, value=req)
            ws2.cell(row=row, column=3, value=hint)
        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 70

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        caption = (
            "\U0001f4cb <b>Шаблон для пакетной загрузки событий</b>\n\n"
            "1. Заполните строки (строка 2 \u2014 пример, её можно удалить)\n"
            "2. Сохраните файл и отправьте его в этот чат\n"
            "3. Мы проверим и отправим на модерацию\n\n"
            "\U0001f4cc Лист <b>Подсказки</b> \u2014 описание каждого поля"
        )
        await message.reply_document(
            document=buf,
            filename="minskdvizh_events_template.xlsx",
            caption=caption,
            parse_mode="HTML",
        )
    except ImportError:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(BATCH_TEMPLATE_HEADERS)
        writer.writerow(BATCH_TEMPLATE_EXAMPLE)
        csv_bytes = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
        await message.reply_document(
            document=csv_bytes,
            filename="minskdvizh_events_template.csv",
            caption="\U0001f4cb <b>Шаблон для пакетной загрузки (CSV)</b>\n\nЗаполните и отправьте обратно.",
            parse_mode="HTML",
        )


def _parse_batch_date(raw: str) -> tuple[str, str]:
    """Парсит дату или период. Возвращает (event_date_value, error_str).
    Поддерживает форматы:
      - YYYY-MM-DD (ISO, в т.ч. datetime из Excel: "2026-03-29 00:00:00")
      - ДД.ММ.ГГГГ
      - ДД.ММ.ГГГГ-ДД.ММ.ГГГГ (период)
    """
    import re as _re
    from datetime import date as _date, datetime as _dt
    raw = (raw or "").strip()
    if not raw:
        return "", "пустая дата"

    # Excel/openpyxl часто возвращает datetime как строку "2026-03-29 00:00:00"
    # или "2026-03-29 19:00:00" — берём только дату
    excel_dt = _re.match(r"^(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}(:\d{2})?$", raw)
    if excel_dt:
        raw = excel_dt.group(1)

    if _re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
        try:
            d = _date.fromisoformat(raw)
            if d < _date.today():
                return "", f"дата {raw} в прошлом"
            return raw, ""
        except Exception:
            return "", f"невалидная дата {raw}"

    pm = _re.match(r"^(\d{1,2}\.\d{1,2}\.\d{4})-(\d{1,2}\.\d{1,2}\.\d{4})$", raw)
    if pm:
        def _pd(s):
            d, mo, y = s.split(".")
            return _date(int(y), int(mo), int(d))
        try:
            d1, d2 = _pd(pm.group(1)), _pd(pm.group(2))
        except Exception:
            return "", f"невалидный период {raw}"
        if d2 < d1:
            return "", f"конец периода раньше начала: {raw}"
        if d1 < _date.today():
            return "", f"дата начала {raw} в прошлом"
        if (d2 - d1).days > 90:
            return "", f"период > 90 дней: {raw}"
        return f"{d1.isoformat()}|{d2.isoformat()}", ""

    m = _re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", raw)
    if m:
        day, month, year = m.groups()
        try:
            d = _date(int(year), int(month), int(day))
        except Exception:
            return "", f"невалидная дата {raw}"
        if d < _date.today():
            return "", f"дата {raw} в прошлом"
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}", ""

    return "", f"не распознан формат даты: {raw}"


def _parse_batch_time(raw: str) -> tuple[str, str]:
    """Парсит время ЧЧ:ММ или ЧЧ:ММ-ЧЧ:ММ. Возвращает (show_time, end_time).
    Также обрабатывает Excel-формат "19:00:00" → "19:00".
    """
    import re as _re
    raw = (raw or "").strip()
    if not raw:
        return "", ""
    # Excel может отдать "19:00:00" — обрезаем секунды
    raw = _re.sub(r"^(\d{1,2}:\d{2}):\d{2}$", r"\1", raw)
    rng = _re.match(r"^(\d{1,2}:\d{2})-(\d{1,2}:\d{2})$", raw)
    if rng:
        return rng.group(1), rng.group(2)
    if _re.match(r"^\d{1,2}:\d{2}$", raw):
        return raw, ""
    return raw, ""


def _read_batch_file(data: bytes, filename: str) -> tuple[list[dict], str]:
    """Читает xlsx/xls/csv и возвращает (rows, error)."""
    fname = filename.lower()
    try:
        if fname.endswith(".csv"):
            text = data.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            return [dict(r) for r in reader], ""
        elif fname.endswith(".xlsx") or fname.endswith(".xls"):
            import openpyxl
            buf = io.BytesIO(data)
            wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return [], "Файл пустой"
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            result = []
            for row in rows[1:]:
                if all((v is None or str(v).strip() == "") for v in row):
                    continue
                result.append({headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                                for i in range(min(len(headers), len(row)))})
            return result, ""
        else:
            return [], f"Неподдерживаемый формат: {filename}"
    except ImportError:
        return [], "openpyxl не установлен — поддерживается только CSV"
    except Exception as e:
        return [], f"Ошибка чтения файла: {e}"


async def handle_batch_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Хендлер документа — принимает xlsx/csv с событиями, парсит, отправляет на модерацию."""
    import html as _html
    user = update.effective_user
    doc = update.message.document
    fname = doc.file_name or ""

    if not (fname.lower().endswith(".xlsx") or fname.lower().endswith(".xls") or fname.lower().endswith(".csv")):
        return

    log_user_action(user.id, user.username, user.first_name, "batch_upload", fname)

    if doc.file_size and doc.file_size > 5 * 1024 * 1024:
        await update.message.reply_text("\u274c Файл слишком большой. Максимум 5 МБ.")
        return

    await update.message.chat.send_action(action="upload_document")
    status_msg = await update.message.reply_text("\u23f3 Читаю файл...")

    try:
        file = await doc.get_file()
        buf = io.BytesIO()
        await file.download_to_memory(buf)
        data = buf.getvalue()
    except Exception as e:
        await status_msg.edit_text(f"\u274c Не удалось скачать файл: {e}")
        return

    rows, err = _read_batch_file(data, fname)
    if err:
        await status_msg.edit_text(f"\u274c {err}")
        return
    if not rows:
        await status_msg.edit_text("\u274c Файл пустой или не содержит данных.")
        return
    if len(rows) > 100:
        await status_msg.edit_text("\u274c Максимум 100 событий за раз.")
        return

    def _norm_key(k):
        return (k or "").strip().lower().replace(" ", "_")

    ok_count = 0
    skip_count = 0
    errors = []
    pending_ids = []
    now_str = datetime.now(MINSK_TZ).strftime("%Y-%m-%d %H:%M:%S")

    for row_num, raw_row in enumerate(rows, start=2):
        row = {_norm_key(k): v for k, v in raw_row.items()}

        title = row.get("title", "").strip()
        if not title or len(title) < 3:
            errors.append(f"Строка {row_num}: пустое или короткое название")
            skip_count += 1
            continue

        date_raw = row.get("event_date", "") or row.get("date", "")
        event_date, date_err = _parse_batch_date(date_raw)
        if date_err:
            errors.append(f"Строка {row_num} «{title[:25]}»: {date_err}")
            skip_count += 1
            continue

        time_raw = row.get("show_time", "") or row.get("time", "")
        show_time, end_time = _parse_batch_time(time_raw)

        place = row.get("place", "").strip()
        if not place:
            errors.append(f"Строка {row_num} «{title[:25]}»: не указано место")
            skip_count += 1
            continue

        cat_raw = (row.get("category", "") or "").strip().lower()
        category = BATCH_CATEGORY_MAP.get(cat_raw, "other")

        details = (row.get("details", "") or "")[:300]
        description = (row.get("description", "") or "")[:1000]
        address = (row.get("address", "") or "").strip()
        price = (row.get("price", "") or "").strip()
        source_url = (row.get("source_url", "") or row.get("url", "") or "").strip()
        # is_promo: "1", "да", "yes", "true" → 1, всё остальное → 0
        promo_raw = (row.get("is_promo", "") or "").strip().lower()
        is_promo = 1 if promo_raw in ("1", "да", "yes", "true", "+") else 0

        dup = check_duplicate_event(title, event_date, place)
        if dup:
            errors.append(f"Строка {row_num} «{title[:25]}»: дубликат (уже есть в афише/очереди)")
            skip_count += 1
            continue

        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO pending_events
                        (user_id, username, first_name, title, event_date, show_time, end_time,
                         place, address, category, details, description, price, source_url,
                         is_promo, status, created_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'pending',?)
                """, (
                    user.id, user.username, user.first_name,
                    title, event_date, show_time, end_time,
                    place, address, category,
                    details, description, price, source_url,
                    is_promo, now_str,
                ))
                conn.commit()
                pending_ids.append(cursor.lastrowid)
                ok_count += 1
        except Exception as e:
            errors.append(f"Строка {row_num} «{title[:25]}»: ошибка БД — {e}")
            skip_count += 1

    if ok_count > 0:
        reply = (
            f"\u2705 <b>Загрузка завершена!</b>\n\n"
            f"\U0001f4e5 Принято на модерацию: <b>{ok_count}</b>\n"
        )
        if skip_count:
            reply += f"\u26a0\ufe0f Пропущено: <b>{skip_count}</b>\n"
        if errors:
            err_lines = "\n".join(f"• {_html.escape(e)}" for e in errors[:10])
            if len(errors) > 10:
                err_lines += f"\n• ...и ещё {len(errors)-10} ошибок"
            reply += f"\n<b>Причины пропуска:</b>\n{err_lines}"
        reply += "\n\n\U0001f50d События проверит модератор и добавит в афишу."
    else:
        reply = "\u274c <b>Ни одно событие не принято.</b>\n\n"
        if errors:
            err_lines = "\n".join(f"• {_html.escape(e)}" for e in errors[:15])
            reply += f"<b>Причины:</b>\n{err_lines}"

    await status_msg.edit_text(reply, parse_mode="HTML")

    if ok_count > 0 and pending_ids:
        try:
            admin_text = (
                f"\U0001f4e6 <b>Пакетная загрузка от @{_html.escape(user.username or '')} "
                f"({_html.escape(user.first_name or '')}, ID {user.id})</b>\n\n"
                f"\U0001f4e5 Добавлено в очередь: <b>{ok_count}</b> событий\n"
                f"\u26a0\ufe0f Пропущено: <b>{skip_count}</b>\n"
                f"\U0001f4c1 Файл: {_html.escape(fname)}\n\n"
                f"Откройте /pending для просмотра."
            )
            await context.bot.send_message(
                chat_id=ADMIN_ID, text=admin_text, parse_mode="HTML"
            )
        except Exception as e:
            logger.warning(f"Не удалось уведомить админа о batch upload: {e}")


async def batch_template_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /template — отправляет шаблон xlsx для пакетной загрузки."""
    user = update.effective_user
    log_user_action(user.id, user.username, user.first_name, "batch_template")
    await send_batch_template(update.message)



# ---------------------- main ----------------------


def build_application() -> Application:
    """Создаёт и настраивает Application — используется и из main() и из start.py."""
    if not TOKEN:
        raise RuntimeError("TELEGRAM_BOT_TOKEN не задан в окружении")

    init_db()

    application = Application.builder().token(TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("today", today_command))
    application.add_handler(CommandHandler("subs", show_subscriptions))
    application.add_handler(CommandHandler("admin", admin_menu))
    application.add_handler(CommandHandler("pending", pending_command))
    application.add_handler(CommandHandler("stats", show_stats))
    application.add_handler(CommandHandler("download_db", download_db))
    application.add_handler(CommandHandler("ustats", show_ustats))
    application.add_handler(CommandHandler("update", update_parsers))
    application.add_handler(CommandHandler("donate", custom_donate))
    application.add_handler(CommandHandler("support", donate_command))
    application.add_handler(CommandHandler("app", app_command))
    application.add_handler(CommandHandler("post_channel", post_channel_command))
    application.add_handler(InlineQueryHandler(inline_query_handler))
    application.add_handler(CommandHandler("about", about))
    application.add_handler(CallbackQueryHandler(button_handler))
    application.add_handler(PreCheckoutQueryHandler(precheckout_callback))
    application.add_handler(MessageHandler(filters.SUCCESSFUL_PAYMENT, successful_payment_callback))
    application.add_handler(CommandHandler("dedup", dedup_command))
    application.add_handler(CommandHandler("delete_event", delete_event_command))
    application.add_handler(CommandHandler("template", batch_template_command))
    application.add_handler(MessageHandler(filters.Document.FileExtension("xlsx"), handle_batch_file))
    application.add_handler(MessageHandler(filters.Document.FileExtension("xls"), handle_batch_file))
    application.add_handler(MessageHandler(filters.Document.FileExtension("csv"), handle_batch_file))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    return application


def main():
    """Запуск через polling — для локальной разработки."""
    application = build_application()
    setup_scheduler(application)
    logger.info("🚀 Бот запущен (polling)")
    application.run_polling(allowed_updates=["message", "callback_query", "inline_query", "chosen_inline_result", "pre_checkout_query"])


if __name__ == "__main__":
    main()